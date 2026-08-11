#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
单一模型评估脚本（仅支持目录模式）

用法：
  python eval_single.py --input-dir <数据目录> --score-col score \
    --name "模型名称" -o output/models/

输入目录下收集所有 *.csv 与 *.parquet（按文件名排序），每个文件单独评估一份，
再把所有文件的行纵向拼接合并评估一份 "all"（全量样本整体一份）。

version 规则（无 --version 参数）：
  - 每个分文件：version = 文件名 stem（例 train.parquet → "train"）
  - 合并版：version = "all"

输出：
  <名称>_<version>_eval.json  +  <名称>_<version>_eval.md  +  <名称>_<version>_eval.xlsx
"""

import argparse, json, sys
from pathlib import Path
import pandas as pd
import numpy as np
from pandas.api import types as pd_types
from sklearn.metrics import roc_auc_score
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from datetime import datetime


# 支持的输入文件扩展名（小写，含点）
_SUPPORTED_EXTS = {".csv", ".parquet"}


def _read_table(path: Path) -> pd.DataFrame:
    """按扩展名读 CSV / parquet，其他扩展名报错。"""
    ext = path.suffix.lower()
    if ext == ".csv":
        return pd.read_csv(path)
    if ext == ".parquet":
        return pd.read_parquet(path)
    raise ValueError(
        f"不支持的输入格式: {path.suffix} (仅支持 {sorted(_SUPPORTED_EXTS)})。路径: {path}"
    )


def _list_table_files(directory: Path) -> list:
    """收集目录下所有 CSV / parquet 文件（按文件名排序）。"""
    files = [p for p in directory.iterdir()
             if p.is_file() and p.suffix.lower() in _SUPPORTED_EXTS]
    return sorted(files, key=lambda p: p.name)



# ============================================================
# 样式常量
# ============================================================
HF = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HFILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
CF = Font(name="微软雅黑", size=10)
BD = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
TITLE_FONT = Font(name="微软雅黑", bold=True, size=13, color="2F5496")
SUB_FONT = Font(name="微软雅黑", bold=True, size=11, color="2F5496")


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HF; cell.fill = HFILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BD


def style_data_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = CF; cell.border = BD


# ============================================================
# 指标计算
# ============================================================
def compute_auc(sub, score_col):
    if sub['label'].nunique() < 2:
        return None
    return float(roc_auc_score(sub['label'], sub[score_col]))


def compute_ks(sub, score_col):
    s = sub.sort_values(score_col, ascending=False)
    cp = (s['label'] == 1).cumsum() / max((s['label'] == 1).sum(), 1)
    cn = (s['label'] == 0).cumsum() / max((s['label'] == 0).sum(), 1)
    return float(abs(cp - cn).max())


def compute_classification_metrics(sub, score_col):
    """准确率、精确率、召回率、F1（以0.5为阈值）"""
    y_true = sub['label'].values
    y_pred = (sub[score_col].values >= 0.5).astype(int)
    tp = ((y_pred == 1) & (y_true == 1)).sum()
    tn = ((y_pred == 0) & (y_true == 0)).sum()
    fp = ((y_pred == 1) & (y_true == 0)).sum()
    fn = ((y_pred == 0) & (y_true == 1)).sum()
    total = len(y_true)
    accuracy = (tp + tn) / total if total > 0 else None
    precision = tp / (tp + fp) if (tp + fp) > 0 else None
    recall = tp / (tp + fn) if (tp + fn) > 0 else None
    f1 = 2 * precision * recall / (precision + recall) if (precision and recall and (precision + recall) > 0) else None
    return {
        'accuracy': round(float(accuracy), 6) if accuracy is not None else None,
        'precision': round(float(precision), 6) if precision is not None else None,
        'recall': round(float(recall), 6) if recall is not None else None,
        'f1': round(float(f1), 6) if f1 is not None else None,
    }


def decile_buckets(sub, score_col, biz_cols, n_bins=10):
    """N等频降序分桶（默认10）。"""
    s = sub.sort_values(score_col, ascending=False).copy()
    s['decile'] = pd.cut(range(len(s)), bins=n_bins, labels=False) + 1
    s['decile'] = n_bins + 1 - s['decile']  # n_bins=最高分
    overall_lr = float(s['label'].mean())
    total_pos = int(s['label'].sum())
    cum_pos = 0
    result = []
    for d in range(n_bins, 0, -1):
        b = s[s['decile'] == d]
        bucket_pos = int(b['label'].sum())
        cum_pos += bucket_pos
        bucket_lr = float(b['label'].mean())
        row = {'decile': d, 'count': len(b),
               'score_min': round(float(b[score_col].min()), 4),
               'score_max': round(float(b[score_col].max()), 4),
               'label_rate': round(bucket_lr, 6),
               'lift': round(bucket_lr / overall_lr, 4) if overall_lr > 0 else None,
               'recall': round(bucket_pos / total_pos, 4) if total_pos > 0 else None,
               'cum_recall': round(cum_pos / total_pos, 4) if total_pos > 0 else None}
        for bc in biz_cols:
            row[bc] = round(float(b[bc].mean()), 4) if bc in b.columns else None
        result.append(row)
    return result


def build_segment_metrics(df, score_col, biz_cols, segment_cols, n_bins=10):
    """按分群列拆分计算全部指标。

    全量始终评估；每个分群列独立分群（多列不交叉），segment 名为 "{col}={value}"。
    数值型分群列按等频3分桶（固定，与 n_bins 无关；duplicates='drop' 处理重复值，
    实际桶数可能少于3），segment 名形如 "{col}=[low, high]"；NaN 单独作为 "{col}=nan" 一档。
    n_bins 仅影响模型打分的排序性分桶（decile_buckets）。

    Returns:
        (results_dict, segment_names) — segment_names 为有序列表（不含 '全量'）
    """
    results = {}
    segment_names = []

    results['全量'] = _compute_subset_metrics(df, score_col, biz_cols, n_bins)

    for col in segment_cols:
        if pd.api.types.is_numeric_dtype(df[col]):
            print(f"[INFO] 分群列 '{col}' 为数值类型, 按等频3分桶拆分")
            for seg_name, s in _iter_numeric_segments(df, col, n_bins=3):
                segment_names.append(seg_name)
                results[seg_name] = _compute_subset_metrics(s, score_col, biz_cols, n_bins)
        else:
            for val in sorted(df[col].astype(str).unique()):
                seg_name = f"{col}={val}"
                segment_names.append(seg_name)
                s = df[df[col].astype(str) == val]
                results[seg_name] = _compute_subset_metrics(s, score_col, biz_cols, n_bins)

    return results, segment_names


def _iter_numeric_segments(df, col, n_bins=10):
    """对数值列做等频N分桶, 按 (seg_name, subset_df) 顺序 yield。

    用 qcut 等频分桶, duplicates='drop' 处理重复值过多的情况（实际桶数可能少于 n_bins）。
    seg_name 形如 "{col}=[low, high]"。NaN 单独作为 "{col}=nan" 一档。

    Yields:
        (seg_name, subset_df)
    """
    bins = pd.qcut(df[col], q=n_bins, duplicates='drop')
    codes = bins.cat.codes
    for i, cat in enumerate(bins.cat.categories):
        s = df[codes == i]
        if len(s) > 0:
            yield f"{col}={cat}", s
    nan_mask = codes == -1
    if nan_mask.sum() > 0:
        yield f"{col}=nan", df[nan_mask]


def _compute_subset_metrics(s, score_col, biz_cols, n_bins=10):
    """对单档子集计算全部指标。"""
    n = len(s)
    if n < n_bins:
        return {'count': n, 'label_rate': None, 'auc': None, 'ks': None,
                'accuracy': None, 'precision': None, 'recall': None, 'f1': None,
                'biz_avg': {}, 'buckets': []}
    cls_metrics = compute_classification_metrics(s, score_col)
    return {
        'count': n,
        'label_rate': round(float(s['label'].mean()), 6),
        'auc': compute_auc(s, score_col),
        'ks': compute_ks(s, score_col) if n > 50 else None,
        'accuracy': cls_metrics['accuracy'],
        'precision': cls_metrics['precision'],
        'recall': cls_metrics['recall'],
        'f1': cls_metrics['f1'],
        'biz_avg': {bc: round(float(s[bc].mean()), 4) for bc in biz_cols if bc in s.columns},
        'buckets': decile_buckets(s, score_col, biz_cols, n_bins),
    }


# ============================================================
# 输出生成
# ============================================================
def build_json(model_meta, tags, df_info, metrics, source_file):
    """构建标准化评估 JSON"""
    return {
        "kb_version": "2.0.0",
        "source_file": str(source_file),
        "generated_at": datetime.now().isoformat(),
        "model_meta": model_meta,
        "tags": tags,
        "data_profile": {
            "observation_window": model_meta.get("observation_window", ""),
            "feature_count": model_meta.get("feature_count", None),
            "data_splits": [
                {"split_name": "全量", "sample_count": df_info["total_count"],
                 "label_rate": df_info["label_rate"]}
            ]
        },
        "metric_by_segment": {
            seg: {
                "count": m['count'],
                "label_rate": m['label_rate'],
                "auc": m['auc'],
                "ks": m['ks'],
                "accuracy": m.get('accuracy'),
                "precision": m.get('precision'),
                "recall": m.get('recall'),
                "f1": m.get('f1'),
                "biz_avg": m['biz_avg']
            }
            for seg, m in metrics.items()
        },
        "performance": {
            "score_buckets": {
                seg: m['buckets'] for seg, m in metrics.items()
            },
            "feature_importance": model_meta.get("feature_importance", [])
        },
        "hyperparameters": model_meta.get("hyperparameters", {})
    }


def build_md(model_name, version, metrics, segment_names):
    """生成 Markdown 报告"""
    full = metrics['全量']
    md = f"# {model_name} {version} 评估报告\n\n"
    md += f"> AUC={full['auc']:.4f}, KS={full['ks']:.4f}, 样本量={full['count']:,}, label率={full['label_rate']:.4%}\n\n"

    # 指标表格
    md += "## 指标 按分群\n\n"
    md += "| 分群 | 样本量 | label率 | AUC | KS |\n"
    md += "|------|--------|---------|-----|-----|\n"
    for seg in segment_names + ['全量']:
        m = metrics[seg]
        md += f"| {seg} | {m['count']:,} | {m['label_rate']:.4%} | {m['auc']:.4f} | {m['ks']:.4f} |\n"

    # 分桶
    md += "\n## 分桶排序性（全量）\n\n"
    biz_keys = [k for k in full['biz_avg'].keys()] if full['biz_avg'] else []
    header = "| 分桶 | 人数 | 分数区间 | label率 | lift | 召回率 | 累计召回 |"
    header += " | ".join(biz_keys) + " |\n" if biz_keys else "\n"
    sep = "|------|------|----------|--------|------|--------|----------|"
    sep += "---|" * len(biz_keys) + "\n" if biz_keys else "\n"
    md += header + sep
    for b in full['buckets']:
        lift_str = f"{b.get('lift', 0):.2f}" if b.get('lift') is not None else "-"
        recall_str = f"{b.get('recall', 0):.1%}" if b.get('recall') is not None else "-"
        cr_str = f"{b.get('cum_recall', 0):.1%}" if b.get('cum_recall') is not None else "-"
        row = f"| {b['decile']} | {b['count']} | [{b['score_min']}, {b['score_max']}] | {b['label_rate']:.4%} | {lift_str} | {recall_str} | {cr_str} |"
        row += "".join(f" {b.get(bk, 'N/A')} |" for bk in biz_keys) if biz_keys else ""
        row += "\n"
        md += row

    # 模型信息
    md += "\n## 模型信息\n\n"
    md += f"- **模型名称**: {model_name} {version}\n"
    md += f"- **评估日期**: {datetime.now().strftime('%Y-%m-%d')}\n"

    return md


def _is_rate_col(col_name):
    """判断是否为率类指标（cnt 结尾、label_rate、req_cnt、loan_cnt 等）"""
    name_lower = col_name.lower()
    return any(kw in name_lower for kw in ['rate', 'label', 'cnt', 'count'])

def _is_amt_col(col_name):
    """判断是否为金额类指标（amt 结尾、req_amt、loan_amt 等）"""
    name_lower = col_name.lower()
    return any(kw in name_lower for kw in ['amt', 'amount', '金额'])

def _format_bucket_value(cell, col_name, value):
    """根据列类型设置数值格式和样式"""
    cell.value = value
    if value is None:
        return
    if _is_rate_col(col_name):
        cell.number_format = '0.0%'  # 率：百分数，1位小数
    elif _is_amt_col(col_name):
        cell.number_format = '#,##0'  # 金额：整数，千分位
    elif col_name in ('label率', 'label_rate'):
        cell.number_format = '0.0%'

def _apply_blue_color_scale(ws, col_idx, min_row, max_row):
    """蓝色渐变条件格式"""
    col_letter = get_column_letter(col_idx)
    cell_range = f"{col_letter}{min_row}:{col_letter}{max_row}"
    ws.conditional_formatting.add(cell_range,
        ColorScaleRule(
            start_type='min', start_color='FFFFFF',
            mid_type='percentile', mid_value=50, mid_color='9DC3E6',
            end_type='max', end_color='2F5496'
        )
    )

def _apply_green_color_scale(ws, col_idx, min_row, max_row):
    """绿色渐变条件格式——label率专用"""
    col_letter = get_column_letter(col_idx)
    cell_range = f"{col_letter}{min_row}:{col_letter}{max_row}"
    ws.conditional_formatting.add(cell_range,
        ColorScaleRule(
            start_type='min', start_color='FFFFFF',
            mid_type='percentile', mid_value=50, mid_color='66BB6A',
            end_type='max', end_color='2E7D32'
        )
    )

def build_xlsx(model_name, version, metrics, segment_names, biz_cols, filepath):
    """生成 Excel"""
    wb = openpyxl.Workbook()

    # ---- Sheet 1: AUC/KS ----
    ws1 = wb.active
    ws1.title = "1-AUC_KS_by分群"
    ws1.cell(row=1, column=1, value="口径：AUC=ROC曲线下面积(越接近1越好) | KS=max(TPR-FPR)(越大区分度越高) | 分桶=按打分降序十等频").font = Font(name="微软雅黑", size=9, color="888888")
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5+len(biz_cols))
    ws1.cell(row=2, column=1, value=f"{model_name} {version} — AUC/KS 按分群").font = TITLE_FONT
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5 + len(biz_cols))

    # 结论句
    full = metrics.get('全量', {})
    auc_v = full.get('auc', 0)
    ks_v = full.get('ks', 0)
    n_v = full.get('count', 0)
    lr_v = full.get('label_rate', 0) if full.get('label_rate') else 0
    if auc_v >= 0.8:
        grade = '优秀'
    elif auc_v >= 0.7:
        grade = '良好'
    elif auc_v >= 0.6:
        grade = '一般'
    else:
        grade = '较差'
    f1_v = full.get('f1', 0) or 0
    acc_v = full.get('accuracy', 0) or 0
    conclusion = f'结论：全量 AUC={auc_v:.4f}，KS={ks_v:.4f}，准确率={acc_v:.1%}，F1={f1_v:.4f}，样本量={n_v:,}，标签率={lr_v:.1%}。模型区分度{grade}。'
    ws1.cell(row=2, column=1, value=conclusion).font = Font(name="微软雅黑", bold=True, size=11, color="C00000")
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5 + len(biz_cols))

    headers = ['分群', '样本量', 'label率', 'AUC', 'KS'] + list(biz_cols)
    r = 4
    for c, h in enumerate(headers, 1):
        ws1.cell(row=r, column=c, value=h)
    style_header_row(ws1, r, len(headers))
    r += 1

    for seg in segment_names + ['全量']:
        m = metrics[seg]
        vals = [seg, m['count'], m['label_rate'], m['auc'], m['ks']]
        vals += [m['biz_avg'].get(bc) for bc in biz_cols]
        for c, v in enumerate(vals, 1):
            ws1.cell(row=r, column=c, value=v)
        style_data_row(ws1, r, len(headers))
        r += 1
    ws1.freeze_panes = ws1.cell(row=5, column=1)

    # ---- Sheet 2: 分桶 ----
    ws2 = wb.create_sheet("2-分桶排序性by分群")
    ws2.cell(row=1, column=1, value="口径：label率=桶内正样本占比 | lift=桶label率/全量label率 | 召回率=桶正样本/总正样本 | 累计召回=从高到低累计捕获正样本占比").font = Font(name="微软雅黑", size=9, color="888888")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5+len(biz_cols))
    ws2.cell(row=2, column=1, value=f"{model_name} {version} — 分桶排序性（10等频）").font = TITLE_FONT
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5 + len(biz_cols))

    bucket_h = ['分群', '分桶', '人数', '分数区间', 'label率', 'lift', '召回率', '累计召回'] + list(biz_cols)
    r = 3
    for seg in segment_names + ['全量']:
        ws2.cell(row=r, column=1, value=f"分群: {seg} (N={metrics[seg]['count']})").font = SUB_FONT
        r += 1
        for c, h in enumerate(bucket_h, 1):
            ws2.cell(row=r, column=c, value=h)
        style_header_row(ws2, r, len(bucket_h))
        r += 1
        seg_data_start = r
        for b in metrics[seg]['buckets']:
            # decile
            ws2.cell(row=r, column=1, value=seg).font = CF
            ws2.cell(row=r, column=1).border = BD
            # 分桶号
            cell_d = ws2.cell(row=r, column=2, value=b['decile'])
            cell_d.font = CF; cell_d.border = BD
            # 人数
            cell_n = ws2.cell(row=r, column=3, value=b['count'])
            cell_n.font = CF; cell_n.border = BD
            # 分数区间
            cell_s = ws2.cell(row=r, column=4, value=f"[{b['score_min']}, {b['score_max']}]")
            cell_s.font = CF; cell_s.border = BD
            # label率
            cell_l = ws2.cell(row=r, column=5, value=b['label_rate'])
            _format_bucket_value(cell_l, 'label率', b['label_rate'])
            cell_l.font = CF; cell_l.border = BD
            # lift
            cell_lift = ws2.cell(row=r, column=6, value=b.get('lift'))
            cell_lift.font = CF; cell_lift.border = BD
            cell_lift.number_format = '0.00'
            # 召回率
            cell_recall = ws2.cell(row=r, column=7, value=b.get('recall'))
            cell_recall.font = CF; cell_recall.border = BD
            cell_recall.number_format = '0.0%'
            # 累计召回
            cell_cr = ws2.cell(row=r, column=8, value=b.get('cum_recall'))
            cell_cr.font = CF; cell_cr.border = BD
            cell_cr.number_format = '0.0%'
            # 业务指标
            for ci, bc in enumerate(biz_cols, 9):
                v = b.get(bc)
                cell_b = ws2.cell(row=r, column=ci, value=v)
                _format_bucket_value(cell_b, bc, v)
                cell_b.font = CF; cell_b.border = BD
            r += 1
        seg_data_end = r - 1
        # 条件格式：label率=绿色，lift/召回率/累计召回=蓝色，业务指标=蓝色
        data_ranges_green = [(5, seg_data_start, seg_data_end)]  # label率 列(绿色)
        data_ranges_blue = [
            (6, seg_data_start, seg_data_end),  # lift 列
            (7, seg_data_start, seg_data_end),  # 召回率 列
            (8, seg_data_start, seg_data_end),  # 累计召回 列
        ]
        for ci, bc in enumerate(biz_cols, 9):
            if _is_rate_col(bc) or _is_amt_col(bc):
                data_ranges_blue.append((ci, seg_data_start, seg_data_end))
        r += 1
    ws2.freeze_panes = ws2.cell(row=4, column=1)

    # 应用条件格式
    for col_idx, start_row, end_row in data_ranges_green:
        _apply_green_color_scale(ws2, col_idx, start_row, end_row)
    for col_idx, start_row, end_row in data_ranges_blue:
        _apply_blue_color_scale(ws2, col_idx, start_row, end_row)

    # ---- Sheet 3: 全量分桶 ----
    if metrics.get('全量') and metrics['全量'].get('buckets'):
        ws3 = wb.create_sheet("3-全量分桶")
        ws3.cell(row=1, column=1, value=f"{model_name} {version} — 全量分桶明细").font = TITLE_FONT
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5 + len(biz_cols))

        flat_h = ['分桶', '人数', '分数下限', '分数上限', 'label率', 'lift', '召回率', '累计召回'] + list(biz_cols)
        r = 3
        for c, h in enumerate(flat_h, 1):
            ws3.cell(row=r, column=c, value=h)
        style_header_row(ws3, r, len(flat_h))
        r += 1
        data_start = r
        for b in metrics['全量']['buckets']:
            vals = [
                b['decile'], b['count'], b['score_min'], b['score_max'],
                b['label_rate'], b.get('lift'), b.get('recall'), b.get('cum_recall')
            ] + [b.get(h) for h in biz_cols]
            for ci, v in enumerate(vals, 1):
                cell = ws3.cell(row=r, column=ci, value=v)
                _format_bucket_value(cell, flat_h[ci-1], v)
                cell.font = CF; cell.border = BD
            r += 1
        data_end = r - 1
        ws3.freeze_panes = ws3.cell(row=4, column=1)
        # 条件格式：label率=绿色, lift/召回/累计召回=蓝色，业务指标=蓝色
        for ci, h in enumerate(flat_h, 1):
            if h == 'label率':
                _apply_green_color_scale(ws3, ci, data_start, data_end)
            elif h in ('lift', '召回率', '累计召回') or _is_rate_col(h) or _is_amt_col(h):
                _apply_blue_color_scale(ws3, ci, data_start, data_end)

    wb.save(filepath)
    return wb


# ============================================================
# Main
# ============================================================
def _prepare_dataframe(df, score_col, segment_cols_arg, metric_cols_arg, label_col):
    """解析分群列 / 业务指标列 / 标签列，返回 (df, segment_cols, biz_cols, label_col)。

    多列分群时每列独立作为一个分群维度（不交叉拼接），segment 名由
    build_segment_metrics 生成形如 "{col}={value}"（离散列）或 "{col}=[low, high]"
    （数值列等频10分桶）。

    Args:
        df: 原始 DataFrame
        score_col: 打分列名
        segment_cols_arg: 分群列名列表（已解析）或 None（不分群）
        metric_cols_arg: 业务指标列名列表（已解析）或 None（不汇总业务指标）
        label_col: 标签列名

    Returns:
        (df, segment_cols, biz_cols, label_col)
    """
    if segment_cols_arg:
        missing_seg = [c for c in segment_cols_arg if c not in df.columns]
        if missing_seg:
            print(f"[WARN] 分群列不存在于数据中, 已忽略: {missing_seg}")
        segment_cols = [c for c in segment_cols_arg if c in df.columns]
    else:
        segment_cols = []

    if metric_cols_arg:
        missing_metric = [c for c in metric_cols_arg if c not in df.columns]
        if missing_metric:
            print(f"[WARN] 业务指标列不存在于数据中, 已忽略: {missing_metric}")
        biz_cols = [c for c in metric_cols_arg
                    if c in df.columns
                    and c not in (score_col, label_col)
                    and c not in segment_cols
                    and pd.api.types.is_numeric_dtype(df[c])]
    else:
        biz_cols = []

    if label_col not in df.columns:
        print(f"ERROR: 找不到标签列 '{label_col}'，可用列: {list(df.columns)}")
        sys.exit(1)
    df['label'] = df[label_col]

    # 防御性剔除: 标签缺失/非法样本不参与评估(尤其 OOT 评估), 避免 AUC/KS 因 NaN 报错或口径污染。
    n_invalid_label = int((~df['label'].isin([0, 1])).sum())
    if n_invalid_label:
        df = df[df['label'].isin([0, 1])].reset_index(drop=True)
        print(f"[INFO] 剔除 {n_invalid_label} 行 label 缺失/非法样本后评估, 剩余 {len(df)} 行")

    return df, segment_cols, biz_cols, label_col


def _eval_single_csv(df, source_file, out_dir, score_col, name, version,
                     model_type, hyperparams,
                     segment_cols_arg, metric_cols_arg, label_col,
                     n_bins=10, silent=False):
    """对单份 DataFrame 执行评估并落三件套。

    Returns:
        metrics dict（含 '全量' 等键），用于上层合并
    """
    df, segment_cols, biz_cols, label_col = _prepare_dataframe(
        df, score_col, segment_cols_arg, metric_cols_arg, label_col)

    if not silent:
        print(f"数据: {len(df)}行, 打分列={score_col}, 标签列={label_col}")
        print(f"分群列: {segment_cols if segment_cols else '无'}, 业务指标: {biz_cols}, 分桶数: {n_bins}")

    metrics, segment_names = build_segment_metrics(df, score_col, biz_cols, segment_cols, n_bins)

    model_meta = {
        "name": name,
        "version": version,
        "model_type": model_type,
        "target_type": "binary_classification",
        "eval_date": datetime.now().strftime("%Y-%m-%d"),
        "hyperparameters": json.loads(hyperparams) if hyperparams else {},
        "feature_importance": [],
        "feature_count": None,
        "observation_window": ""
    }
    tags = [name, version, model_type]
    df_info = {"total_count": len(df), "label_rate": round(float(df['label'].mean()), 6)}

    base_name = f"{name}_{version}"
    eval_json = build_json(model_meta, tags, df_info, metrics, source_file)
    json_path = out_dir / f"{base_name}_eval.json"
    json_path.write_text(json.dumps(eval_json, indent=2, ensure_ascii=False), encoding='utf-8')
    if not silent:
        print(f"[OK] {json_path}")

    md_content = build_md(name, version, metrics, segment_names)
    md_path = out_dir / f"{base_name}_eval.md"
    md_path.write_text(md_content, encoding='utf-8')
    if not silent:
        print(f"[OK] {md_path}")

    xlsx_path = out_dir / f"{base_name}_eval.xlsx"
    build_xlsx(name, version, metrics, segment_names, biz_cols, xlsx_path)
    if not silent:
        print(f"[OK] {xlsx_path}")

    return metrics, segment_names, biz_cols


def main():
    parser = argparse.ArgumentParser(description="单一模型评估（仅支持目录模式）")
    parser.add_argument("--input-dir", required=True, help="输入目录（目录下所有 CSV / parquet 各评估一份 + all 合并一份）")
    parser.add_argument("--score-col", required=True, help="打分列名（如 score_v3）")
    parser.add_argument("--name", default="模型名", help="模型名称")
    parser.add_argument("--model-type", default="xgboost", help="模型类型")
    parser.add_argument("--segment-cols", default=None, help="分群列名（多个逗号分隔，不传则不分群）")
    parser.add_argument("--metric-cols", default=None, help="业务指标列名（多个逗号分隔，不传则不汇总）")
    parser.add_argument("-o", "--output-dir", default=None, help="输出目录（默认：与输入目录同目录）")
    parser.add_argument("--label-col", default="label", help="标签列名")
    parser.add_argument("--hyperparams", default=None, help="超参JSON字符串")
    parser.add_argument("--n-bins", type=int, default=10, help="等频分桶数（默认10）")
    args = parser.parse_args()

    score_col = args.score_col
    data_path = Path(args.input_dir)
    if not data_path.is_dir():
        print(f"ERROR: --input-dir 必须指向目录， got: {data_path}")
        print("本脚本仅支持目录模式：目录下每个 CSV/parquet 单独评估一份 + all 合并一份。")
        sys.exit(1)

    out = Path(args.output_dir) if args.output_dir else data_path
    out.mkdir(parents=True, exist_ok=True)

    segment_cols_arg = [c.strip() for c in args.segment_cols.split(',')] if args.segment_cols else None
    metric_cols_arg = [c.strip() for c in args.metric_cols.split(',')] if args.metric_cols else None

    table_files = _list_table_files(data_path)
    if not table_files:
        print(f"ERROR: 目录 {data_path} 下无 CSV / parquet 文件")
        sys.exit(1)

    print(f"[目录模式] 发现 {len(table_files)} 个文件: {[f.name for f in table_files]}")

    pooled_dfs = []
    for tbl_file in table_files:
        stem = tbl_file.stem
        print(f"\n--- 评估 {tbl_file.name} (version={stem}) ---")
        df = _read_table(tbl_file)
        pooled_dfs.append(df)
        _eval_single_csv(
            df=df, source_file=str(tbl_file), out_dir=out,
            score_col=score_col, name=args.name, version=stem,
            model_type=args.model_type,
            hyperparams=args.hyperparams,
            segment_cols_arg=segment_cols_arg, metric_cols_arg=metric_cols_arg,
            label_col=args.label_col,
            n_bins=args.n_bins,
        )

    # 合并 all：纵向拼接所有文件行后重算指标（全量样本整体一份）
    print(f"\n--- 合并评估 all (version=all, 全量样本整体一份) ---")
    pooled_df = pd.concat(pooled_dfs, ignore_index=True)
    all_metrics, _, _ = _eval_single_csv(
        df=pooled_df, source_file=str(data_path), out_dir=out,
        score_col=score_col, name=args.name, version="all",
        model_type=args.model_type,
        hyperparams=args.hyperparams,
        segment_cols_arg=segment_cols_arg, metric_cols_arg=metric_cols_arg,
        label_col=args.label_col,
        n_bins=args.n_bins,
        silent=True,
    )
    full = all_metrics['全量']
    print(f"[OK] all 合并: AUC={full['auc']:.4f}, KS={full['ks']:.4f}, "
          f"样本量={full['count']:,}, label率={full['label_rate']:.4%}")

    print(f"\n{'='*50}")
    print(f"目录评估完成: {args.name}")
    print(f"单版本: {len(table_files)} 份, all 合并: 1 份 (version=all)")
    print(f"输出: {out.absolute()}")



if __name__ == "__main__":
    main()
