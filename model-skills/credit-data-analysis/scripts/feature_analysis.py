#!/usr/bin/env python3
"""信贷特征分析脚本（ModelEvo credit-data-analysis skill 版）。

读取数据文件，生成包含以下分析结果的 Excel 文件：
  1. 样本分布    2. 特征分布    3. 覆盖率    4. 均值    5. 最小值
  6. 最大值      7. 标准差      8. Nunique   9. PSI     10. IV

用法：
  python3 feature_analysis.py --data-file ka_df.feather \\
      --feature-list feature-list.csv \\
      --feature-extra ascore_fpd7_v3 \\
      --base-month 2025-04 --iv-label fpd7 \\
      --output-dir <产物目录>

支持 .feather / .csv / .parquet（按扩展名自动选读取方式）。
产物：Excel + Markdown + _manifest.json（参数溯源），落 --output-dir（默认当前目录）。

特征列来源（v2.5 修复）：主入口 --feature-list（精确选列，与全框架特征清单唯一真相一致）；
无清单时回退 --feature-start/--feature-end 区间法（DEPRECATED，独立体检模式兼容），
两者均缺报错提示迁移。

pipeline 模式（--split-config <feature_config.yaml>）：
  - 本 skill 是建模 pipeline 内部的特征分析环节（development Stage 0）。
  - 从 feature_config.yaml 的 model.split.oot_range 推导 PSI 基准月 = 第一个 OOT 月
    （用户可 --base-month 覆盖）；该基准月须经用户确认后执行。
  - 不做三档切分、不落盘 splits（切分后置到 training/tuning/evaluation 消费时即时进行）、
    不产 IV/PSI 筛选 csv（训练过程不通过 IV/PSI 指标筛选特征）。
"""

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

import _bootstrap  # noqa: F401  注入 _modelevo-shared/scripts（gen_feature_list 等共享模块）

import pandas as pd
import numpy as np
import warnings
import yaml

warnings.filterwarnings("ignore")

PRODUCED_BY = "skills/credit-data-analysis"
MANIFEST_SCHEMA_VERSION = 1

# ============================================================
# 0. 默认配置（CLI 参数可覆盖）
# ============================================================
DEFAULTS = dict(
    data_file="ka_df.feather",
    output_file="特征分析结果.xlsx",
    output_dir=".",
    time_col="fsx_time",
    feature_start="tx_model_2_score",
    feature_end="mob4_v5_score",
    feature_extra="",               # 逗号分隔的额外特征列名
    base_month="2025-04",           # PSI 基准月份
    iv_label="fpd7",                # IV 计算所用的风险标签，留空则自动选择
    risk_prefixes="fpd,dpd",        # 风险标签前缀，逗号分隔；iv_label 为空时用于自动识别
    risk_labels="",                 # 手动指定风险标签列（逗号分隔），留空则按前缀自动识别
    psi_bins=10,
    iv_bins=10,
    invalid_values="-1,-2,-9,-99,-999,-9999,-99999",  # 无效值哨兵集合，逗号分隔
)


def parse_args():
    ap = argparse.ArgumentParser(description="信贷特征分析脚本")
    ap.add_argument("--data-file", default=DEFAULTS["data_file"])
    ap.add_argument("--output-file", default=DEFAULTS["output_file"])
    ap.add_argument("--output-dir", default=DEFAULTS["output_dir"],
                    help="产物（Excel + Markdown + _manifest.json）输出目录，默认当前目录")
    ap.add_argument("--time-col", default=DEFAULTS["time_col"])
    ap.add_argument("--feature-list", default=None,
                    help="特征清单 CSV/TXT 路径（CSV 取 feature_name 列 / TXT 按行）；提供时精确选列，"
                         "替代 --feature-start/--feature-end 区间法（区间法已弃用，独立体检模式兼容保留）")
    ap.add_argument("--feature-start", default=DEFAULTS["feature_start"],
                    help="[DEPRECATED] 特征列范围的起始列名；--feature-list 提供时忽略")
    ap.add_argument("--feature-end", default=DEFAULTS["feature_end"],
                    help="[DEPRECATED] 特征列范围的结束列名；--feature-list 提供时忽略")
    ap.add_argument("--feature-extra", default=DEFAULTS["feature_extra"],
                    help="额外特征列名，逗号分隔（两种来源均追加）")
    ap.add_argument("--base-month", default=None,
                    help="PSI 基准月份，格式 YYYY-MM；pipeline 模式未指定时默认取第一个 OOT 月")
    ap.add_argument("--split-config", default=None,
                    help="pipeline 模式：feature_config.yaml 路径，用于推导 PSI 基准月（model.split.oot_range 首月）")
    ap.add_argument("--iv-label", default=DEFAULTS["iv_label"],
                    help="IV 计算所用的风险标签，留空则自动选择第一个可用的")
    ap.add_argument("--risk-prefixes", default=DEFAULTS["risk_prefixes"],
                    help="风险标签前缀，逗号分隔")
    ap.add_argument("--risk-labels", default=DEFAULTS["risk_labels"],
                    help="手动指定风险标签列，逗号分隔；覆盖自动识别")
    ap.add_argument("--psi-bins", type=int, default=DEFAULTS["psi_bins"])
    ap.add_argument("--iv-bins", type=int, default=DEFAULTS["iv_bins"])
    ap.add_argument("--invalid-values", default=DEFAULTS["invalid_values"],
                    help="无效值哨兵集合，逗号分隔；命中这些值的特征将在报告中标记提醒，建议建模时替换为空值")
    return ap.parse_args()


# ============================================================
# 1. 加载数据 & 准备列
# ============================================================
def _read_data(path: str) -> pd.DataFrame:
    """按扩展名选择读取方式：.feather / .csv / .parquet。"""
    if not os.path.exists(path):
        raise FileNotFoundError(f"数据文件不存在: {path}")
    suffix = Path(path).suffix.lower()
    if suffix == ".feather":
        return pd.read_feather(path)
    if suffix == ".parquet":
        return pd.read_parquet(path)
    if suffix == ".csv":
        return pd.read_csv(path)
    raise ValueError(
        f"不支持的文件格式: {suffix}。支持 .feather / .csv / .parquet（pandas 可读格式）"
    )


def load_and_prepare(args):
    print("[1/4] 加载数据...")
    df = _read_data(args.data_file)

    # 时间列 & 月份切片
    if args.time_col not in df.columns:
        raise KeyError(
            f"时间列 {args.time_col} 不存在。可选: {list(df.columns[:10])}..."
        )
    if not pd.api.types.is_datetime64_any_dtype(df[args.time_col]):
        # csv 读入后常为字符串/object，自动尝试解析为 datetime
        try:
            df[args.time_col] = pd.to_datetime(df[args.time_col])
        except (ValueError, TypeError) as e:
            raise ValueError(
                f"时间列 {args.time_col} 无法解析为日期，请确认其内容为日期格式: {e}"
            )
    df["month"] = df[args.time_col].dt.to_period("M").astype(str)
    months = sorted(df["month"].unique())

    # 特征列：--feature-list 精确选列（主入口）→ 未提供时区间法（DEPRECATED，独立体检兼容）
    cols = df.columns.tolist()
    feature_source = "feature-list"
    feature_list_missing = []
    if args.feature_list:
        try:
            from gen_feature_list import load_feature_list

            list_feat = load_feature_list(args.feature_list)
        except Exception as e:
            raise ValueError(f"读取 --feature-list 失败: {args.feature_list!r}: {e}")
        feature_cols = [c for c in list_feat if c in df.columns]
        # 清单中缺失的列仅 WARN（容忍列漂移），不报错
        feature_list_missing = [c for c in list_feat if c not in df.columns]
        if feature_list_missing:
            print(f"  ⚠ 特征清单 {len(feature_list_missing)} 列不在样本中（列漂移，已忽略）: "
                  f"{', '.join(feature_list_missing[:20])}")
        if not feature_cols:
            raise ValueError(
                f"特征清单无任何列存在于样本: {args.feature_list!r}"
            )
    elif args.feature_start in cols and args.feature_end in cols:
        feature_source = "interval"
        start_idx = cols.index(args.feature_start)
        end_idx = cols.index(args.feature_end)
        if start_idx > end_idx:
            raise ValueError(f"起始列在结束列之后: {args.feature_start} 出现在 {args.feature_end} 之后")
        feature_cols = cols[start_idx : end_idx + 1]
    else:
        raise ValueError(
            "特征列来源缺失：请提供 --feature-list（推荐主入口）或 --feature-start/--feature-end"
            "（DEPRECATED 区间法，独立体检兼容）；两者均未提供时无法确定特征范围"
        )
    if args.feature_extra:
        for extra in args.feature_extra.split(","):
            extra = extra.strip()
            if extra and extra not in feature_cols:
                feature_cols.append(extra)

    # 风险标签列
    if args.risk_labels:
        risk_labels = [x.strip() for x in args.risk_labels.split(",")]
    else:
        prefixes = [p.strip() for p in args.risk_prefixes.split(",")]
        risk_labels = [c for c in df.columns
                       if any(c.startswith(p) for p in prefixes)]

    # IV 标签：若未指定则从风险标签中自动选择
    iv_label = args.iv_label
    if not iv_label and risk_labels:
        iv_label = risk_labels[0]
        print(f"  → IV 标签自动选择: {iv_label}")

    # 无效值哨兵集合
    invalid_values = []
    for v in args.invalid_values.split(","):
        v = v.strip()
        if v:
            try:
                invalid_values.append(float(v))
            except ValueError:
                print(f"  ⚠ 无效值集合含非数值项，已忽略: {v!r}")
    if invalid_values:
        print(f"  无效值哨兵集合: {invalid_values}")

    print(f"  特征来源: {feature_source}（{'清单' if feature_source == 'feature-list' else '区间法'}"
          f"，共 {len(feature_cols)} 特征，清单缺失 {len(feature_list_missing)} 列）")
    print(f"  特征数: {len(feature_cols)}")
    print(f"  风险标签: {risk_labels}")
    print(f"  月份: {months}")
    print(f"  样本量: {len(df):,}")
    return df, months, feature_cols, risk_labels, iv_label, invalid_values, feature_source, feature_list_missing


# ============================================================
# 1.5 PSI 基准月解析（pipeline 模式：默认取第一个 OOT 月）
# ============================================================
def _month_from_range_start(value: str) -> str:
    """把 split 区间起始日期规范为 YYYY-MM。

    兼容 8 位 YYYYMMDD 与 YYYY-MM-DD 两种写法（与 feature_config 校验口径一致）。
    """
    s = str(value).strip().replace("-", "").replace("/", "")
    if len(s) < 6:
        raise ValueError(f"split 区间起始日期无法解析为月份: {value!r}")
    return f"{s[:4]}-{s[4:6]}"


def resolve_base_month(args, months) -> str:
    """确定 PSI 基准月。

    优先级：显式 --base-month > pipeline 模式 split_config 的 oot_range 首月 > 默认值。
    pipeline 模式推导出的基准月须经用户确认（LLM 在编排层向用户确认后传入 --base-month）。
    """
    if args.base_month:
        return args.base_month
    if args.split_config:
        try:
            cfg = yaml.safe_load(open(args.split_config, encoding="utf-8")) or {}
            oot_range = (cfg.get("model") or {}).get("split", {}).get("oot_range")
        except (OSError, IOError, yaml.YAMLError) as e:
            raise ValueError(f"读取 --split-config 失败: {args.split_config!r}: {e}")
        if not oot_range or len(oot_range) < 1:
            raise ValueError(
                "pipeline 模式推导 PSI 基准月失败: feature_config.yaml 缺少 model.split.oot_range"
            )
        base_month = _month_from_range_start(oot_range[0])
        print(f"  → pipeline 模式: PSI 基准月 = 第一个 OOT 月 {base_month} "
              f"(oot_range 起始 {oot_range[0]!r})，请用户确认")
        if base_month not in months:
            print(f"  ⚠ 注意: 基准月 {base_month} 不在数据月份范围内，PSI 各月均与无样本基准对比，结果仅供参考")
        return base_month
    print(f"  ⚠ 未指定 --base-month 且无 --split-config，使用默认基准月 {DEFAULTS['base_month']}")
    return DEFAULTS["base_month"]


# ============================================================
# 2. 辅助函数
# ============================================================
def _get_bins(series, n_bins):
    s = series.dropna()
    if len(s) < n_bins * 2 or s.nunique() < 2:
        return None
    try:
        _, edges = pd.qcut(s, n_bins, duplicates="drop", retbins=True)
        if len(edges) < 2:
            return None
        return edges
    except (ValueError, IndexError):
        return None


def calc_psi(base_series, actual_series, bins=10):
    base = base_series.dropna()
    actual = actual_series.dropna()
    if len(base) < bins * 2 or len(actual) < bins * 2:
        return np.nan
    edges = _get_bins(base, bins)
    if edges is None:
        return np.nan
    base_dist = pd.cut(base, edges, include_lowest=True).value_counts(normalize=True).sort_index()
    actual_dist = pd.cut(actual, edges, include_lowest=True).value_counts(normalize=True).sort_index()
    all_idx = base_dist.index.union(actual_dist.index)
    base_dist = base_dist.reindex(all_idx, fill_value=0.0001)
    actual_dist = actual_dist.reindex(all_idx, fill_value=0.0001)
    return float(np.sum((actual_dist.values - base_dist.values)
                        * np.log(actual_dist.values / base_dist.values)))


def calc_iv(series, label, bins=10):
    mask = series.notna() & label.notna()
    x = series[mask]
    y = label[mask]
    if len(x) < bins * 2 or y.nunique() < 2:
        return np.nan
    edges = _get_bins(x, bins)
    if edges is None:
        return np.nan
    binned = pd.cut(x, edges, include_lowest=True)
    grouped_good = (y == 0).groupby(binned).sum()
    grouped_bad = (y == 1).groupby(binned).sum()
    total_good = grouped_good.sum()
    total_bad = grouped_bad.sum()
    if total_good == 0 or total_bad == 0:
        return np.nan
    good_pct = (grouped_good / total_good).clip(lower=0.0001)
    bad_pct = (grouped_bad / total_bad).clip(lower=0.0001)
    woe = np.log(good_pct / bad_pct)
    return float(np.sum((good_pct - bad_pct) * woe))


def check_invalid_values(df, feature_cols, invalid_values):
    """检测特征列中是否含无效值哨兵（如 -1/-2/-999/-9999 等）。

    命中规则：特征取值落在无效值集合内的样本数 > 0。
    输出：每特征的命中哨兵值集合、命中样本数、命中占比、是否建议替换为空值。
    这些值往往是"无数据/拒贷/异常"的占位符，若不替换为空值，建模时模型会
    学到虚假的取值边界（尤其树模型/分箱），故标记提醒，由人决策是否清洗。
    """
    if not invalid_values:
        return None
    rows = []
    for fc in feature_cols:
        s = df[fc]
        if not pd.api.types.is_numeric_dtype(s):
            continue
        hit = set()
        for v in invalid_values:
            try:
                if (s == v).any():
                    hit.add(v)
            except (TypeError, ValueError):
                continue
        if hit:
            hit_list = sorted(hit)
            mask = s.isin(hit_list)
            n_hit = int(mask.sum())
            rows.append({
                "特征": fc,
                "命中无效值": ",".join(str(int(v)) if float(v).is_integer() else str(v) for v in hit_list),
                "命中样本数": n_hit,
                "命中占比": round(n_hit / len(s), 6),
                "建议": "建议建模时替换为空值(NaN)",
            })
    if not rows:
        return None
    return pd.DataFrame(rows)


# ============================================================
# 3. 计算所有分析结果
# ============================================================
def compute_all(df, months, feature_cols, risk_labels, iv_label, invalid_values, args):
    print("[2/4] 计算分析结果...")

    # --- 3.1 样本分布 ---
    print("  样本分布...")
    sample_total, sample_overdue, sample_rate = {}, {}, {}
    for month in months:
        mdf = df[df["month"] == month]
        sample_total[month] = {}
        sample_overdue[month] = {}
        sample_rate[month] = {}
        for rl in risk_labels:
            valid = mdf[rl].notna()
            total = valid.sum()
            overdue = (mdf.loc[valid, rl] == 1).sum()
            sample_total[month][rl] = int(total)
            sample_overdue[month][rl] = int(overdue)
            sample_rate[month][rl] = round(overdue / total, 6) if total > 0 else 0.0

    sample_total_df = pd.DataFrame.from_dict(sample_total, orient="index", columns=risk_labels)
    sample_total_df.index.name = "月份"
    sample_overdue_df = pd.DataFrame.from_dict(sample_overdue, orient="index", columns=risk_labels)
    sample_overdue_df.index.name = "月份"
    sample_rate_df = pd.DataFrame.from_dict(sample_rate, orient="index", columns=risk_labels)
    sample_rate_df.index.name = "月份"

    # --- 3.2 特征分布 ---
    print("  特征分布...")
    feat_dist_data = {}
    for fc in feature_cols:
        s = df[fc]
        feat_dist_data[fc] = {
            "覆盖率": s.notna().mean(),
            "平均值": s.mean(),
            "最小值": s.min(),
            "1%": s.quantile(0.01),
            "5%": s.quantile(0.05),
            "25%": s.quantile(0.25),
            "50%": s.quantile(0.50),
            "75%": s.quantile(0.75),
            "95%": s.quantile(0.95),
            "99%": s.quantile(0.99),
            "最大值": s.max(),
        }
    feat_dist_df = pd.DataFrame.from_dict(feat_dist_data, orient="index")
    feat_dist_df.index.name = "特征"

    # --- 3.3-3.8 分月统计表 ---
    monthly_funcs = {
        "覆盖率": lambda s: s.notna().mean(),
        "均值": lambda s: s.mean(),
        "最小值": lambda s: s.min(),
        "最大值": lambda s: s.max(),
        "标准差": lambda s: s.std(),
        "Nunique": lambda s: s.nunique(),
    }
    monthly_results = {}
    for sheet_name, func in monthly_funcs.items():
        print(f"  {sheet_name}...")
        data = {fc: df.groupby("month")[fc].apply(func) for fc in feature_cols}
        monthly_results[sheet_name] = pd.DataFrame(data).T
        monthly_results[sheet_name].index.name = "特征"

    # --- 3.9 PSI ---
    print(f"  PSI (以 {args.base_month} 为基准)...")
    base_mask = df["month"] == args.base_month
    psi_data = {}
    for fc in feature_cols:
        base_s = df.loc[base_mask, fc]
        row = {}
        for month in months:
            if month == args.base_month:
                row[month] = 0.0
            else:
                row[month] = calc_psi(base_s, df.loc[df["month"] == month, fc], args.psi_bins)
        psi_data[fc] = row
    psi_df = pd.DataFrame.from_dict(psi_data, orient="index")
    psi_df.index.name = "特征"

    # --- 3.10 IV ---
    print(f"  IV (标签={iv_label})...")
    label_series = df[iv_label]
    iv_data = {}
    for fc in feature_cols:
        row = {}
        for month in months:
            mask = df["month"] == month
            row[month] = calc_iv(df.loc[mask, fc], label_series[mask], args.iv_bins)
        iv_data[fc] = row
    iv_df = pd.DataFrame.from_dict(iv_data, orient="index")
    iv_df.index.name = "特征"

    # --- 3.11 无效值检查 ---
    print("  无效值检查...")
    invalid_df = check_invalid_values(df, feature_cols, invalid_values)
    if invalid_df is not None and len(invalid_df):
        print(f"  ⚠ 发现 {len(invalid_df)} 个特征含无效值哨兵（如 -1/-2/-999/-9999 等），"
              f"详情见「无效值检查」sheet，建议建模时替换为空值(NaN)")
        top_inv = invalid_df.sort_values("命中占比", ascending=False).head(5)
        for _, r in top_inv.iterrows():
            print(f"    - {r['特征']}: 命中 {r['命中无效值']} 占比 {r['命中占比']:.2%}")
    else:
        print("  ✓ 未发现无效值哨兵特征")

    return (sample_total_df, sample_overdue_df, sample_rate_df,
            feat_dist_df, monthly_results, psi_df, iv_df, invalid_df)


# ============================================================
# 4. 写入 Excel
# ============================================================
def write_excel(output_file, risk_labels, months,
                sample_total_df, sample_overdue_df, sample_rate_df,
                feat_dist_df, monthly_results, psi_df, iv_df, invalid_df):
    print("[3/4] 写入 Excel...")
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    sheet_order = ["特征分布", "覆盖率", "均值", "最小值", "最大值", "标准差", "Nunique", "PSI", "IV"]
    sheet_dfs_map = {
        "特征分布": feat_dist_df,
        "覆盖率": monthly_results["覆盖率"], "均值": monthly_results["均值"],
        "最小值": monthly_results["最小值"], "最大值": monthly_results["最大值"],
        "标准差": monthly_results["标准差"], "Nunique": monthly_results["Nunique"],
        "PSI": psi_df, "IV": iv_df,
    }

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for name in sheet_order:
            sheet_dfs_map[name].to_excel(writer, sheet_name=name, index=True)
        if invalid_df is not None and len(invalid_df):
            invalid_df.to_excel(writer, sheet_name="无效值检查", index=False)

    wb = load_workbook(output_file)
    ws = wb.create_sheet("样本分布", 0)

    def write_matrix(ws, start_row, title, df):
        ws.cell(row=start_row, column=1, value=title)
        ws.cell(row=start_row + 1, column=1, value="月份")
        for j, col in enumerate(df.columns):
            ws.cell(row=start_row + 1, column=j + 2, value=col)
        for i, (idx, row) in enumerate(df.iterrows()):
            ws.cell(row=start_row + 2 + i, column=1, value=str(idx))
            for j, val in enumerate(row):
                ws.cell(row=start_row + 2 + i, column=j + 2, value=val)

    write_matrix(ws, 1, "样本数", sample_total_df)
    write_matrix(ws, len(months) + 4, "逾期数", sample_overdue_df)
    write_matrix(ws, 2 * len(months) + 7, "逾期率", sample_rate_df)

    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, len(risk_labels) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    for name in sheet_order:
        wss = wb[name]
        wss.column_dimensions["A"].width = 18
        for col_idx in range(2, wss.max_column + 1):
            wss.column_dimensions[get_column_letter(col_idx)].width = 14

    wb.save(output_file)


# ============================================================
# 4.5 写入 Markdown 报告（与 Excel 同源，人工可读）
# ============================================================
def write_markdown(output_file, args, months, feature_cols, risk_labels,
                   sample_total_df, sample_overdue_df, sample_rate_df,
                   feat_dist_df, monthly_results, psi_df, iv_df, invalid_df,
                   feature_source="interval", feature_list_missing=()):
    print("[3.5/4] 写入 Markdown 报告...")
    lines = []
    lines.append("# 信贷特征分析报告")
    lines.append("")
    lines.append(f"> 数据文件: `{args.data_file}`")
    lines.append(f"> 时间列: `{args.time_col}` ｜ PSI 基准月: `{args.base_month}` ｜ IV 标签: `{args.iv_label or '（自动）'}`")
    if args.split_config:
        lines.append(f"> pipeline 模式: `{args.split_config}`（PSI 基准月 = 第一个 OOT 月，须用户确认）")
    if args.feature_list:
        lines.append(f"> 特征来源: 特征清单 `{args.feature_list}`（命中 {len(feature_cols)} 列，清单缺失 {len(feature_list_missing)} 列）")
    else:
        lines.append("> 特征来源: 区间法（DEPRECATED，独立体检模式兼容）")
    lines.append(f"> 样本量: `{int(sample_total_df.values.sum() if sample_total_df.size else 0):,}` ｜ 月份: `{months[0]}` ~ `{months[-1]}`（{len(months)} 个月）")
    lines.append("")

    # --- 样本分布 ---
    lines.append("## 一、样本分布")
    lines.append("")
    for title, df in [("样本数", sample_total_df), ("逾期数", sample_overdue_df), ("逾期率", sample_rate_df)]:
        lines.append(f"### {title}")
        lines.append("")
        lines.append("| 月份 | " + " | ".join(df.columns) + " |")
        lines.append("|---|" + "---|" * len(df.columns))
        for idx, row in df.iterrows():
            cells = " | ".join(f"{v:.4f}" if isinstance(v, float) else str(v) for v in row)
            lines.append(f"| {idx} | {cells} |")
        lines.append("")
    if risk_labels:
        latest = sample_rate_df.index[-1]
        avg = sample_rate_df.loc[latest]
        lines.append(f"> 最新月 `{latest}` 逾期率: " + "; ".join(
            f"`{rl}={avg[rl]:.4f}`" for rl in risk_labels if rl in avg))
        lines.append("")

    # --- 特征分布 ---
    lines.append("## 二、特征分布（全时段）")
    lines.append("")
    lines.append("| 特征 | 覆盖率 | 平均值 | 最小值 | 25% | 50% | 75% | 最大值 |")
    lines.append("|---|------:|------:|------:|----:|----:|----:|------:|")
    for fc, row in feat_dist_df.iterrows():
        lines.append(
            f"| {fc} | {row['覆盖率']:.2%} | {row['平均值']:.4g} | {row['最小值']:.4g} "
            f"| {row['25%']:.4g} | {row['50%']:.4g} | {row['75%']:.4g} | {row['最大值']:.4g} |"
        )
    lines.append("")

    # --- PSI 摘要（标红告警）---
    lines.append(f"## 三、分月 PSI（基准月 `{args.base_month}`）")
    lines.append("")
    lines.append("| 特征 | " + " | ".join(months) + " |")
    lines.append("|---|" + "---|" * len(months))
    for fc in psi_df.index:
        cells = " | ".join(
            f"{v:.4f}" if isinstance(v, float) and not pd.isna(v) else "—" for v in psi_df.loc[fc]
        )
        lines.append(f"| {fc} | {cells} |")
    lines.append("")
    psi_warn = (psi_df > 0.10).any(axis=1)
    warn_list = psi_warn[psi_warn].index.tolist()
    if warn_list:
        lines.append(f"> ⚠ **PSI 告警**（>0.10 的特征，共 {len(warn_list)} 个）: {', '.join(map(str, warn_list[:20]))}")
    else:
        lines.append("> ✓ 无 PSI>0.10 告警特征")
    lines.append("")

    # --- IV 摘要 ---
    lines.append(f"## 四、分月 IV（标签 `{args.iv_label}`）")
    lines.append("")
    lines.append("| 特征 | " + " | ".join(months) + " |")
    lines.append("|---|" + "---|" * len(months))
    for fc in iv_df.index:
        cells = " | ".join(
            f"{v:.4f}" if isinstance(v, float) and not pd.isna(v) else "—" for v in iv_df.loc[fc]
        )
        lines.append(f"| {fc} | {cells} |")
    lines.append("")

    # --- 无效值检查 ---
    lines.append("## 五、无效值检查")
    lines.append("")
    if invalid_df is not None and len(invalid_df):
        lines.append("| 特征 | 命中无效值 | 命中样本数 | 命中占比 |")
        lines.append("|---|------|------:|------:|")
        for _, r in invalid_df.iterrows():
            lines.append(f"| {r['特征']} | {r['命中无效值']} | {int(r['命中样本数'])} | {r['命中占比']:.2%} |")
        lines.append("")
        lines.append("> ⚠ 上述特征含哨兵值占位（无数据/拒贷/异常），建模前建议在 data-cleaning 阶段替换为 NaN。")
    else:
        lines.append("> ✓ 未发现无效值哨兵特征。")
    lines.append("")

    md_path = str(Path(output_file).with_suffix(".md"))
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  Markdown 报告: {md_path}")


# ============================================================
# 5. 落盘 manifest（参数溯源，供复现与追溯）
# ============================================================
def write_manifest(output_dir: Path, files: list, args, feature_source="interval",
                   feature_list_missing=()) -> None:
    params = {
        "data_file": args.data_file,
        "feature_extra": args.feature_extra,
        "feature_source": feature_source,
        "base_month": args.base_month,
        "split_config": args.split_config,
        "iv_label": args.iv_label,
        "time_col": args.time_col,
        "risk_prefixes": args.risk_prefixes,
        "risk_labels": args.risk_labels,
        "psi_bins": args.psi_bins,
        "iv_bins": args.iv_bins,
        "invalid_values": args.invalid_values,
    }
    if args.feature_list:
        params["feature_list"] = args.feature_list
        params["feature_list_missing"] = list(feature_list_missing)
    else:
        # 兼容保留（区间法 DEPRECATED）
        params["feature_start"] = args.feature_start
        params["feature_end"] = args.feature_end
    manifest = {
        "schema_version": MANIFEST_SCHEMA_VERSION,
        "produced_by": PRODUCED_BY,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "params": params,
        "files": sorted(files),
    }
    (output_dir / "_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )


# ============================================================
# 6. 主入口
# ============================================================
def main():
    args = parse_args()
    (df, months, feature_cols, risk_labels, iv_label, invalid_values,
     feature_source, feature_list_missing) = load_and_prepare(args)
    # PSI 基准月：pipeline 模式默认取第一个 OOT 月（须用户确认），显式 --base-month 覆盖
    args.base_month = resolve_base_month(args, months)
    results = compute_all(df, months, feature_cols, risk_labels, iv_label, invalid_values, args)
    (sample_total_df, sample_overdue_df, sample_rate_df,
     feat_dist_df, monthly_results, psi_df, iv_df, invalid_df) = results

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    output_file = str(out_dir / args.output_file)

    write_excel(output_file, risk_labels, months,
                sample_total_df, sample_overdue_df, sample_rate_df,
                feat_dist_df, monthly_results, psi_df, iv_df, invalid_df)
    write_markdown(output_file, args, months, feature_cols, risk_labels,
                   sample_total_df, sample_overdue_df, sample_rate_df,
                   feat_dist_df, monthly_results, psi_df, iv_df, invalid_df,
                   feature_source=feature_source, feature_list_missing=feature_list_missing)
    write_manifest(out_dir, [args.output_file, Path(args.output_file).with_suffix(".md").name],
                   args, feature_source=feature_source, feature_list_missing=feature_list_missing)

    print("[4/4] 完成!")
    print(f"  输出文件: {output_file}")
    print(f"  Markdown 报告: {str(Path(output_file).with_suffix('.md'))}")
    print(f"  样本分布: {len(months)} 月 × {len(risk_labels)} 标签")
    print(f"  特征分布: {feat_dist_df.shape[0]} 特征 × {feat_dist_df.shape[1]} 统计量")
    for name in ["覆盖率", "均值", "最小值", "最大值", "标准差", "Nunique", "PSI", "IV"]:
        df_out = (monthly_results[name] if name in monthly_results
                  else psi_df if name == "PSI" else iv_df)
        print(f"  {name}: {df_out.shape[0]} 特征 × {df_out.shape[1]} 月份")
    if invalid_df is not None and len(invalid_df):
        print(f"  无效值检查: {len(invalid_df)} 个特征命中哨兵值（见「无效值检查」sheet）")


if __name__ == "__main__":
    sys.exit(main() or 0)
