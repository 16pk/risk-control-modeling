# -*- coding: utf-8 -*-
"""model_report — session 级 6-sheet 综合报告。

输入 <session_dir>, 扫所有上游产物, 产 6 个 sheet 的 xlsx 落到 <session_dir>/。

6 个 sheet (与 SKILL.md §4.1 对齐):
  1. 模型概览       — task-spec/_manifest.json (需求规格 + 路由 + 样本概况 + 切分配置)
  2. 样本分析       — data-profile/_manifest.json + _split_manifest.json (分时段 + 三档切分 + 稳定性)
  3. 特征质量       — feature-analysis/analysis/{_manifest.json, iv_table.csv, psi_table.csv, stats.csv}
  4. 三档评估       — new-models/*/evaluation/*_{split}_eval.json (多 run × 8 指标, 按 split 分块)
  5. 分桶排序性对比  — new-models/*/evaluation/*_{split}_eval.json performance.score_buckets['全量'] (多 run 并排)
  6. 特征清单       — AUC 最高的 run 的 explainability/{feature-importance.csv, shap-summary.csv}

不受 development / orchestration 调度, 仅用户主动调起。

用法:
    python build_report.py --session-dir <session_dir> [-o <output.xlsx>]
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ===== 样式常量 (参考 classification-model-evaluation/scripts/eval_single.py) =====
HF = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="1A3060")
CF = Font(name="微软雅黑", size=10)
BD = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1A3060")
SUB_FONT = Font(name="微软雅黑", size=11, color="4472C4")
WARN_FONT = Font(name="微软雅黑", size=10, italic=True, color="C0392B")
SRC_FONT = Font(name="微软雅黑", size=9, italic=True, color="7F7F7F")
SRC_FILL = PatternFill("solid", fgColor="F2F2F2")
DEF_FONT = Font(name="微软雅黑", size=9, color="333333")
DEF_FILL = PatternFill("solid", fgColor="FAFAFA")
# section 标签行样式 (灰底蓝字合并, 用于 KV sheet 分段)
SECTION_FILL = PatternFill("solid", fgColor="D9D9D9")
SECTION_FONT = Font(name="微软雅黑", size=11, bold=True, color="1A3060")
# split 标签行样式 (同 section, 用于多 split 子表堆叠)
SPLIT_FILL = PatternFill("solid", fgColor="D9D9D9")
SPLIT_FONT = Font(name="微软雅黑", size=11, bold=True, color="1A3060")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
SRC_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
DEF_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ===== DataBar 条件格式 helpers (参考 classification-model-comparison/scripts/compare_models.py) =====

def _green_scale(ws, col_idx: int, min_row: int, max_row: int) -> None:
    """绿色 DataBar (KS / 召回率 / 正样本率 / label率)."""
    col_letter = get_column_letter(col_idx)
    rng = f"{col_letter}{min_row}:{col_letter}{max_row}"
    ws.conditional_formatting.add(
        rng,
        DataBarRule(start_type="min", end_type="max", color="2E7D32", showValue=True),
    )


def _blue_scale(ws, col_idx: int, min_row: int, max_row: int) -> None:
    """蓝色 DataBar (AUC / 重要性 / SHAP / 召回率 / 累计召回)."""
    col_letter = get_column_letter(col_idx)
    rng = f"{col_letter}{min_row}:{col_letter}{max_row}"
    ws.conditional_formatting.add(
        rng,
        DataBarRule(start_type="min", end_type="max", color="1A3060", showValue=True),
    )


def _red_scale(ws, col_idx: int, min_row: int, max_row: int) -> None:
    """红色 DataBar (PSI / 缺失率)."""
    col_letter = get_column_letter(col_idx)
    rng = f"{col_letter}{min_row}:{col_letter}{max_row}"
    ws.conditional_formatting.add(
        rng,
        DataBarRule(start_type="min", end_type="max", color="C00000", showValue=True),
    )


# ===== IO helpers =====

def _read_json(path: Path) -> Optional[dict]:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        print(f"[build_report] WARN: {path} 解析失败: {e}", file=sys.stderr)
        return None


def _read_csv_rows(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _fmt_pct(x: Any) -> str:
    try:
        return f"{float(x) * 100:.2f}%"
    except (TypeError, ValueError):
        return "—"


def _fmt_num(x: Any) -> str:
    try:
        return f"{int(x):,}"
    except (TypeError, ValueError):
        return str(x) if x is not None else "—"


def _fmt_auc(x: Any) -> str:
    try:
        return f"{float(x):.4f}"
    except (TypeError, ValueError):
        return "—"


def _to_float(x: Any) -> Any:
    """转 float; None / 不可转 → '—'。"""
    if x is None:
        return "—"
    try:
        return float(x)
    except (TypeError, ValueError):
        return "—"


def _to_int(x: Any) -> Any:
    """转 int; None / 不可转 → '—'。兼容 '1234.0' 字符串。"""
    if x is None:
        return "—"
    try:
        return int(x)
    except (TypeError, ValueError):
        try:
            return int(float(x))
        except (TypeError, ValueError):
            return "—"


def _to_str(x: Any) -> str:
    """转 str; None → '—'。"""
    if x is None:
        return "—"
    if isinstance(x, (list, tuple)):
        return ", ".join(str(i) for i in x) if x else "—"
    return str(x)


# ===== Run discovery =====

def _discover_runs(session_dir: Path) -> Tuple[List[Dict[str, Any]], List[Tuple[str, str]]]:
    """扫 <session_dir>/new-models/*/ 下所有 run, 收集 config/manifest/evals。

    返回 (runs_data, skipped):
      runs_data = [{run_dir, run_name, cfg, model_manifest, evals: {train/test/oot: dict|None}}, ...]
      skipped = [(run_name, reason), ...]
    """
    new_models = session_dir / "new-models"
    if not new_models.exists():
        return ([], [])

    run_dirs = sorted(
        [d for d in new_models.iterdir() if d.is_dir()],
        key=lambda d: d.name,
    )

    runs_data: List[Dict[str, Any]] = []
    skipped: List[Tuple[str, str]] = []

    for d in run_dirs:
        cfg = _read_json(d / "config.json")
        evals = {
            sp: _read_json(d / "evaluation" / f"{d.name}_{sp}_eval.json")
            for sp in ("train", "test", "oot", "all")
        }
        if not cfg:
            skipped.append((d.name, "config.json 缺失"))
            continue
        if not any(evals.values()):
            skipped.append((d.name, "evaluation/ 下 train/test/oot/all eval JSON 全部缺失"))
            continue
        model_manifest = _read_json(d / "model" / "_manifest.json")
        runs_data.append({
            "run_dir": d,
            "run_name": d.name,
            "cfg": cfg,
            "model_manifest": model_manifest,
            "evals": evals,
        })

    return (runs_data, skipped)


def _find_best_run(runs_data: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    """返回 oot AUC 最高的 run (oot 缺失则 fallback 到 train AUC, 都缺返回 None)。"""
    best = None
    best_auc = -1.0
    for r in runs_data:
        metrics = r["cfg"].get("runtime", {}).get("metrics", {}) or {}
        oot_auc = (metrics.get("oot") or {}).get("auc")
        train_auc = (metrics.get("train") or {}).get("auc")
        auc = None
        if oot_auc is not None:
            try:
                auc = float(oot_auc)
            except (TypeError, ValueError):
                auc = None
        if auc is None and train_auc is not None:
            try:
                auc = float(train_auc)
            except (TypeError, ValueError):
                auc = None
        if auc is not None and auc > best_auc:
            best_auc = auc
            best = r
    return best


# ===== Sheet writer (通用 — 表格式) =====

def _write_table_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    headers: List[str],
    rows: List[List[Any]],
    warn: Optional[str] = None,
    source_note: Optional[str] = None,
    number_formats: Optional[Dict[int, str]] = None,
    definitions: Optional[List[str]] = None,
) -> None:
    """通用: 单标题 + 数据来源说明(可选) + 指标定义说明(可选) + 表头 + 数据行 + warning。"""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    n_cols = max(len(headers), 1)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    cur_row = 2
    if source_note:
        ws.cell(row=cur_row, column=1, value=f"数据来源: {source_note}")
        ws.cell(row=cur_row, column=1).font = SRC_FONT
        ws.cell(row=cur_row, column=1).fill = SRC_FILL
        ws.cell(row=cur_row, column=1).alignment = SRC_ALIGN
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
        ws.row_dimensions[cur_row].height = 28
        cur_row += 1

    if definitions:
        for line in definitions:
            ws.cell(row=cur_row, column=1, value=line)
            ws.cell(row=cur_row, column=1).font = DEF_FONT
            ws.cell(row=cur_row, column=1).fill = DEF_FILL
            ws.cell(row=cur_row, column=1).alignment = DEF_ALIGN
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
            cur_row += 1

    header_row = cur_row

    if not headers:
        ws.cell(row=header_row, column=1, value=warn or "数据缺失").font = WARN_FONT
        ws.column_dimensions["A"].width = 60
        return

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = HF
        cell.fill = HFILL
        cell.alignment = CENTER
        cell.border = BD

    data_start = header_row + 1
    for r, row in enumerate(rows, data_start):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = CF
            cell.border = BD
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                cell.alignment = RIGHT
                if number_formats and c in number_formats:
                    cell.number_format = number_formats[c]
            else:
                cell.alignment = LEFT

    for c, h in enumerate(headers, 1):
        col_letter = get_column_letter(c)
        max_len = len(str(h))
        for row in rows:
            v = row[c - 1] if c - 1 < len(row) else ""
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 50)

    ws.freeze_panes = f"A{data_start}"

    if warn:
        nr = len(rows) + data_start + 1
        ws.cell(row=nr, column=1, value=f"⚠ {warn}").font = WARN_FONT
        ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=len(headers))


# ===== Sheet writer (KV 式 — 分段 key/value) =====

def _write_kv_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    sections: List[Tuple[str, List[Tuple[str, Any]]]],
    source_note: Optional[str] = None,
    definitions: Optional[List[str]] = None,
    warn: Optional[str] = None,
    number_formats_by_key: Optional[Dict[str, str]] = None,
    return_next_row: bool = False,
) -> None:
    """KV 式 sheet: 标题 + 来源 + 定义 + 多段( section_header + kv rows )。

    sections = [(section_name, [(key, value), ...]), ...]
    若 return_next_row=True, 返回写完后的下一行号 (供调用方继续追加内容)。
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    n_cols = 2  # 字段 / 值
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    cur_row = 2
    if source_note:
        ws.cell(row=cur_row, column=1, value=f"数据来源: {source_note}")
        ws.cell(row=cur_row, column=1).font = SRC_FONT
        ws.cell(row=cur_row, column=1).fill = SRC_FILL
        ws.cell(row=cur_row, column=1).alignment = SRC_ALIGN
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
        ws.row_dimensions[cur_row].height = 28
        cur_row += 1

    if definitions:
        for line in definitions:
            ws.cell(row=cur_row, column=1, value=line)
            ws.cell(row=cur_row, column=1).font = DEF_FONT
            ws.cell(row=cur_row, column=1).fill = DEF_FILL
            ws.cell(row=cur_row, column=1).alignment = DEF_ALIGN
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
            cur_row += 1

    # 表头
    for c, h in enumerate(["字段", "值"], 1):
        cell = ws.cell(row=cur_row, column=c, value=h)
        cell.font = HF
        cell.fill = HFILL
        cell.alignment = CENTER
        cell.border = BD
    cur_row += 1

    for sec_name, kv_list in sections:
        # section 标签行 (灰底蓝字合并)
        ws.cell(row=cur_row, column=1, value=sec_name)
        ws.cell(row=cur_row, column=1).font = SECTION_FONT
        ws.cell(row=cur_row, column=1).fill = SECTION_FILL
        ws.cell(row=cur_row, column=1).alignment = CENTER
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
        for c in range(1, n_cols + 1):
            ws.cell(row=cur_row, column=c).fill = SECTION_FILL
            ws.cell(row=cur_row, column=c).border = BD
        cur_row += 1

        for key, val in kv_list:
            kc = ws.cell(row=cur_row, column=1, value=key)
            kc.font = CF
            kc.alignment = LEFT
            kc.border = BD
            vc = ws.cell(row=cur_row, column=2, value=val)
            vc.font = CF
            vc.border = BD
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                vc.alignment = RIGHT
                if number_formats_by_key and key in number_formats_by_key:
                    vc.number_format = number_formats_by_key[key]
            else:
                vc.alignment = LEFT
            cur_row += 1

    if warn:
        ws.cell(row=cur_row, column=1, value=f"⚠ {warn}").font = WARN_FONT
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
        cur_row += 1

    # 列宽: 字段列 30, 值列 80
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80
    ws.freeze_panes = "A2"

    if return_next_row:
        return cur_row
    return None


# ===== Sheet 1: 模型概览 (from task-spec/_manifest.json) =====

def build_sheet1_overview(wb: Workbook, session_dir: Path) -> None:
    sheet_name = "1-模型概览"
    title = "1. 模型概览 (需求规格 + 路由 + 数据源)"

    ts_manifest = _read_json(session_dir / "task-spec" / "_manifest.json")

    if not ts_manifest:
        _write_table_sheet(
            wb, sheet_name, title, [], [],
            warn="task-spec/_manifest.json 缺失, 无法生成模型概览",
            source_note=f"{session_dir}/task-spec/_manifest.json (缺失)",
        )
        return

    routing = ts_manifest.get("routing", {}) or {}
    routing_basis = routing.get("routing_basis", {}) or {}
    req = ts_manifest.get("requirements", {}) or {}
    who = req.get("who", {}) or {}
    what = req.get("what", {}) or {}
    how_good = req.get("how_good", {}) or {}
    constraints = req.get("constraints", {}) or {}
    how_to_use = req.get("how_to_use", {}) or {}

    sections: List[Tuple[str, List[Tuple[str, Any]]]] = [
        ("▌ 基础信息", [
            ("模型简称 (model_name)", ts_manifest.get("model_name", "—")),
            ("session 时间戳", ts_manifest.get("timestamp", "—")),
            ("产出方", ts_manifest.get("produced_by", "—")),
        ]),
        ("▌ 路由溯源 (来自 model-task-routing)", [
            ("task_type", routing.get("task_type", "—")),
            ("Q1 预测目标", routing_basis.get("q1_target", "—")),
            ("Q2 是否有干预动作", routing_basis.get("q2_intervention", "—")),
            ("Q3 是否有实验/对照数据", routing_basis.get("q3_experiment_data", "—")),
            ("用户原始诉求", routing.get("user_raw_request", "—")),
            ("数据源 (routing 阶段)", routing.get("data_source", "—")),
            ("路由时间", routing.get("routed_at", "—")),
        ]),
        ("▌ 需求规格 — WHO (给谁打分)", [
            ("客群 (cohort)", who.get("cohort", "—")),
            ("筛选条件 (filter)", who.get("filter", "—")),
        ]),
        ("▌ 需求规格 — WHAT (预测什么)", [
            ("预测目标 (target)", what.get("target", "—")),
            ("任务类型", what.get("task_type", "—")),
            ("标签列", what.get("label_col", "—")),
            ("时间窗口 (window)", what.get("window", "—")),
        ]),
        ("▌ 需求规格 — HOW GOOD (效果目标)", [
            ("AUC 目标", how_good.get("auc_target", "—")),
            ("基线 (baseline)", how_good.get("baseline", "—")),
        ]),
        ("▌ 需求规格 — CONSTRAINTS (约束)", [
            ("最大特征数", constraints.get("max_features", "—")),
            ("分箱均匀性要求", constraints.get("binning_uniformity", "—")),
        ]),
        ("▌ 需求规格 — HOW TO USE (使用方式)", [
            ("使用模式", how_to_use.get("mode", "—")),
        ]),
        ("▌ 数据源", [
            ("样本表 (source_table)", ts_manifest.get("source_table", "—")),
            ("特征表 (feature_table)", ts_manifest.get("feature_table", "—")),
            ("join_keys", _to_str(ts_manifest.get("join_keys"))),
            ("id_cols", _to_str(ts_manifest.get("id_cols"))),
            ("时间列 (dt_col)", ts_manifest.get("dt_col", "—")),
            ("补充字段 (supplemental_features)", _to_str(ts_manifest.get("supplemental_features"))),
            ("样本配置 yaml", ts_manifest.get("sample_config_yaml", "—")),
        ]),
    ]

    _write_kv_sheet(
        wb, sheet_name, title, sections,
        source_note=f"{session_dir}/task-spec/_manifest.json",
    )


# ===== Sheet 2: 样本分析 (from data-profile/_manifest.json + _split_manifest.json) =====

def build_sheet2_sample(wb: Workbook, session_dir: Path) -> None:
    sheet_name = "2-样本分析"
    title = "2. 样本分析 (总体 + 分时段 + 三档切分 + 稳定性 + 充足性)"

    dp_manifest = _read_json(session_dir / "data-profile" / "_manifest.json")
    split_manifest = _read_json(session_dir / "data-profile" / "_split_manifest.json")

    if not dp_manifest:
        _write_table_sheet(
            wb, sheet_name, title, [], [],
            warn="data-profile/_manifest.json 缺失, 无法生成样本分析",
            source_note=f"{session_dir}/data-profile/_manifest.json (缺失)",
        )
        return

    sample_summary = dp_manifest.get("sample_summary", {}) or {}
    time_segments = dp_manifest.get("time_segments", []) or []
    stability = dp_manifest.get("stability", {}) or {}
    sufficiency = dp_manifest.get("sample_sufficiency", {}) or {}
    split = dp_manifest.get("split", {}) or {}
    ranges = split.get("ranges", {}) or {}
    actual_ratios = split.get("actual_ratios", {}) or {}
    splits = split.get("splits", {}) or {}
    null_counts = sample_summary.get("null_counts", {}) or {}

    # 实际列名: 优先从 task-spec 存档的 id_cols / dt_col 解析; 兜底按默认列名(fuid/f_p_date)
    ts_manifest = _read_json(session_dir / "task-spec" / "_manifest.json") or {}
    _ts_id_cols = ts_manifest.get("id_cols") or []
    id_col_label = _ts_id_cols[0] if isinstance(_ts_id_cols, (list, tuple)) and _ts_id_cols else (
        str(_ts_id_cols).split(",")[0].strip() if isinstance(_ts_id_cols, str) and _ts_id_cols else "fuid"
    )
    time_col_label = str(ts_manifest.get("dt_col") or "f_p_date")

    # 第一部分: KV 段 (总体样本 + 稳定性 + 充足性 + 切分元信息)
    sections: List[Tuple[str, List[Tuple[str, Any]]]] = [
        ("▌ 总体样本概况", [
            ("总样本量", _to_int(sample_summary.get("total_samples"))),
            ("正样本数", _to_int(sample_summary.get("positive_samples"))),
            ("负样本数", _to_int(sample_summary.get("negative_samples"))),
            ("正样本率", _to_float(sample_summary.get("positive_rate"))),
            ("正负比", sample_summary.get("positive_negative_ratio", "—")),
            (f"{time_col_label} 范围", _to_str(sample_summary.get("pday_range"))),
            (f"{time_col_label} 唯一数", _to_int(sample_summary.get("pday_unique_count"))),
            (f"{time_col_label} 值列表", _to_str(sample_summary.get("pday_values"))),
            ("用户唯一数", _to_int(sample_summary.get("user_unique"))),
            (f"重复 {id_col_label}+{time_col_label}", _to_int(sample_summary.get("dup_user_pday"))),
            (f"{id_col_label} 缺失数", _to_int(null_counts.get(id_col_label))),
            ("label 缺失数", _to_int(null_counts.get("label"))),
            (f"{time_col_label} 缺失数", _to_int(null_counts.get(time_col_label))),
        ]),
        ("▌ 稳定性 (跨 pday 正样本率)", [
            ("最高正样本率", _to_float(stability.get("positive_rate_max"))),
            ("最低正样本率", _to_float(stability.get("positive_rate_min"))),
            ("波动幅度 (pp)", _to_float(stability.get("volatility_pp"))),
            ("标准差 (pp)", _to_float(stability.get("std_pp"))),
            ("稳定性判定", stability.get("judgment", "—")),
            ("说明", stability.get("note", "—")),
        ]),
        ("▌ 样本充足性", [
            ("判定", sufficiency.get("judgment", "—")),
            ("正样本 ≥ 10k?", sufficiency.get("positive_meets_10k", "—")),
            ("正样本数", _to_int(sufficiency.get("positive_count"))),
            ("总样本 ≥ 50k?", sufficiency.get("total_meets_50k", "—")),
            ("总样本数", _to_int(sufficiency.get("total_count"))),
        ]),
        ("▌ 切分元信息", [
            ("切分方法", split.get("method", "—")),
            ("train pday 范围", _to_str(ranges.get("train"))),
            ("test pday 范围", _to_str(ranges.get("test"))),
            ("oot pday 范围", _to_str(ranges.get("oot"))),
            ("train 实际占比", _to_float(actual_ratios.get("train"))),
            ("test 实际占比", _to_float(actual_ratios.get("test"))),
            ("oot 实际占比", _to_float(actual_ratios.get("oot"))),
            ("跨切分正样本率差异 (pp)", _to_float(split.get("cross_split_pos_rate_diff_pp"))),
            ("dropped_rows", _to_int(split.get("dropped_rows"))),
            ("切分说明", split.get("note", "—")),
            ("用户已确认", dp_manifest.get("user_confirmed", "—")),
        ]),
    ]

    number_formats = {
        "正样本率": "0.00%",
        "最高正样本率": "0.00%",
        "最低正样本率": "0.00%",
        "波动幅度 (pp)": "0.00",
        "标准差 (pp)": "0.00",
        "train 实际占比": "0.00%",
        "test 实际占比": "0.00%",
        "oot 实际占比": "0.00%",
        "跨切分正样本率差异 (pp)": "0.00",
    }

    src = f"{session_dir}/data-profile/_manifest.json + {session_dir}/data-profile/_split_manifest.json"

    next_row = _write_kv_sheet(
        wb, sheet_name, title, sections,
        source_note=src,
        number_formats_by_key=number_formats,
        return_next_row=True,
    )
    if next_row is None:
        next_row = 2

    # 在 KV sheet 末尾追加两个表格: 分时段 + 三档切分
    ws = wb[sheet_name]
    cur_row = next_row + 1  # KV 写完后再空 1 行
    n_cols = 5
    # section 标签行
    ws.cell(row=cur_row, column=1, value=f"▌ 分时段样本 (按 {time_col_label} 分组)").font = SECTION_FONT
    ws.cell(row=cur_row, column=1).fill = SECTION_FILL
    ws.cell(row=cur_row, column=1).alignment = CENTER
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
    for c in range(1, n_cols + 1):
        ws.cell(row=cur_row, column=c).fill = SECTION_FILL
        ws.cell(row=cur_row, column=c).border = BD
    cur_row += 1

    # 表头
    seg_headers = [time_col_label, "样本量", "正样本数", "正样本率", "正负比"]
    for c, h in enumerate(seg_headers, 1):
        cell = ws.cell(row=cur_row, column=c, value=h)
        cell.font = HF
        cell.fill = HFILL
        cell.alignment = CENTER
        cell.border = BD
    cur_row += 1

    # 数据行
    data_start = cur_row
    for seg in time_segments:
        row_vals = [
            seg.get("pday", "—"),
            _to_int(seg.get("samples")),
            _to_int(seg.get("positive")),
            _to_float(seg.get("positive_rate")),
            seg.get("pos_neg_ratio", "—"),
        ]
        for c, v in enumerate(row_vals, 1):
            cell = ws.cell(row=cur_row, column=c, value=v)
            cell.font = CF
            cell.border = BD
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                cell.alignment = RIGHT
                if c == 2 or c == 3:
                    cell.number_format = "#,##0"
                elif c == 4:
                    cell.number_format = "0.00%"
            else:
                cell.alignment = LEFT
        cur_row += 1
    data_end = cur_row - 1

    cur_row += 1  # 空行

    # ---- 表 2: 三档切分 ----
    # section 标签行
    ws.cell(row=cur_row, column=1, value="▌ Train / Eval / OOT 切分").font = SECTION_FONT
    ws.cell(row=cur_row, column=1).fill = SECTION_FILL
    ws.cell(row=cur_row, column=1).alignment = CENTER
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
    for c in range(1, n_cols + 1):
        ws.cell(row=cur_row, column=c).fill = SECTION_FILL
        ws.cell(row=cur_row, column=c).border = BD
    cur_row += 1

    # 表头
    split_headers = ["split", "样本量", "正样本数", "正样本率", "pday 范围"]
    for c, h in enumerate(split_headers, 1):
        cell = ws.cell(row=cur_row, column=c, value=h)
        cell.font = HF
        cell.fill = HFILL
        cell.alignment = CENTER
        cell.border = BD
    cur_row += 1

    # 数据行
    data_start = cur_row
    for sp_name in ("train", "test", "oot"):
        sp_data = splits.get(sp_name, {}) or {}
        prange = sp_data.get("pday_range")
        row_vals = [
            sp_name,
            _to_int(sp_data.get("rows")),
            _to_int(sp_data.get("positive")),
            _to_float(sp_data.get("positive_rate")),
            _to_str(prange),
        ]
        for c, v in enumerate(row_vals, 1):
            cell = ws.cell(row=cur_row, column=c, value=v)
            cell.font = CF
            cell.border = BD
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                cell.alignment = RIGHT
                if c == 2 or c == 3:
                    cell.number_format = "#,##0"
                elif c == 4:
                    cell.number_format = "0.00%"
            else:
                cell.alignment = LEFT
        cur_row += 1
    data_end = cur_row - 1

    # 列宽调整 (覆盖 KV sheet 的 30/80, 改为适配表格)
    # 实际上 column_dimensions 是 sheet 级, 不能分区域设; 取折中:
    # A 列 30, B-E 列各 15-20
    ws.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 18


# ===== Sheet 3: 特征质量 (全量特征, from feature-analysis CSVs) =====

def build_sheet3_quality(wb: Workbook, session_dir: Path) -> None:
    sheet_name = "3-特征质量"
    title = "3. 特征质量 (全量特征 — IV / 单变量 AUC / PSI / 缺失率 / 基础统计)"

    fa_dir = session_dir / "sample-features" / "feature-analysis" / "analysis"
    fa_manifest = _read_json(fa_dir / "_manifest.json")
    iv_rows = _read_csv_rows(fa_dir / "iv_table.csv")
    psi_rows = _read_csv_rows(fa_dir / "psi_table.csv")
    stats_rows = _read_csv_rows(fa_dir / "stats.csv")

    if not fa_manifest:
        _write_table_sheet(
            wb, sheet_name, title, [], [],
            warn="feature-analysis/analysis/_manifest.json 缺失, 无法生成特征质量",
            source_note=f"{fa_dir}/_manifest.json (缺失)",
        )
        return

    src = (
        f"{fa_dir}/iv_table.csv (IV / 单变量 AUC) + "
        f"{fa_dir}/psi_table.csv (PSI) + "
        f"{fa_dir}/stats.csv (缺失率 / 基础统计)"
    )

    # ---- 全量特征质量表 (无 KV 概况段, 样本分析已在 Sheet 2) ----
    # 合并 3 个 CSV: iv_table (feature, iv, auc, n_bins_effective)
    #             + psi_table (feature, psi, warn)
    #             + stats (feature, dtype, count, missing_rate, unique, mean, std, min, q25, median, q75, max)
    iv_map = {r.get("feature"): r for r in iv_rows if r.get("feature")}
    psi_map = {r.get("feature"): r for r in psi_rows if r.get("feature")}
    stats_map = {r.get("feature"): r for r in stats_rows if r.get("feature")}

    # 取并集, 以 stats 为主(包含所有特征), 顺序按 IV 降序 (缺失 IV 排末尾)
    all_features: List[str] = []
    seen = set()
    for r in stats_rows:
        f = r.get("feature")
        if f and f not in seen:
            all_features.append(f)
            seen.add(f)
    for r in iv_rows:
        f = r.get("feature")
        if f and f not in seen:
            all_features.append(f)
            seen.add(f)
    for r in psi_rows:
        f = r.get("feature")
        if f and f not in seen:
            all_features.append(f)
            seen.add(f)

    # 按 IV 降序排序 (缺失 IV 的特征排末尾, 保持原 union 顺序)
    all_features.sort(
        key=lambda f: (
            _to_float(iv_map.get(f, {}).get("iv")) is None,
            -(_to_float(iv_map.get(f, {}).get("iv")) or 0.0),
        )
    )

    # 表头: # / feature / dtype / IV / 单变量 AUC / PSI / PSI 预警 / 有效分箱 / 缺失率 / unique / mean / std / min / q25 / median / q75 / max
    headers = [
        "#", "feature", "dtype", "IV", "单变量 AUC", "PSI", "PSI 预警", "有效分箱",
        "缺失率", "unique", "mean", "std", "min", "q25", "median", "q75", "max",
    ]
    number_formats = {
        4: "0.0000", 5: "0.0000", 6: "0.0000", 8: "0", 9: "0.00%", 10: "#,##0",
        11: "0.000000", 12: "0.000000", 13: "0.000000", 14: "0.000000",
        15: "0.000000", 16: "0.000000", 17: "0.000000",
    }

    table_rows: List[List[Any]] = []
    for i, f in enumerate(all_features, 1):
        iv_r = iv_map.get(f, {})
        psi_r = psi_map.get(f, {})
        st_r = stats_map.get(f, {})
        row = [
            i,
            f,
            st_r.get("dtype", "—"),
            _to_float(iv_r.get("iv")) if iv_r.get("iv") not in (None, "", "nan") else _to_float(None),
            _to_float(iv_r.get("auc")) if iv_r.get("auc") not in (None, "", "nan") else _to_float(None),
            _to_float(psi_r.get("psi")) if psi_r.get("psi") not in (None, "", "nan") else _to_float(None),
            psi_r.get("warn", "—"),
            _to_int(iv_r.get("n_bins_effective")),
            _to_float(st_r.get("missing_rate")) if st_r.get("missing_rate") not in (None, "", "nan") else _to_float(None),
            _to_int(st_r.get("unique")),
            _to_float(st_r.get("mean")) if st_r.get("mean") not in (None, "", "nan") else _to_float(None),
            _to_float(st_r.get("std")) if st_r.get("std") not in (None, "", "nan") else _to_float(None),
            _to_float(st_r.get("min")) if st_r.get("min") not in (None, "", "nan") else _to_float(None),
            _to_float(st_r.get("q25")) if st_r.get("q25") not in (None, "", "nan") else _to_float(None),
            _to_float(st_r.get("median")) if st_r.get("median") not in (None, "", "nan") else _to_float(None),
            _to_float(st_r.get("q75")) if st_r.get("q75") not in (None, "", "nan") else _to_float(None),
            _to_float(st_r.get("max")) if st_r.get("max") not in (None, "", "nan") else _to_float(None),
        ]
        table_rows.append(row)

    # 直接用 _write_table_sheet 写全量特征质量表 (无 KV, 无 definitions)
    _write_table_sheet(
        wb, sheet_name, title, headers, table_rows,
        source_note=src,
        number_formats=number_formats,
    )

    # DataBar: IV(col 4, 蓝) / PSI(col 6, 红) / 缺失率(col 9, 红)
    ws = wb[sheet_name]
    data_start = 2
    data_end = ws.max_row
    if data_end >= data_start:
        _blue_scale(ws, 4, data_start, data_end)
        _red_scale(ws, 6, data_start, data_end)
        _red_scale(ws, 9, data_start, data_end)

    # 列宽
    col_widths = {
        "A": 6, "B": 60, "C": 10, "D": 12, "E": 12, "F": 12, "G": 10, "H": 10,
        "I": 12, "J": 10, "K": 16, "L": 16, "M": 14, "N": 14, "O": 14, "P": 14, "Q": 14,
    }
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w
    ws.freeze_panes = "C2"


# ===== Sheet 4: 三档评估 (多 run × 8 指标, 按 split 分块) =====

def build_sheet4_metrics(
    wb: Workbook,
    runs_data: List[Dict[str, Any]],
    skipped: List[Tuple[str, str]],
    session_dir: Path,
) -> None:
    sheet_name = "4-三档评估"
    title = "4. 三档评估 (多 run × 8 指标, 按 split 分块 — train / test / oot / all)"

    headers = ["run_name", "样本量", "正样本率", "AUC", "KS", "准确率", "精确率", "召回率", "F1"]
    number_formats = {
        2: "#,##0", 3: "0.00%",
        4: "0.0000", 5: "0.0000", 6: "0.0000", 7: "0.0000", 8: "0.0000", 9: "0.0000",
    }

    definitions = [
        "布局说明:",
        "  · 本 sheet 按 split 分 4 个子表 (train / test / oot / all), 纵向堆叠, 以 split 标签行(灰底)分隔。",
        "  · 每个子表内: 一行一 run, 列为 8 个评估指标 (count/label_rate/auc/ks/accuracy/precision/recall/f1)。",
        "  · all = train+test+oot 合并(全集)评估, 指标口径与单 split 一致。",
        "",
        "指标定义与计算口径:",
        "  · 样本量 (count): 该 split 的样本总数。",
        "  · 正样本率 (label_rate): 正样本数 / 样本量; 各 run 应相同(同源数据)。",
        "  · AUC: ROC 曲线下面积, 范围 [0, 1], 越大越好。",
        "  · KS: 正负样本累积分布最大差值, 范围 [0, 1], 越大越好。",
        "  · 准确率 (accuracy): (TP+TN)/(TP+TN+FP+FN); 不平衡时可能虚高。",
        "  · 精确率 (precision): TP/(TP+FP), 预测为正中实际为正的比例。",
        "  · 召回率 (recall): TP/(TP+FN), 实际为正中被预测为正的比例。",
        "  · F1: 精确率与召回率的调和平均。",
    ]

    splits = ["train", "test", "oot", "all"]
    split_blocks: Dict[str, List[Tuple[Dict[str, Any], Optional[dict]]]] = {}
    missing_notes: List[str] = []
    for sp in splits:
        block: List[Tuple[Dict[str, Any], Optional[dict]]] = []
        for r in runs_data:
            ev = r["evals"].get(sp)
            if not ev:
                missing_notes.append(f"{r['run_name']} 缺 {sp}_eval.json")
                block.append((r, None))
                continue
            seg_map = ev.get("metric_by_segment", {}) or {}
            seg_metrics = seg_map.get("全量", {})
            if not seg_metrics:
                missing_notes.append(f"{r['run_name']} 的 {sp}_eval.json 无 metric_by_segment.全量")
                block.append((r, None))
                continue
            block.append((r, seg_metrics))
        split_blocks[sp] = block

    warn = None
    if missing_notes:
        warn = f"以下 run/split 的评估指标缺失, 在对应子表内留空: {', '.join(missing_notes)}"

    # 直接构造 sheet (本 sheet 是多子表堆叠, 布局特殊, 不走 _write_table_sheet)
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    n_cols = len(headers)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    cur_row = 2
    # 行 2: 数据来源
    src_text = (
        f"{session_dir}/new-models/<run_name>/evaluation/<run_name>_<split>_eval.json "
        "→ metric_by_segment['全量']"
    )
    ws.cell(row=cur_row, column=1, value=f"数据来源: {src_text}")
    ws.cell(row=cur_row, column=1).font = SRC_FONT
    ws.cell(row=cur_row, column=1).fill = SRC_FILL
    ws.cell(row=cur_row, column=1).alignment = SRC_ALIGN
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
    ws.row_dimensions[cur_row].height = 28
    cur_row += 1

    # 行 3+: 指标定义
    for line in definitions:
        ws.cell(row=cur_row, column=1, value=line)
        ws.cell(row=cur_row, column=1).font = DEF_FONT
        ws.cell(row=cur_row, column=1).fill = DEF_FILL
        ws.cell(row=cur_row, column=1).alignment = DEF_ALIGN
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
        cur_row += 1

    # 3 个 split 子表纵向堆叠
    for sp in splits:
        block = split_blocks[sp]
        # split 标签行
        ws.cell(row=cur_row, column=1, value=f"▼ {sp} 档 (共 {len(block)} 个 run)")
        ws.cell(row=cur_row, column=1).font = SPLIT_FONT
        ws.cell(row=cur_row, column=1).fill = SPLIT_FILL
        ws.cell(row=cur_row, column=1).alignment = CENTER
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
        for c in range(1, n_cols + 1):
            ws.cell(row=cur_row, column=c).fill = SPLIT_FILL
            ws.cell(row=cur_row, column=c).border = BD
        cur_row += 1

        # 表头
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=cur_row, column=c, value=h)
            cell.font = HF
            cell.fill = HFILL
            cell.alignment = CENTER
            cell.border = BD
        cur_row += 1

        # 数据行
        data_start = cur_row
        for r, seg_metrics in block:
            if seg_metrics is None:
                ws.cell(row=cur_row, column=1, value=r["run_name"]).font = CF
                ws.cell(row=cur_row, column=1).border = BD
                ws.cell(row=cur_row, column=1).alignment = LEFT
                for c in range(2, n_cols + 1):
                    ws.cell(row=cur_row, column=c, value="—").font = WARN_FONT
                    ws.cell(row=cur_row, column=c).border = BD
                    ws.cell(row=cur_row, column=c).alignment = RIGHT
            else:
                row_vals = [
                    r["run_name"],
                    _to_int(seg_metrics.get("count")),
                    _to_float(seg_metrics.get("label_rate")),
                    _to_float(seg_metrics.get("auc")),
                    _to_float(seg_metrics.get("ks")),
                    _to_float(seg_metrics.get("accuracy")),
                    _to_float(seg_metrics.get("precision")),
                    _to_float(seg_metrics.get("recall")),
                    _to_float(seg_metrics.get("f1")),
                ]
                for c, v in enumerate(row_vals, 1):
                    cell = ws.cell(row=cur_row, column=c, value=v)
                    cell.font = CF
                    cell.border = BD
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        cell.alignment = RIGHT
                        if c in number_formats:
                            cell.number_format = number_formats[c]
                    else:
                        cell.alignment = LEFT
            cur_row += 1
        data_end = cur_row - 1
        # Sheet 4 不画 DataBar (用户要求: 仅看指标数值, 不用条形可视化)

        cur_row += 1  # 子表之间空一行

    # 列宽
    for c, h in enumerate(headers, 1):
        col_letter = get_column_letter(c)
        max_len = len(str(h))
        for row_idx in range(2, cur_row):
            v = ws.cell(row=row_idx, column=c).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 50)

    ws.freeze_panes = "A2"

    if warn:
        ws.cell(row=cur_row, column=1, value=f"⚠ {warn}").font = WARN_FONT
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)

    if skipped:
        skipped_str = ", ".join(f"{n}({reason})" for n, reason in skipped)
        cur_row += 1
        ws.cell(row=cur_row, column=1, value=f"⚠ 被跳过的 run: {skipped_str}").font = WARN_FONT
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)


# ===== Sheet 5: 分桶排序性对比 (严格参考 classification-model-comparison Sheet 2 格式) =====

# 灰底表头样式 (与 compare_models.py style_header_red 一致)
GRAY_FILL = PatternFill("solid", fgColor="D9D9D9")
GRAY_FONT = Font(name="微软雅黑", bold=True, size=11, color="333333")
# 灰底分组标题样式 (metric group header)
GRP_GRAY_FONT = Font(name="微软雅黑", bold=True, size=10, color="1A3060")
# Lift 子表数值颜色 (>1 绿, <1 红, =1 / 基线 灰)
LIFT_GREEN_FONT = Font(name="微软雅黑", size=10, color="006100")
LIFT_RED_FONT = Font(name="微软雅黑", size=10, color="9C0006")
LIFT_GRAY_FONT = Font(name="微软雅黑", size=10, color="333333")


def build_sheet5_buckets(
    wb: Workbook,
    runs_data: List[Dict[str, Any]],
    skipped: List[Tuple[str, str]],
    session_dir: Path,
) -> None:
    sheet_name = "5-分桶排序性对比"
    title = "5. 分桶排序性对比 (多 run 并排, 仅 oot + all — 同 comparison Sheet 2 格式)"

    # 收集每个 run 的 buckets (per split, 仅 oot + all)
    runs_with_buckets: List[Tuple[Dict[str, Any], Dict[str, List[dict]]]] = []
    for r in runs_data:
        buckets_by_split: Dict[str, List[dict]] = {}
        for sp in ("oot", "all"):
            ev = r["evals"].get(sp)
            if ev:
                sb = ((ev.get("performance") or {}).get("score_buckets") or {}).get("全量", [])
                if sb:
                    buckets_by_split[sp] = sb
        if buckets_by_split:
            runs_with_buckets.append((r, buckets_by_split))

    if not runs_with_buckets:
        _write_table_sheet(
            wb, sheet_name, title, [], [],
            warn="所有 run 均无 performance.score_buckets['全量'] 数据, 无法生成分桶对比",
            source_note=f"{session_dir}/new-models/<run_name>/evaluation/<run_name>_<split>_eval.json (缺失 buckets)",
        )
        return

    n_runs = len(runs_with_buckets)
    n_cols = 2 + 3 * n_runs  # 分桶 + 人数 + (label率, 召回率, 累计召回) per run

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    cur_row = 2
    # 数据来源
    src_text = (
        f"{session_dir}/new-models/<run_name>/evaluation/<run_name>_<split>_eval.json "
        "→ performance.score_buckets['全量']"
    )
    ws.cell(row=cur_row, column=1, value=f"数据来源: {src_text}")
    ws.cell(row=cur_row, column=1).font = SRC_FONT
    ws.cell(row=cur_row, column=1).fill = SRC_FILL
    ws.cell(row=cur_row, column=1).alignment = SRC_ALIGN
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
    ws.row_dimensions[cur_row].height = 28
    cur_row += 1

    # 指标计算逻辑 (灰底小字, 同 comparison Sheet 2 R1)
    ws.cell(row=cur_row, column=1,
            value="指标计算逻辑: 分桶=各模型按自己打分降序十分位(Decile10=最高分) | label率=桶内正样本占比 | 召回率=桶正样本/总正样本 | 累计召回=从高到低累计捕获正样本占比")
    ws.cell(row=cur_row, column=1).font = Font(name="微软雅黑", size=9, color="888888")
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
    cur_row += 1

    # 基线版本声明 (灰底小字, 同 comparison Sheet 2 R2)
    # 基线 = runs_with_buckets 的第 1 个 run (与 comparison skill 默认取 names[0] 一致)
    baseline_run_name = runs_with_buckets[0][0]["run_name"]
    ws.cell(row=cur_row, column=1,
            value=f"基线版本={baseline_run_name} | Lift=各版本指标 / 基线同指标 (>1 表示优于基线)")
    ws.cell(row=cur_row, column=1).font = Font(name="微软雅黑", size=9, color="888888")
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
    cur_row += 1

    # 每个 split 一个子表 (主表 + Lift 子表), 仅 oot + all
    for sp in ("oot", "all"):
        # 该 split 下各 run 的 buckets (可能 None); 记 (run_name, buckets_or_None)
        run_buckets_list: List[Tuple[str, Optional[List[dict]]]] = []
        for r, bbs in runs_with_buckets:
            run_buckets_list.append((r["run_name"], bbs.get(sp)))

        # 若所有 run 都缺该 split, 跳过
        if all(b is None for _, b in run_buckets_list):
            continue

        # split 标签行
        ws.cell(row=cur_row, column=1, value=f"▼ {sp} 档 (共 {len(run_buckets_list)} 个 run)")
        ws.cell(row=cur_row, column=1).font = SPLIT_FONT
        ws.cell(row=cur_row, column=1).fill = SPLIT_FILL
        ws.cell(row=cur_row, column=1).alignment = CENTER
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
        for c in range(1, n_cols + 1):
            ws.cell(row=cur_row, column=c).fill = SPLIT_FILL
            ws.cell(row=cur_row, column=c).border = BD
        cur_row += 1

        # === 主表: metric group headers (灰底合并, per-metric 分组) ===
        # 列顺序: 分桶(1) 人数(2) | label率×n_runs | 召回率×n_runs | 累计召回×n_runs
        # 前 2 列 (分桶/人数) 留空, 加边框
        ws.cell(row=cur_row, column=1, value=None).border = BD
        ws.cell(row=cur_row, column=2, value=None).border = BD
        ws.cell(row=cur_row, column=1).fill = GRAY_FILL
        ws.cell(row=cur_row, column=2).fill = GRAY_FILL
        col = 3
        for grp_label in ("label率(正样本率)", "召回率", "累计召回"):
            grp_end = col + n_runs - 1
            cell = ws.cell(row=cur_row, column=col, value=grp_label)
            cell.font = GRP_GRAY_FONT
            cell.fill = GRAY_FILL
            cell.alignment = CENTER
            ws.merge_cells(start_row=cur_row, start_column=col, end_row=cur_row, end_column=grp_end)
            for c in range(col, grp_end + 1):
                ws.cell(row=cur_row, column=c).fill = GRAY_FILL
                ws.cell(row=cur_row, column=c).border = BD
            col = grp_end + 1
        cur_row += 1

        # === 模型名子表头 (灰底, 每组下重复 n_runs 个模型名, 基线加 "(基线)" 后缀) ===
        sub_headers = ["分桶", "人数"]
        # label率 group
        for run_name, _ in run_buckets_list:
            sub_headers.append(f"{run_name} (基线)" if run_name == baseline_run_name else run_name)
        # 召回率 group
        for run_name, _ in run_buckets_list:
            sub_headers.append(f"{run_name} (基线)" if run_name == baseline_run_name else run_name)
        # 累计召回 group
        for run_name, _ in run_buckets_list:
            sub_headers.append(f"{run_name} (基线)" if run_name == baseline_run_name else run_name)
        for c, h in enumerate(sub_headers, 1):
            cell = ws.cell(row=cur_row, column=c, value=h)
            cell.font = GRAY_FONT
            cell.fill = GRAY_FILL
            cell.alignment = CENTER
            cell.border = BD
        cur_row += 1

        # === 主表数据行: 10 deciles (10 → 1) ===
        # 用首个有 buckets 的 run 提供 decile / count (与 comparison 一致)
        first_with_buckets = next((b for _, b in run_buckets_list if b is not None), None) or []
        max_buckets = len(first_with_buckets) if first_with_buckets else 0
        if max_buckets == 0:
            max_buckets = 10

        data_start = cur_row
        for decile_idx in range(max_buckets):
            # decile + count from first run with buckets
            if decile_idx < len(first_with_buckets):
                b0 = first_with_buckets[decile_idx]
                decile_v = b0.get("decile")
                count_v = b0.get("count")
            else:
                decile_v = "—"
                count_v = "—"
            cell = ws.cell(row=cur_row, column=1, value=decile_v)
            cell.font = CF; cell.border = BD
            cell.alignment = CENTER
            cell2 = ws.cell(row=cur_row, column=2, value=count_v)
            cell2.font = CF; cell2.border = BD
            cell2.alignment = RIGHT
            if isinstance(count_v, (int, float)) and not isinstance(count_v, bool):
                cell2.number_format = "#,##0"

            col = 3
            # label率 group
            for _, buckets in run_buckets_list:
                if buckets is not None and decile_idx < len(buckets):
                    v = buckets[decile_idx].get("label_rate")
                else:
                    v = None
                cell = ws.cell(row=cur_row, column=col, value=v if v is not None else "—")
                cell.font = CF; cell.border = BD
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    cell.alignment = RIGHT
                    cell.number_format = "0.00%"
                else:
                    cell.alignment = CENTER
                col += 1
            # 召回率 group
            for _, buckets in run_buckets_list:
                if buckets is not None and decile_idx < len(buckets):
                    v = buckets[decile_idx].get("recall")
                else:
                    v = None
                cell = ws.cell(row=cur_row, column=col, value=v if v is not None else "—")
                cell.font = CF; cell.border = BD
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    cell.alignment = RIGHT
                    cell.number_format = "0.00%"
                else:
                    cell.alignment = CENTER
                col += 1
            # 累计召回 group
            for _, buckets in run_buckets_list:
                if buckets is not None and decile_idx < len(buckets):
                    v = buckets[decile_idx].get("cum_recall")
                else:
                    v = None
                cell = ws.cell(row=cur_row, column=col, value=v if v is not None else "—")
                cell.font = CF; cell.border = BD
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    cell.alignment = RIGHT
                    cell.number_format = "0.00%"
                else:
                    cell.alignment = CENTER
                col += 1
            cur_row += 1
        data_end = cur_row - 1

        # === DataBar: 主表 label率(绿) / 召回率(蓝) / 累计召回(蓝), per metric group, 各列独立 ===
        if data_end >= data_start:
            col = 3
            for _ in range(n_runs):  # label率 group
                _green_scale(ws, col, data_start, data_end)
                col += 1
            for _ in range(n_runs):  # 召回率 group
                _blue_scale(ws, col, data_start, data_end)
                col += 1
            for _ in range(n_runs):  # 累计召回 group
                _blue_scale(ws, col, data_start, data_end)
                col += 1

        # === Lift 子表 (仅当 n_runs > 1 且基线 split 有 buckets 时) ===
        baseline_buckets = next(
            (b for rn, b in run_buckets_list if rn == baseline_run_name and b is not None), None
        )
        if n_runs > 1 and baseline_buckets:
            cur_row += 1  # 主表与 Lift 子表之间空一行
            # Lift 子标题
            ws.cell(row=cur_row, column=1,
                    value=f"各版本 vs {baseline_run_name} Lift（label率 / 召回率 / 累计召回 相除）")
            ws.cell(row=cur_row, column=1).font = SUB_FONT
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)
            cur_row += 1

            # Lift metric group headers (灰底合并)
            ws.cell(row=cur_row, column=1, value=None).border = BD
            ws.cell(row=cur_row, column=2, value=None).border = BD
            ws.cell(row=cur_row, column=1).fill = GRAY_FILL
            ws.cell(row=cur_row, column=2).fill = GRAY_FILL
            col = 3
            for grp_label in ("label率 Lift", "召回率 Lift", "累计召回 Lift"):
                grp_end = col + n_runs - 1
                cell = ws.cell(row=cur_row, column=col, value=grp_label)
                cell.font = GRP_GRAY_FONT
                cell.fill = GRAY_FILL
                cell.alignment = CENTER
                ws.merge_cells(start_row=cur_row, start_column=col, end_row=cur_row, end_column=grp_end)
                for c in range(col, grp_end + 1):
                    ws.cell(row=cur_row, column=c).fill = GRAY_FILL
                    ws.cell(row=cur_row, column=c).border = BD
                col = grp_end + 1
            cur_row += 1

            # Lift 模型名子表头 (同主表)
            lift_sub_headers = ["分桶", "人数"]
            for _ in range(3):  # 3 metric groups
                for run_name, _ in run_buckets_list:
                    lift_sub_headers.append(f"{run_name} (基线)" if run_name == baseline_run_name else run_name)
            for c, h in enumerate(lift_sub_headers, 1):
                cell = ws.cell(row=cur_row, column=c, value=h)
                cell.font = GRAY_FONT
                cell.fill = GRAY_FILL
                cell.alignment = CENTER
                cell.border = BD
            cur_row += 1

            # Lift 数据行: 基线列 = "-", 其他 = ratio (绿>1 / 红<1 / 灰=1)
            for decile_idx in range(len(baseline_buckets)):
                b0 = baseline_buckets[decile_idx] if decile_idx < len(baseline_buckets) else {}
                decile_v = b0.get("decile", "—")
                count_v = b0.get("count", "—")
                cell = ws.cell(row=cur_row, column=1, value=decile_v)
                cell.font = CF; cell.border = BD; cell.alignment = CENTER
                cell2 = ws.cell(row=cur_row, column=2, value=count_v)
                cell2.font = CF; cell2.border = BD; cell2.alignment = RIGHT
                if isinstance(count_v, (int, float)) and not isinstance(count_v, bool):
                    cell2.number_format = "#,##0"

                col = 3
                for metric_key in ("label_rate", "recall", "cum_recall"):
                    base_v = b0.get(metric_key, 0) if isinstance(b0, dict) else 0
                    base_v = base_v if isinstance(base_v, (int, float)) and not isinstance(base_v, bool) else 0
                    for run_name, buckets in run_buckets_list:
                        if run_name == baseline_run_name:
                            cell = ws.cell(row=cur_row, column=col, value="-")
                            cell.font = LIFT_GRAY_FONT
                            cell.border = BD
                            cell.alignment = CENTER
                        else:
                            if buckets is not None and decile_idx < len(buckets):
                                v = buckets[decile_idx].get(metric_key, 0)
                                v = v if isinstance(v, (int, float)) and not isinstance(v, bool) else 0
                            else:
                                v = 0
                            dv = v / base_v if base_v > 0 else 0
                            cell = ws.cell(row=cur_row, column=col, value=dv)
                            if dv > 1:
                                cell.font = LIFT_GREEN_FONT
                            elif dv < 1:
                                cell.font = LIFT_RED_FONT
                            else:
                                cell.font = LIFT_GRAY_FONT
                            cell.border = BD
                            cell.alignment = RIGHT
                            cell.number_format = "0.00"
                        col += 1
                cur_row += 1
            # Lift 子表不加 DataBar (用户要求: Lift 仅看数值)

        cur_row += 1  # split 之间空一行

    # 列宽
    ws.column_dimensions["A"].width = 8   # 分桶
    ws.column_dimensions["B"].width = 10  # 人数
    # 后续列: 每 run 3 列, 但因 per-metric 分组, 同一 run 的 3 列被拆到不同 metric group
    # 统一设 13 列宽 (兼容模型名 + 数据)
    for c in range(3, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    ws.freeze_panes = "C2"

    if skipped:
        skipped_str = ", ".join(f"{n}({reason})" for n, reason in skipped)
        ws.cell(row=cur_row, column=1, value=f"⚠ 被跳过的 run: {skipped_str}").font = WARN_FONT
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=n_cols)


# ===== Sheet 6: 特征清单 (AUC 最高的 run) =====

def build_sheet6_features(
    wb: Workbook,
    runs_data: List[Dict[str, Any]],
    skipped: List[Tuple[str, str]],
    session_dir: Path,
) -> None:
    sheet_name = "6-特征清单"
    title = "6. 特征清单 (AUC 最高的 run — 特征重要性 + SHAP)"

    best = _find_best_run(runs_data)
    if best is None:
        _write_table_sheet(
            wb, sheet_name, title, [], [],
            warn="无有效 run (config.json.runtime.metrics 全缺), 无法选定 AUC 最高 run",
            source_note=f"{session_dir}/new-models/<run_name>/config.json (无有效 run)",
        )
        return

    run_dir = best["run_dir"]
    run_name = best["run_name"]
    cfg = best["cfg"]
    mm = best.get("model_manifest") or {}
    runtime = cfg.get("runtime", {}) or {}
    metrics = runtime.get("metrics", {}) or {}
    train_m = metrics.get("train", {}) or {}
    val_m = metrics.get("val", {}) or {}
    oot_m = metrics.get("oot", {}) or {}
    used_params = mm.get("used_params", {}) or {}

    # 决定 AUC 选型说明
    oot_auc = oot_m.get("auc")
    train_auc = train_m.get("auc")
    if oot_auc is not None:
        selection_basis = f"oot AUC = {float(oot_auc):.4f}"
    else:
        selection_basis = f"train AUC = {float(train_auc):.4f}" if train_auc is not None else "(无 AUC)"

    fi_csv = run_dir / "explainability" / "feature-importance.csv"
    shap_csv = run_dir / "explainability" / "shap-summary.csv"
    used_csv = run_dir / "features" / "used-feature-list.csv"

    fi_rows = _read_csv_rows(fi_csv)
    shap_rows = _read_csv_rows(shap_csv)
    used_rows = _read_csv_rows(used_csv)

    # ---- Section 1: 基本信息 (KV) ----
    sections: List[Tuple[str, List[Tuple[str, Any]]]] = [
        ("▌ 基本信息 (AUC 最高的 run)", [
            ("run_name", run_name),
            ("algo", cfg.get("algo", "—")),
            ("suffix", cfg.get("suffix", "—") or "—"),
            ("version", cfg.get("version", "—")),
            ("label", cfg.get("label", "—")),
            ("timestamp", cfg.get("timestamp", "—")),
            ("produced_by", cfg.get("produced_by", "—")),
            ("n_features (runtime)", _to_int(runtime.get("n_features"))),
            ("best_iteration", _to_int(runtime.get("best_iteration"))),
            ("选定依据", selection_basis),
        ]),
        ("▌ 三档 AUC / KS (runtime.metrics)", [
            ("train AUC", _to_float(train_m.get("auc"))),
            ("train KS", _to_float(train_m.get("ks"))),
            ("train Gini", _to_float(train_m.get("gini"))),
            ("test (val) AUC", _to_float(val_m.get("auc"))),
            ("test (val) KS", _to_float(val_m.get("ks"))),
            ("test (val) Gini", _to_float(val_m.get("gini"))),
            ("oot AUC", _to_float(oot_m.get("auc"))),
            ("oot KS", _to_float(oot_m.get("ks"))),
            ("oot Gini", _to_float(oot_m.get("gini"))),
        ]),
    ]

    # 训练参数 (used_params) — 只列标量
    if used_params:
        params_kv: List[Tuple[str, Any]] = []
        for k, v in used_params.items():
            if isinstance(v, (int, float, str, bool)) or v is None:
                params_kv.append((k, v if v is not None else "—"))
        if params_kv:
            sections.append(("▌ 训练参数 (model/_manifest.json used_params)", params_kv))

    sections.append(("▌ 入模特征清单统计", [
        ("used-feature-list.csv 行数", _to_int(len(used_rows))),
        ("feature-importance.csv 行数", _to_int(len(fi_rows))),
        ("shap-summary.csv 行数", _to_int(len(shap_rows))),
    ]))

    src = (
        f"{run_dir}/config.json (基本信息 + runtime.metrics) + "
        f"{run_dir}/model/_manifest.json (used_params) + "
        f"{run_dir}/explainability/feature-importance.csv (特征重要性) + "
        f"{run_dir}/explainability/shap-summary.csv (SHAP) + "
        f"{run_dir}/features/used-feature-list.csv (入模特征清单)"
    )

    number_formats = {
        "train AUC": "0.0000", "train KS": "0.0000", "train Gini": "0.0000",
        "test (val) AUC": "0.0000", "test (val) KS": "0.0000", "test (val) Gini": "0.0000",
        "oot AUC": "0.0000", "oot KS": "0.0000", "oot Gini": "0.0000",
    }

    next_row = _write_kv_sheet(
        wb, sheet_name, title, sections,
        source_note=src,
        number_formats_by_key=number_formats,
        return_next_row=True,
    )
    if next_row is None:
        next_row = 2

    # ---- 在 KV 末尾追加 1 张全量合并表 (特征重要性 + SHAP 同表) ----
    ws = wb[sheet_name]
    cur_row = next_row + 1

    # 特征重要性 (全量, 按 importance 降序)
    fi_sorted = sorted(
        [r for r in fi_rows if r.get("importance") and r["importance"] not in ("nan", "")],
        key=lambda r: float(r["importance"]),
        reverse=True,
    )
    # SHAP (全量, 按 mean_abs_shap 降序) — 用 feature 名做 map, 后面 join
    shap_map = {}
    for r in shap_rows:
        f = r.get("feature")
        if f and r.get("mean_abs_shap") and r["mean_abs_shap"] not in ("nan", ""):
            shap_map[f] = r
    # SHAP 中存在但 fi 缺失的特征, 追加在 fi_sorted 末尾 (按 mean_abs_shap 降序)
    shap_only = sorted(
        [r for r in shap_rows
         if r.get("feature") and r.get("feature") not in {x.get("feature") for x in fi_sorted}
         and r.get("mean_abs_shap") and r["mean_abs_shap"] not in ("nan", "")],
        key=lambda r: float(r["mean_abs_shap"]),
        reverse=True,
    )

    # 合并表头: # / feature / importance / mean_abs_shap / mean_shap
    merged_headers = ["#", "feature", "importance", "mean_abs_shap", "mean_shap"]
    merged_data = []
    for i, r in enumerate(fi_sorted, 1):
        f = r.get("feature", "—")
        sh = shap_map.get(f) or {}
        merged_data.append([
            i,
            f,
            _to_float(r.get("importance")),
            _to_float(sh.get("mean_abs_shap")) if sh else "—",
            _to_float(sh.get("mean_shap")) if sh else "—",
        ])
    for j, r in enumerate(shap_only, len(fi_sorted) + 1):
        merged_data.append([
            j,
            r.get("feature", "—"),
            "—",
            _to_float(r.get("mean_abs_shap")),
            _to_float(r.get("mean_shap")),
        ])

    # section 标签行
    ws.cell(row=cur_row, column=1,
            value="▌ 特征重要性 + SHAP (全量, 按 importance 降序; SHAP 列缺失写 —)").font = SECTION_FONT
    ws.cell(row=cur_row, column=1).fill = SECTION_FILL
    ws.cell(row=cur_row, column=1).alignment = CENTER
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)
    for c in range(1, 6):
        ws.cell(row=cur_row, column=c).fill = SECTION_FILL
        ws.cell(row=cur_row, column=c).border = BD
    cur_row += 1

    # 表头
    for c, h in enumerate(merged_headers, 1):
        cell = ws.cell(row=cur_row, column=c, value=h)
        cell.font = HF
        cell.fill = HFILL
        cell.alignment = CENTER
        cell.border = BD
    cur_row += 1

    # 数据
    data_start = cur_row
    for row in merged_data:
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=cur_row, column=c, value=v)
            cell.font = CF
            cell.border = BD
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                cell.alignment = RIGHT
                if c == 1:
                    cell.number_format = "0"
                elif c == 3:
                    cell.number_format = "0.000000"
                elif c in (4, 5):
                    cell.number_format = "0.000000"
            else:
                cell.alignment = LEFT if c == 2 else RIGHT
        cur_row += 1
    data_end = cur_row - 1
    # DataBar 只画在 importance 列 (col 3, 蓝)
    if data_end >= data_start:
        _blue_scale(ws, 3, data_start, data_end)

    # 列宽
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "A2"

    if skipped:
        skipped_str = ", ".join(f"{n}({reason})" for n, reason in skipped)
        cur_row += 1
        ws.cell(row=cur_row, column=1, value=f"⚠ 被跳过的 run: {skipped_str}").font = WARN_FONT
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)


# ===== Main =====

def main() -> None:
    parser = argparse.ArgumentParser(description="session 级 6-sheet 综合报告")
    parser.add_argument(
        "--session-dir", required=True, type=Path,
        help="session 根目录, 如 runs/20260701-110624-draw_willingness/",
    )
    parser.add_argument(
        "-o", "--output", type=Path, default=None,
        help="输出 xlsx 路径; 默认 <session_dir>/{session_name}_report.xlsx",
    )
    args = parser.parse_args()

    session_dir = args.session_dir.resolve()
    if not session_dir.is_dir():
        sys.exit(f"[build_report] session-dir 不存在: {session_dir}")

    runs_data, skipped = _discover_runs(session_dir)
    if not runs_data:
        detail = ""
        if skipped:
            detail = "\n  " + "\n  ".join(f"{n}: {reason}" for n, reason in skipped)
        sys.exit(
            f"[build_report] {session_dir}/new-models/ 下无有效 run "
            f"(需含 config.json + 至少一个 split 的 eval JSON){detail}"
        )

    output = args.output or (session_dir / f"{session_dir.name}_report.xlsx")
    output = output.resolve()

    print(f"[build_report] session_dir = {session_dir}")
    print(f"[build_report] runs        = {[r['run_name'] for r in runs_data]}")
    if skipped:
        print(f"[build_report] skipped     = {skipped}")
    print(f"[build_report] output      = {output}")

    wb = Workbook()
    wb.remove(wb.active)

    build_sheet1_overview(wb, session_dir)
    build_sheet2_sample(wb, session_dir)
    build_sheet3_quality(wb, session_dir)
    build_sheet4_metrics(wb, runs_data, skipped, session_dir)
    build_sheet5_buckets(wb, runs_data, skipped, session_dir)
    build_sheet6_features(wb, runs_data, skipped, session_dir)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    n_sheets = len(wb.sheetnames)
    print(f"[build_report] {n_sheets} sheets written to: {output}")
    print(f"[build_report] sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
