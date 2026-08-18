# -*- coding: utf-8 -*-
"""通用风控模型评估报告生成。

接收 config dict（见 build_default_config），从 CSV 读取打分数据，
生成含 6 个 sheet 的 Excel 评估报告。

用法:
    python generate_report.py config.json
    # 或在 Python 中:
    # from generate_report import main_with_config
    # main_with_config(config)

config 字段:
    csv_path        : str       # 输入 CSV
    template_path   : str       # Excel 模板路径（可选，None 则从零建）
    out_path        : str       # 输出 Excel 路径（缺省: CSV 同目录；CSV 在 scoring/ 子目录时落其父目录）
    model_name      : str       # 报告名（如 "V12mob4"）
    primary_label   : str       # 主标签列名（如 "dpd30_3c"）

    uid_col         : str       # 用户ID列（默认 "fuid"）
    date_col        : str       # 日期列（默认 "fser_date"）

    # 模型: [(名称, 评分列名), ...] —— 第一个为新模型，第二个为基线
    models          : list[tuple]

    # 客群过滤: {列名: 值}，多条件 AND，空字典则不过滤
    filter_conds    : dict      # 如 {"if_tf": 1}

    train_range     : (str, str) # 训练集月范围 如 ("2025-08", "2025-11")
    oot_range       : (str, str) # OOT 集月范围 如 ("2025-12", "2026-04")

    base_month      : str or None  # PSI 基准月，None 取第一个 OOT 月

    # 标签列
    ks_labels       : list[str]  # KS 表所有标签列
    lift_labels     : list[str]  # Lift/SWAP 标签（至少1个）
    seg_labels      : list[str]  # 分段逾期率标签

    lift_bins       : int       # Lift 等频桶数（默认10）
    swap_bins       : int       # SWAP 等频桶数（默认5）
    min_cnt         : int       # 有表现<此值标记 '-'（默认30）
"""
import warnings
warnings.filterwarnings("ignore")

import os
import sys
import math
import json
import argparse
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# metric.py 和 generate_report.py 同目录
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from metric import calc_ks, calc_psi


# ---------------------------------------------------------------- default config
def build_default_config():
    return {
        "csv_path": None,
        "template_path": None,
        "out_path": None,
        "model_name": "Model",
        "primary_label": None,
        "uid_col": "fuid",
        "date_col": "fser_date",
        "models": [],
        "filter_conds": {},
        "train_range": None,
        "oot_range": None,
        "base_month": None,
        "ks_labels": [],
        "lift_labels": [],
        "seg_labels": [],
        "lift_bins": 10,
        "swap_bins": 5,
        "min_cnt": 30,
    }


def infer_columns_and_ranges(csv_path):
    """扫描 CSV，单次读取推断列名、标签列、评分列、日期列、月份范围。

    只对 CSV 做一次截断读取：先取列名，再按列裁剪只读日期候选列（最多 N_PROBE_ROWS 行），
    从中推断月份范围与可用日期列。避免对每个日期候选列重复全量解析。
    """
    N_PROBE_ROWS = 500000  # 月份推断探测行数
    cols = pd.read_csv(csv_path, nrows=0).columns.tolist()

    # 标签列：fpd 或 dpd 开头
    label_cols = [c for c in cols if c.startswith(("fpd", "dpd"))]
    # 评分列：含 score 且不含 label
    score_cols = [c for c in cols if "score" in c.lower() and "label" not in c.lower()]
    # 日期候选列：含 date 或 ser
    date_candidates = [c for c in cols if "date" in c.lower() or "ser" in c.lower()]

    months, date_col, n_rows = [], "fser_date", 0
    if date_candidates:
        # 列裁剪只读日期候选列 + 标签列（一次读取，供月份推断）
        probe = pd.read_csv(
            csv_path, usecols=date_candidates, nrows=N_PROBE_ROWS,
            dtype={c: str for c in date_candidates})
        n_rows = len(probe)
        # 取第一个能解析出年月的候选列作为日期列
        for c in date_candidates:
            try:
                parsed = pd.to_datetime(probe[c], errors="coerce")
                m = sorted(parsed.dt.strftime("%Y-%m").dropna().unique().tolist())
            except Exception:
                continue
            if m:
                months, date_col = m, c
                break

    return {"all_cols": cols, "label_cols": label_cols, "score_cols": score_cols,
            "months": months, "date_col": date_col, "n_rows": n_rows}


def resolve_config(cfg, base_dir=None):
    """校验 + 补充默认值。返回最终 config dict。"""
    c = dict(build_default_config())
    c.update(cfg)

    if not c["csv_path"]:
        raise ValueError("csv_path is required")
    if not c["out_path"]:
        import datetime
        if base_dir:
            out_dir = base_dir
        else:
            csv_dir = os.path.dirname(os.path.abspath(c["csv_path"]))
            # 输入 CSV 位于 scoring/ 子目录时，报告默认落到 session 根（scoring 的父目录），
            # 避免业务评估报告混入 scoring/ 打分产物目录
            if os.path.basename(csv_dir) == "scoring":
                out_dir = os.path.dirname(csv_dir)
            else:
                out_dir = csv_dir
        c["out_path"] = os.path.join(out_dir,
            f"{c['model_name']}评估报告_{datetime.date.today().strftime('%Y%m%d')}.xlsx")

    info = infer_columns_and_ranges(c["csv_path"])
    all_cols = info["all_cols"]
    colset = set(all_cols)

    # 日期列：用户显式指定优先；否则用推断结果（build_default_config 的 "fser_date"
    # 默认值不覆盖推断结果，避免 CSV 实际日期列非 fser_date 时取错列）
    c["date_col"] = (cfg.get("date_col") or info["date_col"])

    # ---- 输入校验：所有用到的列必须在 CSV 中存在，否则给出清晰报错而非中途 KeyError ----
    def _missing(name, col):
        return col not in colset

    # 基础列
    if _missing("uid_col", c["uid_col"]):
        raise ValueError(
            f"uid_col 列 {c['uid_col']!r} 不在 CSV 中。CSV 全部列: {all_cols}")
    if _missing("date_col", c["date_col"]):
        raise ValueError(
            f"date_col 列 {c['date_col']!r} 不在 CSV 中。CSV 全部列: {all_cols}")
    # 模型评分列
    if not c["models"]:
        raise ValueError(
            f"models 为空，请至少提供一个 (名称, 评分列)。"
            f"CSV 中疑似评分列: {info['score_cols']}")
    for m in c["models"]:
        if len(m) < 2 or not m[1]:
            raise ValueError(f"models 条目格式应为 [名称, 评分列]，当前: {m!r}")
        if _missing("score_col", m[1]):
            raise ValueError(
                f"模型 {m[0]!r} 的评分列 {m[1]!r} 不在 CSV 中。"
                f"CSV 中疑似评分列: {info['score_cols']}")
    # filter_conds 的列
    for fc in c["filter_conds"]:
        if _missing("filter_cond", fc):
            raise ValueError(
                f"filter_conds 的过滤列 {fc!r} 不在 CSV 中。CSV 全部列: {all_cols}")

    if not c["ks_labels"]:
        c["ks_labels"] = info["label_cols"]
    if not c["lift_labels"]:
        c["lift_labels"] = c["ks_labels"][:2] if len(c["ks_labels"]) >= 2 else c["ks_labels"][:]
    if not c["seg_labels"]:
        c["seg_labels"] = c["lift_labels"][:]
    if c["primary_label"] is None:
        # 默认用第三个标签或第一个 dpd 开头
        dpd = [l for l in c["ks_labels"] if l.startswith("dpd")]
        c["primary_label"] = dpd[0] if dpd else c["ks_labels"][0] if c["ks_labels"] else "dpd30_3c"

    # 校验所有标签列存在
    _label_groups = {"ks_labels": c["ks_labels"], "lift_labels": c["lift_labels"],
                     "seg_labels": c["seg_labels"]}
    if c["primary_label"]:
        _label_groups["primary_label"] = [c["primary_label"]]
    for gname, cols in _label_groups.items():
        miss = [x for x in cols if x not in colset]
        if miss:
            raise ValueError(
                f"{gname} 中列 {miss} 不在 CSV 中。CSV 全部标签列: {info['label_cols']}")

    if not c["ks_labels"]:
        raise ValueError(
            f"未找到任何标签列(fpd/dpd 开头)，且未在 ks_labels 指定。CSV 全部列: {all_cols}")
    if not c["lift_labels"]:
        raise ValueError(
            f"lift_labels 为空，且无法从 ks_labels 推断。CSV 标签列: {info['label_cols']}")

    all_months = info["months"]
    if not all_months:
        # Fallback
        pass
    if c["train_range"] is None or c["oot_range"] is None:
        mid = len(all_months) // 2
        if c["train_range"] is None:
            c["train_range"] = (all_months[0], all_months[mid - 1])
        else:
            c["train_range"] = tuple(c["train_range"])
        if c["oot_range"] is None:
            c["oot_range"] = (all_months[mid], all_months[-1])
        else:
            c["oot_range"] = tuple(c["oot_range"])
    else:
        c["train_range"] = tuple(c["train_range"])
        c["oot_range"] = tuple(c["oot_range"])

    if c["base_month"] is None:
        c["base_month"] = c["oot_range"][0]

    # models list[tuple] -> list of lists for JSON compat
    c["models"] = [list(m) for m in c["models"]]

    return c


# ---------------------------------------------------------------- tag display names
def tag_disp(name):
    """展示名：fpd7_sx30 -> fpd7, dpd30_3c -> dpd30_3c"""
    for prefix in ["fpd7", "fpd15", "fpd30"]:
        if name.startswith(prefix):
            return prefix
    return name


# ---------------------------------------------------------------- styles
TITLE_FONT = Font(bold=True, size=12)
HEADER_FONT = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
CENTER = Alignment(horizontal="center", vertical="center")


def style_header(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def clear_sheet(ws):
    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mc))
    if ws.max_row is not None and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    if ws.max_column is not None and ws.max_column > 0:
        ws.delete_cols(1, ws.max_column)


def write_row(ws, r, values, start_col=1):
    for i, v in enumerate(values):
        ws.cell(row=r, column=start_col + i).value = v


def set_rate(ws, r, c, value):
    """写逾期率/占比格：value 为 float 比率(0~1)时设置百分数格式 0.00%。"""
    cell = ws.cell(row=r, column=c)
    if isinstance(value, str) or value is None:
        cell.value = value
    else:
        cell.value = float(value)
        cell.number_format = "0.00%"


# ---------------------------------------------------------------- data
def load_data(c):
    df = pd.read_csv(c["csv_path"], parse_dates=[c["date_col"]])
    for col, val in c["filter_conds"].items():
        if col in df.columns:
            df = df[df[col] == val].copy()
    df["fmth"] = df[c["date_col"]].dt.strftime("%Y-%m")
    uid = c["uid_col"]
    df = df.groupby(["fmth", uid], as_index=False).first()
    df = df.sort_values("fmth").reset_index(drop=True)
    months = sorted(df.fmth.unique())
    oot_months = [m for m in months if c["oot_range"][0] <= m <= c["oot_range"][1]]
    return df, months, oot_months


# ---------------------------------------------------------------- stats
def cnt_bad(df, label):
    s = df[label]
    return int(s.notna().sum()), int((s > 0).sum())


def ks_grid(df, months, score_col, ks_labels, min_cnt):
    grid = []
    for m in months:
        sub = df[df.fmth == m]
        row = []
        for lab in ks_labels:
            d = sub[[score_col, lab]].dropna()
            n = len(d)
            if n < min_cnt:
                row.append("-"); continue
            bad = int((d[lab] > 0).sum())
            good = n - bad
            if bad == 0 or good == 0:
                row.append("-"); continue
            ks = calc_ks(d[score_col].values, d[lab].values)
            row.append(round(float(ks), 4) if not np.isnan(ks) else "-")
        grid.append(row)
    return grid


def lift_block(d, score_col, label, n_bins, min_cnt):
    d = d[[score_col, label]].dropna().copy()
    d["_bucket"] = pd.qcut(d[score_col], n_bins, duplicates="drop")
    intv_order = sorted(d["_bucket"].cat.categories, key=lambda x: x.left)
    d["_bucket"] = d["_bucket"].cat.reorder_categories(intv_order, ordered=True)

    total_n = len(d)
    total_bad = int((d[label] > 0).sum())
    overall_rate = (total_bad / total_n) if total_n else np.nan

    rows = []
    for it in intv_order:
        g = d[d._bucket == it]
        n = len(g)
        bad = int((g[label] > 0).sum())
        rate = (bad / n) if n >= min_cnt and n > 0 else None
        if rate is None or overall_rate in (0, np.nan) or np.isnan(overall_rate):
            lift_disp, rate_val = "-", "-"
        else:
            lift_disp = round(rate / overall_rate, 4)
            rate_val = rate
        rows.append((str(it), n, bad, rate_val, lift_disp))
    total_rate = overall_rate if overall_rate and not np.isnan(overall_rate) else "-"
    return rows, (None, total_n, total_bad, total_rate, None)


def swap_cross(d, old_col, new_col, label, n_bins):
    d = d[[old_col, new_col, label]].dropna().copy()
    d["_ob"] = pd.qcut(d[old_col], n_bins, duplicates="drop")
    d["_nb"] = pd.qcut(d[new_col], n_bins, duplicates="drop")
    d["_ob"] = d["_ob"].cat.as_ordered()
    d["_nb"] = d["_nb"].cat.as_ordered()

    old_cats = sorted(d["_ob"].cat.categories, key=lambda x: x.left)
    new_cats = sorted(d["_nb"].cat.categories, key=lambda x: x.left)
    n_tab = pd.crosstab(d["_ob"], d["_nb"]).reindex(index=old_cats, columns=new_cats).fillna(0).astype(int)
    dbad = d[d[label] > 0]
    bad_tab = pd.crosstab(dbad["_ob"], dbad["_nb"]).reindex(index=old_cats, columns=new_cats).fillna(0).astype(int)

    old_names = []
    new_names = []
    for i, it in enumerate(old_cats):
        s = d.loc[d["_ob"] == it, old_col]
        old_names.append("%d-[%d, %d]" % (i + 1, int(round(s.min())), int(round(s.max()))))
    for i, it in enumerate(new_cats):
        s = d.loc[d["_nb"] == it, new_col]
        new_names.append("%d-[%d, %d]" % (i + 1, int(round(s.min())), int(round(s.max()))))
    return old_names, new_names, n_tab, bad_tab


def build_score_bins(df, months, score_col):
    """构造打分分布分桶, 覆盖 [min,max]。

    步长自适应: 概率分数(全量取值在 (0,1], 如 LightGBM predict_proba 输出)用 0.1 等距
    (10 桶, PSI/分布才有区分度); 其他范围用 10 分等距(spec §6.2 口径)。

    头尾聚合判据：取各桶在每个月份占比的**跨月最小值（最差月）**，任一月头/尾桶
    占比 < 1% 即并入相邻桶，直至所有月份头尾桶占比均 ≥ 1%（spec §6.2 口径）。
    无样本的月份跳过，不参与 min。
    """
    all_s = df[score_col].dropna()
    if len(all_s) == 0:
        return [0.0, 1.0], ["0.0-1.0"]
    smin, smax = float(all_s.min()), float(all_s.max())
    # 概率分数: 取值全在 (0,1] → 0.1 步长; 否则 10 分等距
    if smax <= 1.0 and smin >= 0.0:
        lo, hi, step = 0.0, 1.0, 0.1
    else:
        lo = int(math.floor(smin / 10) * 10)
        hi = int(math.ceil(smax / 10) * 10)
        step = 10
    # 用 Decimal 避免浮点累加误差(0.1 步长)
    import decimal
    bins = []
    x = decimal.Decimal(str(lo))
    dstep = decimal.Decimal(str(step))
    while x <= decimal.Decimal(str(hi)) + dstep / 2:
        bins.append(float(x))
        x += dstep
    if bins[-1] < hi:
        bins.append(float(hi))
    n_intervals = len(bins) - 1
    interval_labels = [pd.Interval(bins[i], bins[i + 1], closed="right")
                       for i in range(n_intervals)]

    # 跨月最大占比：任何一个月占比 > 1% 即保留该桶不聚合
    max_pcts = np.zeros(n_intervals, dtype=float)
    for m in months:
        s = df[df.fmth == m][score_col].dropna()
        if len(s) == 0:
            continue
        counts, _ = np.histogram(s, bins=bins)
        per = counts / counts.sum() if counts.sum() > 0 else np.zeros(n_intervals)
        max_pcts = np.maximum(max_pcts, per)

    head_idx = 0
    while head_idx < n_intervals and max_pcts[head_idx] < 0.01:
        head_idx += 1
    tail_idx = n_intervals - 1
    while tail_idx >= 0 and max_pcts[tail_idx] < 0.01:
        tail_idx -= 1
    if head_idx > tail_idx:
        head_idx, tail_idx = 0, n_intervals - 1

    out_bins = sorted(set([bins[0]] + bins[head_idx + 1:tail_idx + 1] + [bins[-1]]))
    labels = ["%g-%g" % (out_bins[i], out_bins[i + 1]) for i in range(len(out_bins) - 1)]
    return out_bins, labels


def score_dist_count(df, months, score_col, bins):
    bin_labels = ["%g-%g" % (bins[i], bins[i + 1]) for i in range(len(bins) - 1)]
    records = []
    for m in months:
        s = df[df.fmth == m][score_col].dropna()
        if len(s) == 0:
            records.append([0] * (len(bins) - 1))
            continue
        counts, _ = np.histogram(s, bins=bins)
        records.append([int(v) for v in counts])
    mat = pd.DataFrame(records, index=months, columns=bin_labels).T
    return bin_labels, mat


def psi_row_fn(df, months, score_col, bins, base_month):
    base_s = df[df.fmth == base_month][score_col].dropna()
    base_counts, _ = np.histogram(base_s, bins=bins)
    base_pcts = base_counts / base_counts.sum() if base_counts.sum() > 0 else base_counts
    out = []
    for m in months:
        s = df[df.fmth == m][score_col].dropna()
        if len(s) == 0:
            out.append("-"); continue
        counts, _ = np.histogram(s, bins=bins)
        pcts = counts / counts.sum() if counts.sum() > 0 else counts
        if len(pcts) != len(base_pcts):
            out.append("-"); continue
        psi = calc_psi(pcts, base_pcts)
        out.append(round(float(psi), 4))
    return out


def seg_label_block(df, months, score_col, label, bins, bin_labels, min_cnt):
    cnt_mat, bad_mat, rate_mat = [], [], []
    for blab in bin_labels:
        cnt_r, bad_r, rate_r = [], [], []
        for m in months:
            sub = df[df.fmth == m]
            s = sub[[score_col, label]].dropna()
            # 按 bins 边界做数值 mask(不依赖 Interval 对象匹配, 兼容概率分数 float 边界)
            i = bin_labels.index(blab)
            lo, hi = bins[i], bins[i + 1]
            if i == 0:
                mask = s[score_col] <= hi          # 首桶含下界
            else:
                mask = (s[score_col] > lo) & (s[score_col] <= hi)
            n = int(mask.sum())
            bad = int((s.loc[mask, label] > 0).sum())
            cnt_r.append(n); bad_r.append(bad)
            rate_r.append((bad / n) if n >= min_cnt else "-")
        cnt_mat.append(cnt_r); bad_mat.append(bad_r); rate_mat.append(rate_r)
    return cnt_mat, bad_mat, rate_mat


# ---------------------------------------------------------------- sheets
def sheet1_huisu(ws):
    clear_sheet(ws)
    write_row(ws, 1, ["回溯表（元信息，留空）"])
    ws.cell(1, 1).font = TITLE_FONT
    write_row(ws, 3, ["回溯样本逻辑", "（留空，待补充）"])
    write_row(ws, 5, ["回溯表名", "（留空，待补充）"])
    write_row(ws, 7, ["主要字段", "数据类型", "含义"])
    style_header(ws, 7, 1, 3)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 30


def sheet2_modelinfo(ws, df, months, c):
    clear_sheet(ws)
    uid = c["uid_col"]
    primary = c["primary_label"]
    r = 1
    write_row(ws, r, ["建模样本"]); ws.cell(r, 1).font = TITLE_FONT; r += 1
    write_row(ws, r, [None, "（留空，待补充）"]); r += 2
    write_row(ws, r, ["建模标签"]); ws.cell(r, 1).font = TITLE_FONT; r += 1
    write_row(ws, r, [None, primary]); r += 2
    write_row(ws, r, ["训练集/OOT集"]); ws.cell(r, 1).font = TITLE_FONT; r += 1
    write_row(ws, r, [None, "训练集：%s - %s" % (c["train_range"][0], c["train_range"][1])]); r += 1
    write_row(ws, r, [None, "OOT集：%s - %s" % (c["oot_range"][0], c["oot_range"][1])]); r += 2

    write_row(ws, r, ["样本统计"]); ws.cell(r, 1).font = TITLE_FONT; r += 1
    write_row(ws, r, [None, "月份", "用信", f"有表现({primary})", "逾期", "逾期率"])
    style_header(ws, r, 2, 6); r += 1
    for m in months:
        sub = df[df.fmth == m]
        yongxin = int(sub[uid].nunique())
        cnt, bad = cnt_bad(sub, primary)
        rate = (bad / cnt) if cnt >= c["min_cnt"] else "-"
        write_row(ws, r, [None, m, yongxin, cnt, bad])
        set_rate(ws, r, 6, rate); r += 1
    r += 1

    write_row(ws, r, ["入模特征"]); ws.cell(r, 1).font = TITLE_FONT; r += 1
    write_row(ws, r, [None, "（留空，待补充）"])
    ws.column_dimensions["A"].width = 16
    for cc in "BCDEF":
        ws.column_dimensions[cc].width = 18


def sheet3_ks(ws, df, months, c):
    clear_sheet(ws)
    ks_labels = c["ks_labels"]
    disp = [tag_disp(l) for l in ks_labels]
    n_lab = len(ks_labels)
    r = 1
    write_row(ws, r, ["有表现样本数 & 逾期样本数"]); ws.cell(r, 1).font = TITLE_FONT; r += 1
    write_row(ws, r, [None, "fmth"] + disp + disp); style_header(ws, r, 2, 1 + 1 + 2 * n_lab); r += 1
    write_row(ws, r, [None, "", "cnt"] + [None] * (n_lab - 1) + ["bad_cnt"] + [None] * (n_lab - 1)); r += 1
    for m in months:
        sub = df[df.fmth == m]
        cnts = [cnt_bad(sub, lab)[0] for lab in ks_labels]
        bads = [cnt_bad(sub, lab)[1] for lab in ks_labels]
        write_row(ws, r, [None, m] + cnts + bads); r += 1
    r += 1

    write_row(ws, r, ["逾期率"]); ws.cell(r, 1).font = TITLE_FONT; r += 1
    write_row(ws, r, [None, "fmth"] + disp); style_header(ws, r, 2, 1 + 1 + n_lab); r += 1
    for m in months:
        sub = df[df.fmth == m]
        write_row(ws, r, [None, m])
        for j, lab in enumerate(ks_labels):
            cnt, bad = cnt_bad(sub, lab)
            set_rate(ws, r, 3 + j, (bad / cnt) if cnt >= c["min_cnt"] else "-")
        r += 1
    r += 1

    grids = {}
    for name, col in c["models"]:
        grids[name] = ks_grid(df, months, col, ks_labels, c["min_cnt"])
    for name, col in c["models"]:
        write_row(ws, r, [name]); ws.cell(r, 1).font = TITLE_FONT; r += 1
        write_row(ws, r, [None, "fmth"] + disp); style_header(ws, r, 2, 1 + 1 + n_lab); r += 1
        for i, m in enumerate(months):
            write_row(ws, r, [None, m] + grids[name][i]); r += 1
        r += 1

    if len(c["models"]) >= 2:
        n1, n2 = c["models"][0][0], c["models"][1][0]
        write_row(ws, r, [f"Delta ({n1} - {n2})"]); ws.cell(r, 1).font = TITLE_FONT; r += 1
        write_row(ws, r, [None, "fmth"] + disp); style_header(ws, r, 2, 1 + 1 + n_lab); r += 1
        g1, g2 = grids[n1], grids[n2]
        for i, m in enumerate(months):
            row = []
            for j in range(n_lab):
                a, b = g1[i][j], g2[i][j]
                row.append("-" if a == "-" or b == "-" else round(a - b, 4))
            write_row(ws, r, [None, m] + row); r += 1

    ws.column_dimensions["A"].width = 18
    for i in range(2, 2 + 2 * n_lab + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11


def sheet4_featimp(ws):
    clear_sheet(ws)
    write_row(ws, 1, ["入模特征概况"]); ws.cell(1, 1).font = TITLE_FONT
    write_row(ws, 2, [None, "类型", "特征权重", "特征数量"]); style_header(ws, 2, 2, 4)
    write_row(ws, 4, ["（留空，待补充）"])
    write_row(ws, 6, ["Top N 特征重要性"]); ws.cell(6, 1).font = TITLE_FONT
    write_row(ws, 7, [None, "feature_name", "imp", "%imp", "含义", "类型", "fsm字段"])
    style_header(ws, 7, 2, 8)
    write_row(ws, 9, ["（留空，待补充）"])
    ws.column_dimensions["A"].width = 18
    for cc in "BCDEFGH":
        ws.column_dimensions[cc].width = 16


def sheet5_effect(ws, df, oot_months, c):
    clear_sheet(ws)
    models = c["models"]
    lift_labels = c["lift_labels"]
    n_bins_l = c["lift_bins"]
    min_cnt = c["min_cnt"]
    swap_bins = c["swap_bins"]
    sub_o = df[df.fmth.isin(oot_months)]
    cols_per_label = 5
    r = 1

    # 5.1 Lift
    write_row(ws, r, ["等频分桶 Lift（多个OOT月合并）"]); ws.cell(r, 1).font = TITLE_FONT; r += 1
    for name, col in models:
        total_cols = 1 + cols_per_label * len(lift_labels)
        hdr_lab = [None]
        for lab in lift_labels:
            hdr_lab.append(f"以 {tag_disp(lab)} 为标签")
            hdr_lab += [None] * (cols_per_label - 1)
        write_row(ws, r, hdr_lab[:total_cols]); r += 1
        hdr_col = [name]
        for lab in lift_labels:
            hdr_col += [None, "用信人数", "逾期人数", "逾期率", "Lift"]
        write_row(ws, r, hdr_col[:total_cols])
        style_header(ws, r, 1, total_cols); r += 1

        blocks = {}
        for lab in lift_labels:
            rows, total = lift_block(sub_o, col, lab, n_bins_l, min_cnt)
            blocks[lab] = (rows, total)
        max_n = max(len(blocks[l][0]) for l in lift_labels)
        for i in range(max_n):
            vals = []
            for lab in lift_labels:
                ri = blocks[lab][0][i] if i < len(blocks[lab][0]) else (None, None, None, None, None)
                vals.extend([None, ri[0], ri[1], ri[2], None, ri[4]])
            write_row(ws, r, vals)
            for k, lab in enumerate(lift_labels):
                ri = blocks[lab][0][i] if i < len(blocks[lab][0]) else (None, None, None, None, None)
                set_rate(ws, r, 2 + k * cols_per_label + 3, ri[3])
            r += 1
        vals = []
        for lab in lift_labels:
            tot = blocks[lab][1]
            vals.extend([None, "总计", tot[1], tot[2], None, None])
        write_row(ws, r, vals)
        for k, lab in enumerate(lift_labels):
            set_rate(ws, r, 2 + k * cols_per_label + 3, blocks[lab][1][3])
        r += 1

    # 5.2 SWAP
    if len(models) < 2:
        ws.column_dimensions["A"].width = 16
        for i in range(2, 30):
            ws.column_dimensions[get_column_letter(i)].width = 11
        return

    r += 1
    write_row(ws, r, ["SWAP 分析"]); ws.cell(r, 1).font = TITLE_FONT; r += 1
    for lab in lift_labels:
        write_row(ws, r, [None, f"以 {lab} 为标签"]); r += 1
        old_names, new_names, n_tab, bad_tab = swap_cross(
            sub_o, models[1][1], models[0][1], lab, swap_bins)
        nN = len(new_names)
        width = 1 + 1 + 1 + nN + 1

        c1 = [None, "样本数", models[1][0] + "(列)"] + new_names + ["总计"]
        c2 = [None, "坏样本数", models[1][0] + "(列)"] + new_names + ["总计"]
        c3 = [None, f"逾期率({lab})", models[1][0] + "(列)"] + new_names + ["总计"]
        write_row(ws, r, c1); style_header(ws, r, 1, width)
        write_row(ws, r, c2, start_col=width + 1); style_header(ws, r, width + 1, 2 * width)
        write_row(ws, r, c3, start_col=2 * width + 1); style_header(ws, r, 2 * width + 1, 3 * width)
        r += 1

        for i in range(len(old_names) + 1):
            rn = models[0][0] + "(行)" if i == 0 else (old_names[i] if i < len(old_names) else "总计")

            if i < len(old_names):
                nv = list(n_tab.iloc[i].values) + [int(n_tab.iloc[i].sum())]
                bv = list(bad_tab.iloc[i].values) + [int(bad_tab.iloc[i].sum())]
                rv = [(bad_tab.iloc[i, j] / n_tab.iloc[i, j] if n_tab.iloc[i, j] >= min_cnt else "-") for j in range(nN)]
            else:
                nv = list(n_tab.sum(axis=0).values) + [int(n_tab.values.sum())]
                bv = list(bad_tab.sum(axis=0).values) + [int(bad_tab.values.sum())]
                rv = [(bad_tab.iloc[:, j].sum() / n_tab.iloc[:, j].sum() if n_tab.iloc[:, j].sum() >= min_cnt else "-") for j in range(nN)]
            tot_n = sum(int(x) for x in nv[:-1])
            tot_b = sum(int(x) for x in bv[:-1])
            rv_tot = (tot_b / tot_n) if tot_n >= min_cnt else "-"

            write_row(ws, r, [None, rn] + nv, start_col=1)
            write_row(ws, r, [None, rn] + bv, start_col=width + 1)
            write_row(ws, r, [None, rn], start_col=2 * width + 1)
            for j in range(nN):
                set_rate(ws, r, 2 * width + 3 + j, rv[j])
            set_rate(ws, r, 2 * width + 3 + nN, rv_tot)
            r += 1
        r += 1

    ws.column_dimensions["A"].width = 16
    for i in range(2, 35):
        ws.column_dimensions[get_column_letter(i)].width = 11


def sheet6_dist(ws, df, months, c):
    clear_sheet(ws)
    new_name, new_col = c["models"][0]
    bins, bin_labels = build_score_bins(df, months, new_col)
    base_month = c["base_month"]
    seg_labels = c["seg_labels"]
    min_cnt = c["min_cnt"]
    r = 1

    write_row(ws, r, [f"PSI（基于用信样本，{new_name}）"]); ws.cell(r, 1).font = TITLE_FONT; r += 1
    write_row(ws, r, [None, f"psi(基准月={base_month})"] + months)
    style_header(ws, r, 2, 1 + 1 + len(months)); r += 1
    write_row(ws, r, [None, "psi"] + psi_row_fn(df, months, new_col, bins, base_month)); r += 2

    write_row(ws, r, [f"用信用户分布及占比（{new_name} 有分样本）"]); ws.cell(r, 1).font = TITLE_FONT; r += 1
    _, cnt_mat = score_dist_count(df, months, new_col, bins)
    write_row(ws, r, [None, "score_bin"] + months); style_header(ws, r, 2, 1 + 1 + len(months)); r += 1
    for blab in bin_labels:
        write_row(ws, r, [None, blab] + [int(x) for x in cnt_mat.loc[blab].values]); r += 1
    write_row(ws, r, [None, "Total"] + [int(cnt_mat[m].sum()) for m in months]); r += 1
    r += 1
    write_row(ws, r, [None, "score_bin"] + months); style_header(ws, r, 2, 1 + 1 + len(months)); r += 1
    for blab in bin_labels:
        write_row(ws, r, [None, blab])
        for j, m in enumerate(months):
            t = cnt_mat[m].sum()
            set_rate(ws, r, 3 + j, (cnt_mat.loc[blab, m] / t) if t else 0)
        r += 1
    r += 1

    for lab in seg_labels:
        write_row(ws, r, [f"{tag_disp(lab)} 分段逾期率"]); ws.cell(r, 1).font = TITLE_FONT; r += 1
        cnt_m, bad_m, rate_m = seg_label_block(df, months, new_col, lab, bins, bin_labels, min_cnt)
        for title, mat in [("有表现用户数", cnt_m), ("逾期人数", bad_m), ("逾期率", rate_m)]:
            write_row(ws, r, [None, title]); r += 1
            write_row(ws, r, [None, "score_bin"] + months); style_header(ws, r, 2, 1 + 1 + len(months)); r += 1
            for i, blab in enumerate(bin_labels):
                write_row(ws, r, [None, blab])
                if title == "逾期率":
                    for j in range(len(months)):
                        set_rate(ws, r, 3 + j, mat[i][j])
                else:
                    write_row(ws, r, list(mat[i]), start_col=3)
                r += 1
            if title == "逾期率":
                write_row(ws, r, [None, "Total"])
                for j in range(len(months)):
                    tc = sum(cnt_m[k][j] for k in range(len(bin_labels)))
                    tb = sum(bad_m[k][j] for k in range(len(bin_labels)))
                    set_rate(ws, r, 3 + j, (tb / tc) if tc >= min_cnt else "-")
                r += 1
            else:
                tot = [sum(mat[k][j] for k in range(len(bin_labels))) for j in range(len(months))]
                write_row(ws, r, [None, "Total"] + tot); r += 1
            r += 1

    ws.column_dimensions["A"].width = 16
    for i in range(2, 2 + max(len(months), 10) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11


# ---------------------------------------------------------------- workbook
def create_workbook(template_path=None):
    if template_path and os.path.exists(template_path):
        return openpyxl.load_workbook(template_path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in ["回溯表", "建模信息", "KS", "特征重要性", "模型效果", "打分分布"]:
        wb.create_sheet(name)
    return wb


# ---------------------------------------------------------------- main
def main_with_config(cfg, base_dir=None):
    c = resolve_config(cfg, base_dir=base_dir)
    print(">>> 加载数据 ...")
    df, months, oot_months = load_data(c)
    print(f"    总去重样本: {len(df)}, 月份: {len(months)}, OOT月: {oot_months}")
    print(f"    模型: {[m[0] for m in c['models']]}, 标签: {c['ks_labels']}")

    print(">>> 加载/创建模板 ...")
    wb = create_workbook(c.get("template_path"))

    print(">>> Sheet1 回溯表 ...");        sheet1_huisu(wb["回溯表"])
    print(">>> Sheet2 建模信息 ...");       sheet2_modelinfo(wb["建模信息"], df, months, c)
    print(">>> Sheet3 KS ...");            sheet3_ks(wb["KS"], df, months, c)
    print(">>> Sheet4 特征重要性 ...");     sheet4_featimp(wb["特征重要性"])
    print(">>> Sheet5 模型效果 ...");       sheet5_effect(wb["模型效果"], df, oot_months, c)
    print(">>> Sheet6 打分分布 ...");       sheet6_dist(wb["打分分布"], df, months, c)

    wb.save(c["out_path"])
    print(f">>> 已保存: {c['out_path']}")

    sub = df[df.fmth.isin(oot_months)]
    print("\n===== 自检 =====")
    for lab in c["lift_labels"]:
        for name, col in c["models"]:
            d = sub[[col, lab]].dropna()
            bad_cnt = int((d[lab] > 0).sum())
            if bad_cnt > 0 and len(d) > 0:
                ks = calc_ks(d[col].values, d[lab].values)
                print(f"  OOT合并 {name} KS({lab}, n={len(d)}): {ks:.4f}")
            else:
                print(f"  OOT合并 {name} KS({lab}, n={len(d)}): 无逾期样本")
    for name, col in c["models"]:
        for lab in c["lift_labels"]:
            d = sub[[col, lab]].dropna()
            print(f"  Lift汇总 {name}/{lab}: 总样本={len(d)} 逾期={int((d[lab]>0).sum())} 率={((d[lab]>0).sum()/len(d)):.4f}")
    if len(c["models"]) >= 2:
        for lab in c["lift_labels"]:
            both = sub[sub[c["models"][0][1]].notna() & sub[c["models"][1][1]].notna() & sub[lab].notna()]
            print(f"  SWAP {lab} 有表现样本数(两模型均有分): {len(both)}")
    if oot_months and c["lift_labels"]:
        m0 = oot_months[0]; lab0 = c["lift_labels"][0]
        sm = df[df.fmth == m0]
        cnt, bad = cnt_bad(sm, lab0)
        if cnt > 0:
            print(f"  交叉校验 {m0} {lab0}: cnt={cnt} bad={bad} 率={bad/cnt:.4f}")

    # ---- 结果异常自检（异常时向用户反馈, 不静默出报告） ----
    print("\n===== 结果异常自检 =====")
    issues = []
    try:
        for name, col in c["models"]:
            s_all = df[col].dropna()
            if len(s_all) == 0:
                issues.append(f"[WARN] {name} 打分列 {col} 全为空, 报告各表将无有效内容")
                continue
            bins, blab = build_score_bins(df, months, col)
            if len(bins) <= 2:
                issues.append(
                    f"[WARN] {name} 打分分布仅 {len(bins)-1} 个桶, PSI/分桶无区分度"
                    f"(分数范围 {s_all.min():.4f}~{s_all.max():.4f}), 请检查分数列口径(概率 0~1 用 0.1 步长)或是否打分列选错"
                )
            psi_vals = psi_row_fn(df, months, col, bins, c["base_month"])
            numeric_psi = [v for v in psi_vals if isinstance(v, (int, float))]
            if len(numeric_psi) >= 2 and all(abs(v) < 1e-9 for v in numeric_psi):
                issues.append(
                    f"[WARN] {name} 各月 PSI 均为 0, 分布无漂移信息; 请检查分桶是否退化 / 基准月 {c['base_month']} 是否正确"
                )
        for lab in c["lift_labels"]:
            for name, col in c["models"]:
                d = sub[[col, lab]].dropna()
                if len(d) > 0 and (d[lab] > 0).sum() > 0:
                    ks = calc_ks(d[col].values, d[lab].values)
                    if not (0.0 < ks < 1.0):
                        issues.append(f"[WARN] {name}/{lab} OOT KS={ks:.4f} 异常(应在 0~1)")
        # 回溯表各月逾期率全为 0 提示
        for lab in c["ks_labels"]:
            tot_bad = int((df[lab].fillna(0) > 0).sum())
            if tot_bad == 0:
                issues.append(f"[WARN] 标签 {lab} 全量无正样本, 报告逾期率/KS/Lift 无意义")
    except Exception as e:
        issues.append(f"[WARN] 结果自检执行异常: {e}")
    if issues:
        print("\n".join(issues))
        print(">>> ⚠️ 检测到异常指标, 请先核实原因(分数列口径/基准月/样本)再使用报告")
    else:
        print(">>> 全部指标在合理范围")


def main():
    parser = argparse.ArgumentParser(description="风控模型评估报告生成")
    parser.add_argument("config", type=str, help="config JSON 文件路径")
    parser.add_argument("--dir", type=str, default=None, help="基础目录")
    args = parser.parse_args()
    with open(args.config) as f:
        cfg = json.load(f)
    main_with_config(cfg, base_dir=args.dir)


if __name__ == "__main__":
    main()
