#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
模型对比评估脚本（多版本N-way对比）
用法：
  python compare_models.py --jsons V1.json V2.json V3.json -o 对比报告 --fmt all

输出：
  对比报告.json  +  对比报告.md  +  对比报告.xlsx
"""

import argparse, json, os
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter


# ============================================================
# 样式
# ============================================================
HF = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HFILL = PatternFill(start_color="1A3060", end_color="1A3060", fill_type="solid")
CF = Font(name="微软雅黑", size=10)
BD = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
TITLE_FONT = Font(name="微软雅黑", bold=True, size=13, color="1A3060")
SUB_FONT = Font(name="微软雅黑", bold=True, size=11, color="1A3060")
GREEN_FONT = Font(name="微软雅黑", size=10, color="006100")
RED_FONT = Font(name="微软雅黑", size=10, color="9C0006")


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HF; cell.fill = HFILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BD


def style_header_red(ws, row, ncols):
    """浅灰表头——分桶表专用"""
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    gray_font = Font(name="微软雅黑", bold=True, size=11, color="333333")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = gray_font; cell.fill = gray_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BD


def style_data_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = CF; cell.border = BD


# ============================================================
# Core
# ============================================================
def load_model(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_segments(metrics_dict):
    """Get sorted segments excluding '全量'"""
    segs = [k for k in metrics_dict.keys() if k != '全量']
    return sorted(segs) + ['全量']


def _is_rate_col(col_name):
    """率类指标: label_rate / recall / cum_recall (含 rate 或 recall 子串)。
    注意: count/cnt 不是率, 不在此列。"""
    name_lower = col_name.lower()
    return any(kw in name_lower for kw in ['rate', 'recall'])

def _is_amt_col(col_name):
    name_lower = col_name.lower()
    return any(kw in name_lower for kw in ['amt', 'amount'])

def _fmt_cell(cell, col_name, value):
    cell.value = value
    if value is None:
        return
    if _is_rate_col(col_name):
        cell.number_format = '0.00%'
    elif _is_amt_col(col_name):
        cell.number_format = '#,##0'

def _green_scale(ws, col_idx, min_row, max_row):
    col_letter = get_column_letter(col_idx)
    rng = f'{col_letter}{min_row}:{col_letter}{max_row}'
    ws.conditional_formatting.add(rng,
        DataBarRule(start_type='min', end_type='max', color='2E7D32', showValue=True))

def _blue_scale(ws, col_idx, min_row, max_row):
    col_letter = get_column_letter(col_idx)
    rng = f'{col_letter}{min_row}:{col_letter}{max_row}'
    ws.conditional_formatting.add(rng,
        DataBarRule(start_type='min', end_type='max', color='1A3060', showValue=True))



# ============================================================
# N-way comparison
# ============================================================
def build_nway_comparison(models, names):
    """多模型对比"""
    n = len(models)

    # 收集每个模型的 oot 全量指标
    auc_table = {}; ks_table = {}
    segments = set()
    for i, m in enumerate(models):
        met = m.get('metric_by_segment', {})
        segments.update(met.keys())

    segments = sorted(segments, key=lambda x: (x != '全量', x))

    for seg in segments:
        auc_row = {}; ks_row = {}
        for i, m in enumerate(models):
            met = m.get('metric_by_segment', {}).get(seg, {})
            auc_row[names[i]] = met.get('auc')
            ks_row[names[i]] = met.get('ks')
        auc_table[seg] = auc_row; ks_table[seg] = ks_row

    # 分桶
    buckets_by_model = {}
    for i, m in enumerate(models):
        b = m.get('performance', {}).get('score_buckets', {}).get('全量', [])
        buckets_by_model[names[i]] = b

    return {
        "_format": "模型评估对比报告（多版本）",
        "_generated_at": datetime.now().isoformat(),
        "comparison_meta": {
            "models": names,
            "count": n,
            "conclusion": "见各数据集对比表"
        },
        "auc_comparison": auc_table,
        "ks_comparison": ks_table,
        "buckets_by_model": buckets_by_model,
    }


def build_nway_md(cmp):
    md = f"# 多模型对比报告\n\n"
    md += f"> 生成于 {cmp['_generated_at']}\n\n"
    vs_str = ' vs '.join(cmp['comparison_meta']['models'])
    md += f"**对比模型: {vs_str}**\n\n"

    # AUC表
    md += "## AUC 对比（oot）\n\n"
    models_list = cmp['comparison_meta']['models']
    header = "| 客群 |" + "|".join(f" {m} " for m in models_list) + "|\n"
    sep = "|------|" + "|".join(":---:" for _ in models_list) + "|\n"
    md += header + sep
    for seg, row in cmp['auc_comparison'].items():
        vals = "|".join(f" {row.get(m,'-'):.4f} " if row.get(m) is not None else " - " for m in models_list)
        md += f"| {seg} |{vals}|\n"

    # KS表
    md += "\n## KS 对比（oot）\n\n"
    md += header + sep
    for seg, row in cmp['ks_comparison'].items():
        vals = "|".join(f" {row.get(m,'-'):.4f} " if row.get(m) is not None else " - " for m in models_list)
        md += f"| {seg} |{vals}|\n"

    return md


def build_nway_xlsx(cmp, models, names, filepath):
    wb = openpyxl.Workbook()

    # 提取模型状态（已有/新建）
    model_statuses = []
    for m in models:
        status = m.get('model_meta', {}).get('status', '')
        model_statuses.append('已有' if status == 'retired' else '新建')

    # ---- Sheet 1: 指标对比（行=指标, 列=模型） ----
    ws1 = wb.active
    ws1.title = "1-指标对比"
    title_str = '多模型对比: ' + ' vs '.join(names)
    ws1.cell(row=1, column=1, value=title_str).font = TITLE_FONT
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(names))
    # 说明行
    ws1.cell(row=2, column=1,
        value='数据来源：读取model-eval产出的JSON | 对比逻辑：同数据集、同客群、同指标口径对比').font = Font(name="微软雅黑", size=9, color="888888")
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=1 + len(names))

    # 第3行：已有/新建标识
    ws1.cell(row=3, column=1, value='类型').font = Font(name="微软雅黑", bold=True, size=10, color="666666")
    ws1.cell(row=3, column=1).border = BD
    for ci, (name, st) in enumerate(zip(names, model_statuses), 2):
        cell = ws1.cell(row=3, column=ci, value=st)
        cell.font = Font(name="微软雅黑", size=10, color="1A3060" if st == '已有' else "C00000")
        cell.border = BD

    # 从全量提取指标
    seg = '全量'
    metrics = [
        ('AUC', lambda n: cmp['auc_comparison'].get(seg,{}).get(n), '0.0000'),
        ('KS', lambda n: cmp['ks_comparison'].get(seg,{}).get(n), '0.0000'),
    ]

    # 表头
    ws1.cell(row=4, column=1, value='指标').font = HF; ws1.cell(row=4, column=1).fill = HFILL
    ws1.cell(row=4, column=1).alignment = Alignment(horizontal="center", vertical="center"); ws1.cell(row=4, column=1).border = BD
    for ci, name in enumerate(names, 2):
        cell = ws1.cell(row=4, column=ci, value=name)
        cell.font = HF; cell.fill = HFILL; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = BD

    # Lift rows (vs baseline)
    baseline_name = names[0]
    base_aucs = {}
    base_kss = {}
    for ri, (mname, getter, fmt) in enumerate(metrics, 5):
        ws1.cell(row=ri, column=1, value=mname).font = CF; ws1.cell(row=ri, column=1).border = BD
        for ci, name in enumerate(names, 2):
            v = getter(name)
            cell = ws1.cell(row=ri, column=ci, value=v)
            cell.font = CF; cell.border = BD; cell.number_format = fmt
            if mname == 'AUC': base_aucs[name] = v
            if mname == 'KS': base_kss[name] = v

    for lift_label, base_dict, fmt in [('AUC_Lift', base_aucs, '0.0000'), ('KS_Lift', base_kss, '0.0000')]:
        ri = 5 + len(metrics)
        ws1.cell(row=ri, column=1, value=lift_label).font = Font(name="微软雅黑", size=10, bold=True, color="666666")
        ws1.cell(row=ri, column=1).border = BD
        for ci, name in enumerate(names, 2):
            dv = base_dict.get(name, 0)
            if dv is not None and base_dict.get(baseline_name) is not None:
                base_v = base_dict.get(baseline_name, 1) or 1
            dv = dv / base_v if name != baseline_name else 1.0
            cell = ws1.cell(row=ri, column=ci, value=dv if name != baseline_name else 1.0)
            cell.font = Font(name="微软雅黑", size=10, color="006100" if (dv or 0) > 1 else "9C0006" if (dv or 0) < 1 else "333333")
            cell.border = BD; cell.number_format = '0.00'
        ws1.cell(row=ri, column=1, value=lift_label).font = Font(name="微软雅黑", bold=True, size=10, color="666666")
        ws1.cell(row=ri, column=1).border = BD
        metrics.append((lift_label, None, fmt))  # placeholder

    ws1.freeze_panes = ws1.cell(row=4, column=2)

    # ---- Sheet 2: 所有版本并排对比（去掉分数区间，同指标不同版本画一起） ----
    n_models = len(names)
    baseline_name = names[0]
    ws2 = wb.create_sheet('2-分桶并排对比')
    ws2.sheet_view.showGridLines = False
    # 第 1 行：指标计算逻辑（灰色小字）
    ws2.cell(row=1, column=1,
        value='指标计算逻辑: 分桶=各模型按自己打分降序十分位(Decile10=最高分) | label率=桶内正样本占比 | 召回率=桶正样本/总正样本 | 累计召回=从高到低累计捕获正样本占比').font = Font(name="微软雅黑", size=9, color="888888")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + 3 * n_models)
    # 第 2 行：基线版本声明（灰色小字）
    ws2.cell(row=2, column=1,
        value='基线版本={} | Lift=各模型指标 / 基线同指标 (>1 表示优于基线)'.format(baseline_name)).font = Font(name="微软雅黑", size=9, color="888888")
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2 + 3 * n_models)
    cur_row = 3
    biz_cols = set()
    all_buckets = cmp.get('buckets_by_model', {})

    fixed_cols = {'decile','count','score_min','score_max','label_rate','lift','recall','cum_recall'}
    for name in names:
        for b in all_buckets.get(name, []):
            biz_cols.update(k for k in b.keys() if k not in fixed_cols)
    biz_cols = sorted(biz_cols)
    if n_models > 0 and all_buckets:
        first_buckets = None
        for name in names:
            if all_buckets.get(name):
                first_buckets = all_buckets[name]
                break

        if first_buckets:
            # Header: metric groups with gray merged cells, full model names below
            # 列顺序: label率 | 召回率 | 累计召回
            metric_groups = [
                ('label率(正样本率)', 'label'),
                ('召回率', 'recall'),
                ('累计召回', 'cum_recall'),
            ]
            # Top merged row: gray background, metric group name centered
            gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            gray_font = Font(name='微软雅黑', bold=True, size=10, color='1A3060')
            col = 3
            ws2.cell(row=cur_row, column=1, value='').border = BD
            ws2.cell(row=cur_row, column=2, value='').border = BD
            # label率 group
            grp_end = col + n_models - 1
            cell = ws2.cell(row=cur_row, column=col, value='label率(正样本率)')
            cell.font = gray_font; cell.fill = gray_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws2.merge_cells(start_row=cur_row, start_column=col, end_row=cur_row, end_column=grp_end)
            for c in range(col, grp_end+1):
                ws2.cell(row=cur_row, column=c).fill = gray_fill
                ws2.cell(row=cur_row, column=c).border = BD
            col = grp_end + 1
            # 召回率 group
            grp_end = col + n_models - 1
            cell = ws2.cell(row=cur_row, column=col, value='召回率')
            cell.font = gray_font; cell.fill = gray_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws2.merge_cells(start_row=cur_row, start_column=col, end_row=cur_row, end_column=grp_end)
            for c in range(col, grp_end+1):
                ws2.cell(row=cur_row, column=c).fill = gray_fill
                ws2.cell(row=cur_row, column=c).border = BD
            col = grp_end + 1
            # 累计召回 group
            grp_end = col + n_models - 1
            cell = ws2.cell(row=cur_row, column=col, value='累计召回')
            cell.font = gray_font; cell.fill = gray_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws2.merge_cells(start_row=cur_row, start_column=col, end_row=cur_row, end_column=grp_end)
            for c in range(col, grp_end+1):
                ws2.cell(row=cur_row, column=c).fill = gray_fill
                ws2.cell(row=cur_row, column=c).border = BD
            col = grp_end + 1
            cur_row += 1

            # Header row: full model names under each metric group
            headers = ['分桶', '人数']
            # label率 group
            for n in names:
                headers.append('{}{}'.format(n, ' (基线)' if n == baseline_name else ''))
            # 召回率 group
            for n in names:
                headers.append('{}{}'.format(n, ' (基线)' if n == baseline_name else ''))
            # 累计召回 group
            for n in names:
                headers.append('{}{}'.format(n, ' (基线)' if n == baseline_name else ''))
            for c, h in enumerate(headers, 1):
                ws2.cell(row=cur_row, column=c, value=h)
            style_header_red(ws2, cur_row, len(headers))
            cur_row += 1
            data_start = cur_row

            for decile_idx in range(len(first_buckets)):
                b_list = [all_buckets.get(name, []) for name in names]
                b0 = b_list[0][decile_idx] if decile_idx < len(b_list[0]) else {}
                ws2.cell(row=cur_row, column=1, value=b0.get('decile','')).font = CF; ws2.cell(row=cur_row, column=1).border = BD
                cn = ws2.cell(row=cur_row, column=2, value=b0.get('count',0)); cn.font = CF; cn.border = BD
                col = 3
                # 存储基线值用于delta计算
                base_vals = {}                # label率 group
                for i, name in enumerate(names):
                    b = b_list[i][decile_idx] if decile_idx < len(b_list[i]) else {}
                    c = ws2.cell(row=cur_row, column=col, value=b.get('label_rate'))
                    _fmt_cell(c, 'label_rate', b.get('label_rate')); c.font = CF; c.border = BD
                    base_vals[f'{name}_label率'] = b.get('label_rate') or 0
                    col += 1
                # 召回率 group
                for i, name in enumerate(names):
                    b = b_list[i][decile_idx] if decile_idx < len(b_list[i]) else {}
                    c = ws2.cell(row=cur_row, column=col, value=b.get('recall'))
                    c.font = CF; c.border = BD; c.number_format = '0.00%'
                    base_vals[f'{name}_召回率'] = b.get('recall') or 0
                    col += 1
                # 累计召回 group
                for i, name in enumerate(names):
                    b = b_list[i][decile_idx] if decile_idx < len(b_list[i]) else {}
                    c = ws2.cell(row=cur_row, column=col, value=b.get('cum_recall'))
                    c.font = CF; c.border = BD; c.number_format = '0.00%'
                    base_vals[f'{name}_累计召回'] = b.get('cum_recall') or 0
                    col += 1
                cur_row += 1
            data_end = cur_row - 1

            # DataBars: label率(green) + 召回率(blue) + 累计召回(blue)
            col = 3
            for _ in names: _green_scale(ws2, col, data_start, data_end); col += 1
            for _ in names: _blue_scale(ws2, col, data_start, data_end); col += 1
            for _ in names: _blue_scale(ws2, col, data_start, data_end); col += 1

    ws2.freeze_panes = 'C4'

    # ---- Sheet 2 下方：各版本 vs baseline 的 Lift 子表（label率 / 召回率 / 累计召回 三组并排） ----
    if n_models > 1 and first_buckets and all_buckets.get(baseline_name):
        base_buckets = all_buckets[baseline_name]
        lift_metrics = [
            ('label率 Lift', 'label_rate', _green_scale),
            ('召回率 Lift', 'recall', _blue_scale),
            ('累计召回 Lift', 'cum_recall', _blue_scale),
        ]
        cur_row += 1
        ws2.cell(row=cur_row, column=1,
                 value='各版本 vs {} Lift（label率 / 召回率 / 累计召回 相除）'.format(baseline_name)).font = SUB_FONT
        ws2.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=2 + 3 * n_models)
        cur_row += 1

        # 第 1 行组表头（灰色合并）
        gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        gray_font = Font(name='微软雅黑', bold=True, size=10, color='1A3060')
        col = 3
        ws2.cell(row=cur_row, column=1, value='').border = BD
        ws2.cell(row=cur_row, column=2, value='').border = BD
        for grp_label, _, _ in lift_metrics:
            grp_end = col + n_models - 1
            cell = ws2.cell(row=cur_row, column=col, value=grp_label)
            cell.font = gray_font; cell.fill = gray_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws2.merge_cells(start_row=cur_row, start_column=col, end_row=cur_row, end_column=grp_end)
            for c in range(col, grp_end+1):
                ws2.cell(row=cur_row, column=c).fill = gray_fill
                ws2.cell(row=cur_row, column=c).border = BD
            col = grp_end + 1
        cur_row += 1

        # 第 2 行模型名表头
        delta_h = ['分桶', '人数']
        for _ in lift_metrics:
            for n in names:
                delta_h.append('{}{}'.format(n, ' (基线)' if n == baseline_name else ''))
        for c, h in enumerate(delta_h, 1):
            ws2.cell(row=cur_row, column=c, value=h)
        style_header_red(ws2, cur_row, len(delta_h))
        cur_row += 1
        delta_start = cur_row

        for decile_idx in range(len(first_buckets)):
            b0 = first_buckets[decile_idx] if decile_idx < len(first_buckets) else {}
            ws2.cell(row=cur_row, column=1, value=b0.get('decile','')).font = CF; ws2.cell(row=cur_row, column=1).border = BD
            ws2.cell(row=cur_row, column=2, value=b0.get('count',0)).font = CF; ws2.cell(row=cur_row, column=2).border = BD
            col = 3
            for _, metric_key, _ in lift_metrics:
                base_v = base_buckets[decile_idx].get(metric_key, 0) if decile_idx < len(base_buckets) else 0
                for name in names:
                    if name == baseline_name:
                        ws2.cell(row=cur_row, column=col, value='-').font = CF; ws2.cell(row=cur_row, column=col).border = BD
                    else:
                        b = all_buckets.get(name, [])
                        v = b[decile_idx].get(metric_key, 0) if decile_idx < len(b) else 0
                        dv = v / base_v if base_v > 0 else 0
                        cell = ws2.cell(row=cur_row, column=col, value=dv)
                        cell.font = Font(name='微软雅黑', size=10, color='006100' if dv > 1 else '9C0006' if dv < 1 else '333333')
                        cell.border = BD; cell.number_format = '0.00'
                    col += 1
            cur_row += 1
        delta_end = cur_row - 1
        # Lift 子表不加条件格式 (用户要求 Lift 仅看数值, 不画 DataBar)

    # ---- Sheet 3: raw_data（所有版本按行拼接，供透视） ----
    ws3 = wb.create_sheet("3-raw_data")
    raw_cols = set()
    for name in names:
        for b in all_buckets.get(name, []):
            raw_cols.update(k for k in b.keys() if k not in ('score_min','score_max','decile','count'))
    raw_cols = sorted(raw_cols)
    raw_headers = ['模型版本', '分桶', '人数'] + raw_cols
    for c, h in enumerate(raw_headers, 1):
        ws3.cell(row=1, column=c, value=h)
    style_header_row(ws3, 1, len(raw_headers))
    rr = 2
    for name in names:
        for b in all_buckets.get(name, []):
            ws3.cell(row=rr, column=1, value=name).font = CF; ws3.cell(row=rr, column=1).border = BD
            ws3.cell(row=rr, column=2, value=b.get('decile')).font = CF; ws3.cell(row=rr, column=2).border = BD
            ws3.cell(row=rr, column=3, value=b.get('count',0)).font = CF; ws3.cell(row=rr, column=3).border = BD
            for ci, bc in enumerate(raw_cols, 4):
                v = b.get(bc)
                cell = ws3.cell(row=rr, column=ci, value=v)
                _fmt_cell(cell, bc, v); cell.font = CF; cell.border = BD
            rr += 1
    ws3.freeze_panes = ws3.cell(row=2, column=1)

    wb.save(filepath)




# ============================================================
# Main
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="模型对比评估")
    parser.add_argument("--jsons", nargs="+", required=True, help="模型 eval JSON 列表（可多个）")
    parser.add_argument("-o", "--output", default="对比报告", help="输出文件前缀")
    parser.add_argument("--fmt", choices=["json", "md", "xlsx", "all"], default="all", help="输出格式")
    args = parser.parse_args()

    out_dir = Path(args.output).parent if Path(args.output).parent != Path('.') else Path('.')
    out_dir.mkdir(parents=True, exist_ok=True)
    out_prefix = Path(args.output).stem

    models = []
    model_names = []
    for jpath in args.jsons:
        m = load_model(jpath)
        meta = m.get('model_meta', {})
        name = f"{meta.get('name','?')} {meta.get('version','?')}"
        models.append(m)
        model_names.append(name)
        print(f"加载: {Path(jpath).name} → {name}")

    cmp = build_nway_comparison(models, model_names)

    if args.fmt in ("json", "all"):
        jpath = out_dir / f"{out_prefix}.json"
        jpath.write_text(json.dumps(cmp, indent=2, ensure_ascii=False), encoding='utf-8')
        print(f"[OK] {jpath}")

    if args.fmt in ("md", "all"):
        mpath = out_dir / f"{out_prefix}.md"
        mpath.write_text(build_nway_md(cmp), encoding='utf-8')
        print(f"[OK] {mpath}")

    if args.fmt in ("xlsx", "all"):
        xpath = out_dir / f"{out_prefix}.xlsx"
        build_nway_xlsx(cmp, models, model_names, xpath)
        print(f"[OK] {xpath}")

    print(f"\n{len(models)} 个模型对比完成")


if __name__ == "__main__":
    main()
