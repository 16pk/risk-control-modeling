#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""会话级 N-way 对比聚合: 扫描 session_dir 下所有 run 的 eval JSON,
对 oot / all 两档分别调 compare_models.py 产横向对比三件套。

输出目录: <session_dir>/model-comparison/
  model-comparison_oot.{json,md,xlsx}     ← oot 单独对比
  model-comparison_all.{json,md,xlsx}     ← train+test+oot 合并对比
  _manifest.json

数据源 (两处 glob, 合并后按 model_id 去重保序):
  - <session_dir>/new-models/*/evaluation/*_<split>_eval.json
  - <session_dir>/model-recommend/*/evaluation/*_<split>_eval.json

all split 生成方式:
  对每个命中的模型，优先读取已有的 *_all_eval.json（由 eval_single.py --input-dir 产出）；
  若不存在，则用 merge_eval_splits.py 从 train/test/oot 三档合并生成。

用法:
  python aggregate_session_comparison.py --session-dir <session_dir>
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
import tempfile
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence

import openpyxl
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Font as _OpenpyxlFont, PatternFill, Alignment, Border, Side

_REPO_ROOT = Path(__file__).resolve().parents[2]
_COMPARE = _REPO_ROOT / "classification-model-comparison" / "scripts" / "compare_models.py"
_MERGE = _REPO_ROOT / "classification-model-evaluation" / "scripts" / "merge_eval_splits.py"

_SPLITS = ("oot", "all")


def _glob_eval_jsons(session_dir: Path, split: str) -> List[Path]:
    """扫两个来源, 按 model_id 去重保序。

    来源 1: new-models/*/evaluation/*_<split>_eval.json (本 session 训练的新模型)
    来源 2: model-recommend/*/evaluation/*_<split>_eval.json (推荐基线模型)

    model_id = 文件名去掉 `_<split>_eval.json` 后缀。同 model_id 跨来源
    只保留首个 (优先新模型目录, 避免与 model-recommend 重复计入)。
    """
    sources = [
        session_dir / "new-models",
        session_dir / "model-recommend",
    ]
    hits: List[Path] = []
    seen_ids: set = set()
    for base in sources:
        if not base.exists():
            continue
        for eval_dir in sorted(base.glob("*/evaluation")):
            if not eval_dir.is_dir():
                continue
            # 兼容两种命名: 标准 `*_{split}_eval.json` 与自定义路径 `*_{split}_predictions_eval.json`
            candidates = sorted(
                list(eval_dir.glob(f"*_{split}_eval.json"))
                + list(eval_dir.glob(f"*_{split}_predictions_eval.json"))
            )
            for m in candidates:
                if m.name.endswith(f"_{split}_predictions_eval.json"):
                    model_id = m.name.rsplit(f"_{split}_predictions_eval.json", 1)[0]
                else:
                    model_id = m.name.rsplit(f"_{split}_eval.json", 1)[0]
                if model_id in seen_ids:
                    continue
                seen_ids.add(model_id)
                hits.append(m)
    return hits


def _generate_all_eval_jsons(session_dir: Path) -> List[Path]:
    """为每个命中的模型收集 / 生成 *_all_eval.json。

    优先使用 eval_single.py --input-dir 直接产出的 *_all_eval.json；
    若不存在，则从 train/test/oot 三档 eval JSON 合并生成。
    返回所有 *_all_eval.json 路径列表。
    """
    sources = [
        session_dir / "new-models",
        session_dir / "model-recommend",
    ]
    # model_id -> {"train": path, "test": path, "oot": path}
    model_splits: Dict[str, Dict[str, Path]] = {}

    for base in sources:
        if not base.exists():
            continue
        for eval_dir in sorted(base.glob("*/evaluation")):
            if not eval_dir.is_dir():
                continue
            for m in sorted(eval_dir.glob("*_eval.json")):
                fname = m.name
                # 识别 split：train / test / oot / all
                detected_split = None
                for s in ("train", "test", "oot", "all"):
                    if fname.endswith(f"_{s}_eval.json"):
                        detected_split = s
                        break
                if detected_split is None:
                    continue
                model_id = fname.rsplit(f"_{detected_split}_eval.json", 1)[0]
                if model_id not in model_splits:
                    model_splits[model_id] = {}
                if detected_split not in model_splits[model_id]:
                    model_splits[model_id][detected_split] = m

    all_jsons: List[Path] = []
    for model_id, splits in model_splits.items():
        # 优先使用已有的 all（由 eval_single.py --input-dir 直接产出）
        if "all" in splits:
            all_jsons.append(splits["all"])
            continue

        # 否则从 train+test+oot 合并
        jsons = [splits.get(s) for s in ("train", "test", "oot") if s in splits]
        if len(jsons) < 2:
            # 只有 1 档，跳过（无法做有意义的合并对比）
            continue

        first_path = next(iter(splits.values()))
        out_dir = first_path.parent
        out_path = out_dir / f"{model_id}_all_eval.json"

        # 如果已存在且比所有 split 文件都新，跳过
        if out_path.exists():
            newest_split = max(p.stat().st_mtime for p in jsons)
            if out_path.stat().st_mtime >= newest_split:
                all_jsons.append(out_path)
                continue

        print(f"[aggregate] 合并 {model_id}: {[j.name for j in jsons]} → {out_path.name}")
        cmd = [
            sys.executable, str(_MERGE),
            "--jsons", *[str(j) for j in jsons],
            "-o", str(out_path),
        ]
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0:
            print(result.stdout, file=sys.stdout)
            print(result.stderr, file=sys.stderr)
            print(f"[aggregate] ⚠ 合并 {model_id} 失败，跳过", file=sys.stderr)
            continue
        all_jsons.append(out_path)

    return all_jsons


def _run_compare(
    compare_script: Path,
    jsons: List[Path],
    tmp_dir: Path,
    split: str,
) -> Dict[str, Path]:
    """调一次 compare_models.py 产三件套到临时目录，返回 {ext: path}。"""
    out_prefix = tmp_dir / split
    cmd = [
        sys.executable, str(compare_script),
        "--jsons", *[str(j) for j in jsons],
        "-o", str(out_prefix),
        "--fmt", "all",
    ]
    names = ", ".join(j.parent.parent.name + "/" + j.name for j in jsons)
    print(f"[aggregate] {split}: {len(jsons)} 个模型 → {names}")
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print(result.stdout, file=sys.stdout)
        print(result.stderr, file=sys.stderr)
        raise RuntimeError(
            f"[aggregate] compare_models.py 失败 (code={result.returncode}): split={split}"
        )
    files: Dict[str, Path] = {}
    for ext in ("json", "md", "xlsx"):
        f = tmp_dir / f"{split}.{ext}"
        if not f.exists():
            raise FileNotFoundError(
                f"[aggregate] 预期产物未生成: {f}\n"
                f"compare_models stdout: {result.stdout}"
            )
        files[ext] = f
    return files


def aggregate(
    session_dir: Path,
    produced_by: str = "skills/model-comparison",
) -> Dict[str, List[Path]]:
    """对 oot / all 两档分别产 N-way 对比, 落 model-comparison/。"""
    if not _COMPARE.exists():
        raise FileNotFoundError(
            f"[aggregate] 找不到 compare_models.py: {_COMPARE}"
        )
    if not session_dir.is_dir():
        raise FileNotFoundError(f"[aggregate] session_dir 不存在: {session_dir}")

    out_dir = session_dir / "model-comparison"
    out_dir.mkdir(parents=True, exist_ok=True)

    result: Dict[str, Dict[str, Path]] = {}
    skipped: List[Dict[str, str]] = []
    included_runs: List[str] = []

    # Phase 1: run compare_models.py into a temp dir for each split
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        for split in _SPLITS:
            if split == "all":
                jsons = _generate_all_eval_jsons(session_dir)
            else:
                jsons = _glob_eval_jsons(session_dir, split)
            if len(jsons) < 2:
                print(f"[aggregate] 跳过 {split}: 命中 JSON 不足 2 个 ({len(jsons)}), 无法做 N-way 对比")
                skipped.append({
                    "split": split,
                    "reason": f"insufficient_models: {len(jsons)} < 2",
                })
                continue
            result[split] = _run_compare(_COMPARE, jsons, tmp, split)
            for j in jsons:
                run_name = j.parent.parent.name
                if run_name not in included_runs:
                    included_runs.append(run_name)

        # Phase 2: build combined files from temp intermediates
        combined: Dict[str, dict] = {}
        md_parts: List[str] = []
        SPLIT_LABEL_FONT = _OpenpyxlFont(name="微软雅黑", bold=True, size=12, color="2F5496")
        BD = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
        sheet_data = OrderedDict()
        sheet_names_order: List[str] = []

        for split in _SPLITS:
            if split not in result:
                continue
            f = result[split]
            combined[split] = json.loads(f["json"].read_text(encoding="utf-8"))
            md_parts.append(f["md"].read_text(encoding="utf-8"))

            # Write per-split files (model-comparison_{split}.{json,md,xlsx}) for fill_report.py
            per_split_out_json = out_dir / f"model-comparison_{split}.json"
            per_split_out_md = out_dir / f"model-comparison_{split}.md"
            per_split_out_xlsx = out_dir / f"model-comparison_{split}.xlsx"
            per_split_out_json.write_text(
                json.dumps(combined[split], indent=2, ensure_ascii=False), encoding="utf-8")
            per_split_out_md.write_text(f["md"].read_text(encoding="utf-8"), encoding="utf-8")
            per_split_out_xlsx.write_bytes(f["xlsx"].read_bytes())

            src_wb = openpyxl.load_workbook(str(f["xlsx"]))
            for src_sheet in src_wb.worksheets:
                name = src_sheet.title
                if name not in sheet_data:
                    sheet_data[name] = []
                    sheet_names_order.append(name)
                sheet_rows = []
                for row in src_sheet.iter_rows(min_row=1, max_row=src_sheet.max_row):
                    row_data = []
                    for cell in row:
                        row_data.append(dict(
                            value=cell.value, font=cell.font.copy(), fill=cell.fill.copy(),
                            border=cell.border.copy(), alignment=cell.alignment.copy(),
                            number_format=cell.number_format))
                    sheet_rows.append(row_data)
                sheet_data[name].append((split, sheet_rows, src_sheet.column_dimensions))
            src_wb.close()

        # ---- Write combined JSON ----
        combined_json_path = out_dir / "对比报告.json"
        combined_json_path.write_text(
            json.dumps(combined, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"[aggregate] {combined_json_path.name}")

        # ---- Write combined MD ----
        combined_md_path = out_dir / "对比报告.md"
        combined_md_path.write_text("\n\n".join(md_parts), encoding="utf-8")
        print(f"[aggregate] {combined_md_path.name}")

        # ---- Write combined XLSX (oot/all stacked in same sheets, with unified DataBars) ----
        combined_xlsx_path = out_dir / "对比报告.xlsx"
        if sheet_data:
            from openpyxl.utils import get_column_letter
            cwb = openpyxl.Workbook()
            cwb.remove(cwb.active)
            for sname in sheet_names_order:
                ws = cwb.create_sheet(title=sname)
                ws.sheet_view.showGridLines = False
                cur_row = 1
                for si, (split_label, rows, col_dims) in enumerate(sheet_data[sname]):
                    if si > 0:
                        cur_row += 1
                    ws.cell(row=cur_row, column=1, value=split_label).font = SPLIT_LABEL_FONT
                    cur_row += 2

                    # Find real header (column header row starting with '分桶' or '模型版本')
                    # NOTE: 不能用 'or n >= 3' 兜底 — 组表头行(label率/召回率/累计召回)也会命中 n>=3
                    header_idx = None
                    for ri, rd in enumerate(rows):
                        vals = [str(cd["value"]) if cd["value"] else "" for cd in rd]
                        if vals[0] in ("分桶", "模型版本"):
                            header_idx = ri
                            break
                    data_start = cur_row + header_idx + 1 if header_idx is not None else cur_row

                    # ---- DataBars only on non-raw sheets ----
                    is_raw = "raw_data" in sname.lower()

                    # ---- Classify columns by position (header now has short names only) ----
                    # 新列顺序: 分桶(1) 人数(2) | label率(n) | 召回率(n) | 累计召回(n)
                    green_cols: List[int] = []      # label率
                    blue_recall_cols: List[int] = []  # 召回率
                    blue_cum_cols: List[int] = []    # 累计召回
                    if header_idx is not None and not is_raw:
                        hdr = rows[header_idx]
                        # Count distinct model names (excluding Delta columns)
                        seen_short = set()
                        for ci in range(3, len(hdr) + 1):
                            v = str(hdr[ci-1]["value"]).strip() if hdr[ci-1]["value"] else ""
                            if v:
                                # 去掉 "(基线)" 后缀再统计
                                clean = v.replace(' (基线)', '').strip()
                                if clean and not clean.startswith("Δ"):
                                    seen_short.add(clean)
                        n_models = len(seen_short) or 3
                        base = 3  # after 分桶(1) and 人数(2)
                        green_cols = list(range(base, base + n_models))
                        recall_start = base + n_models
                        blue_recall_cols = list(range(recall_start, recall_start + n_models))
                        cum_start = recall_start + n_models
                        blue_cum_cols = list(range(cum_start, cum_start + n_models))

                    # ---- Split rows into sub-tables (主表 vs Lift 子表) ----
                    # 主表: 应用 green/blue DataBars (label率 / 召回率 / 累计召回)
                    # Lift 子表: 不在此处应用 DataBars (下方重建逻辑会处理 3 组 Lift DataBars)
                    sub_ranges: List[tuple] = []  # [(start_row, end_row, is_lift)]
                    if header_idx is not None and not is_raw:
                        sub_start = header_idx + 1
                        for ri in range(sub_start, len(rows)):
                            v0 = str(rows[ri][0]["value"]) if rows[ri][0] and rows[ri][0]["value"] else ""
                            if "各版本 vs" in v0 or "vs baseline" in v0.lower():
                                if ri > sub_start:
                                    sub_ranges.append((sub_start, ri - 1, False))
                                sub_ranges.append((ri + 1, len(rows) - 1, True))
                                break
                        if not sub_ranges:
                            sub_ranges.append((sub_start, len(rows) - 1, False))

                    # ---- Write rows and collect values (main tables only) ----
                    sub_data: List[dict] = []
                    cur_row_start = cur_row
                    for sr, er, is_lift in sub_ranges:
                        if is_lift:
                            sub_data.append(None)
                            continue
                        gv, rv, cv = [], [], []
                        for ri in range(sr, er + 1):
                            if ri >= len(rows):
                                break
                            row_len = len(rows[ri])
                            for ci in green_cols:
                                if ci - 1 >= row_len:
                                    continue
                                v = rows[ri][ci-1]["value"]
                                if isinstance(v, (int, float)) and v is not None:
                                    gv.append(float(v))
                            for ci in blue_recall_cols:
                                if ci - 1 >= row_len:
                                    continue
                                v = rows[ri][ci-1]["value"]
                                if isinstance(v, (int, float)) and v is not None:
                                    rv.append(float(v))
                            for ci in blue_cum_cols:
                                if ci - 1 >= row_len:
                                    continue
                                v = rows[ri][ci-1]["value"]
                                if isinstance(v, (int, float)) and v is not None:
                                    cv.append(float(v))
                        sub_data.append({
                            "start": cur_row_start + sr,
                            "end": cur_row_start + er,
                            "gm": (max(gv) if gv else 1) or 0.001,
                            "rm": (max(rv) if rv else 1) or 0.001,
                            "cm": (max(cv) if cv else 1) or 0.001,
                        })

                    # Write all rows
                    for rd in rows:
                        for ci, cd in enumerate(rd, 1):
                            dst = ws.cell(row=cur_row, column=ci, value=cd["value"])
                            dst.font = cd["font"]; dst.fill = cd["fill"]; dst.border = cd["border"]
                            dst.alignment = cd["alignment"]
                            if cd["number_format"] != "General":
                                dst.number_format = cd["number_format"]
                        cur_row += 1

                    # ---- Apply DataBars: per sub-table (main only, 3 groups) ----
                    has_bars = green_cols or blue_recall_cols or blue_cum_cols
                    if not is_raw and has_bars:
                        for sd_row in sub_data:
                            if sd_row is None:
                                continue
                            ds, de = sd_row["start"], sd_row["end"]
                            if de < ds:
                                continue
                            for ci in green_cols:
                                ws.conditional_formatting.add(
                                    f"{get_column_letter(ci)}{ds}:{get_column_letter(ci)}{de}",
                                    DataBarRule(start_type="num", start_value=0, end_type="num",
                                                end_value=sd_row["gm"], color="2E7D32", showValue=True))
                            for ci in blue_recall_cols:
                                ws.conditional_formatting.add(
                                    f"{get_column_letter(ci)}{ds}:{get_column_letter(ci)}{de}",
                                    DataBarRule(start_type="num", start_value=0, end_type="num",
                                                end_value=sd_row["rm"], color="1A3060", showValue=True))
                            for ci in blue_cum_cols:
                                ws.conditional_formatting.add(
                                    f"{get_column_letter(ci)}{ds}:{get_column_letter(ci)}{de}",
                                    DataBarRule(start_type="num", start_value=0, end_type="num",
                                                end_value=sd_row["cm"], color="1A3060", showValue=True))

                    # ---- Rebuild merged cells for metric group headers (Sheet 2) ----
                    # 新列顺序: 分桶(1) 人数(2) | label率(nm) | 召回率(nm) | 累计召回(nm)
                    if "分桶并排" in sname:
                        gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
                        gray_font2 = _OpenpyxlFont(name='微软雅黑', bold=True, size=10, color='1A3060')
                        # metric_groups: (label_text, n_cols)
                        # 顺序: label率 | 召回率 | 累计召回
                        # Scan all rows for metric group labels (主表第一行组表头)
                        for r in range(1, ws.max_row + 1):
                            v = ws.cell(row=r, column=3).value
                            if v and str(v).strip() == 'label率(正样本率)':
                                metric_label_row = r
                                data_hdr_row = r + 1
                                # Count distinct models from row below (exclude " (基线)" 后缀)
                                short_names = set()
                                for ci in range(3, ws.max_column + 1):
                                    cv = ws.cell(row=data_hdr_row, column=ci).value
                                    if cv:
                                        s = str(cv).strip()
                                        if s:
                                            clean = s.replace(' (基线)', '').strip()
                                            if clean:
                                                short_names.add(clean)
                                nm = len(short_names)
                                if nm == 0:
                                    continue
                                # 构造分组: [(组名, 列数)]
                                groups = [
                                    ('label率(正样本率)', nm),
                                    ('召回率', nm),
                                    ('累计召回', nm),
                                ]

                                col = 3
                                for gname, gcols in groups:
                                    grp_end = col + gcols - 1
                                    c = ws.cell(row=metric_label_row, column=col, value=gname)
                                    c.font = gray_font2; c.fill = gray_fill
                                    c.alignment = Alignment(horizontal='center', vertical='center')
                                    if gcols > 1:
                                        ws.merge_cells(start_row=metric_label_row, start_column=col,
                                                       end_row=metric_label_row, end_column=grp_end)
                                    for ci2 in range(col, grp_end + 1):
                                        cl = ws.cell(row=metric_label_row, column=ci2)
                                        cl.fill = gray_fill; cl.border = BD
                                    col = grp_end + 1

                        # ---- Rebuild merged cells for Lift subtable (3 groups side by side, NO DataBars) ----
                        # Lift 子表结构: 标题行 | 组表头(label率 Lift / 召回率 Lift / 累计召回 Lift) | 模型名行 | 数据行
                        # 用户要求 Lift 子表不画 DataBar, 仅重建组表头合并单元格
                        for r in range(1, ws.max_row + 1):
                            v = ws.cell(row=r, column=3).value
                            if v and str(v).strip() == 'label率 Lift':
                                grp_hdr_row = r
                                # Count distinct models (exclude " (基线)" suffix)
                                data_hdr_row = r + 1
                                short_names = set()
                                for ci in range(3, ws.max_column + 1):
                                    cv = ws.cell(row=data_hdr_row, column=ci).value
                                    if cv:
                                        s = str(cv).strip()
                                        if s:
                                            clean = s.replace(' (基线)', '').strip()
                                            if clean:
                                                short_names.add(clean)
                                nm = len(short_names)
                                if nm == 0:
                                    continue
                                # Clear any existing DataBar rules on Lift subtable range (defensive)
                                data_start = r + 2
                                data_end = data_start
                                while data_end <= ws.max_row:
                                    v1 = ws.cell(row=data_end, column=1).value
                                    if v1 is None or str(v1).strip() == '':
                                        break
                                    data_end += 1
                                data_end -= 1
                                if data_end >= data_start:
                                    lift_start_col = 3
                                    lift_end_col = 3 + 3 * nm - 1
                                    existing_ranges = list(ws.conditional_formatting._cf_rules.keys())
                                    for rng_key in existing_ranges:
                                        rng_obj = rng_key
                                        if (hasattr(rng_obj, 'min_row') and hasattr(rng_obj, 'max_row')
                                                and hasattr(rng_obj, 'min_col') and hasattr(rng_obj, 'max_col')):
                                            if (rng_obj.min_row >= data_start and rng_obj.max_row <= data_end
                                                    and rng_obj.min_col >= lift_start_col and rng_obj.max_col <= lift_end_col):
                                                del ws.conditional_formatting._cf_rules[rng_key]
                                # Rebuild merged cells for 3 group headers (no DataBars)
                                col = 3
                                for gname in ['label率 Lift', '召回率 Lift', '累计召回 Lift']:
                                    grp_end = col + nm - 1
                                    c = ws.cell(row=grp_hdr_row, column=col, value=gname)
                                    c.font = gray_font2; c.fill = gray_fill
                                    c.alignment = Alignment(horizontal='center', vertical='center')
                                    if nm > 1:
                                        ws.merge_cells(start_row=grp_hdr_row, start_column=col,
                                                       end_row=grp_hdr_row, end_column=grp_end)
                                    for ci2 in range(col, grp_end + 1):
                                        cl = ws.cell(row=grp_hdr_row, column=ci2)
                                        cl.fill = gray_fill; cl.border = BD
                                    col = grp_end + 1

                    for cl, cd in col_dims.items():
                        ex = ws.column_dimensions.get(cl)
                        if ex is None or (ex.width or 0) < (cd.width or 0):
                            ws.column_dimensions[cl].width = cd.width

                    # ---- Set sensible column widths (if not yet set) ----
                    # 分桶(1) 人数(2) 略窄; 模型列宽一点
                    if "分桶并排" in sname and header_idx is not None:
                        # 分桶 / 人数
                        if not ws.column_dimensions['A'].width or ws.column_dimensions['A'].width < 6:
                            ws.column_dimensions['A'].width = 6
                        if not ws.column_dimensions['B'].width or ws.column_dimensions['B'].width < 8:
                            ws.column_dimensions['B'].width = 8
                        # 模型列: C 起, 共 n_models*3 列 (label率 / 召回率 / 累计召回)
                        # 模型列宽 16
                        if 'n_models' in dir() and n_models:
                            col_i = 3
                            # label率 group
                            for _ in range(n_models):
                                if not ws.column_dimensions[get_column_letter(col_i)].width or ws.column_dimensions[get_column_letter(col_i)].width < 14:
                                    ws.column_dimensions[get_column_letter(col_i)].width = 16
                                col_i += 1
                            # 召回率 group
                            for _ in range(n_models):
                                if not ws.column_dimensions[get_column_letter(col_i)].width or ws.column_dimensions[get_column_letter(col_i)].width < 14:
                                    ws.column_dimensions[get_column_letter(col_i)].width = 16
                                col_i += 1
                            # 累计召回 group
                            for _ in range(n_models):
                                if not ws.column_dimensions[get_column_letter(col_i)].width or ws.column_dimensions[get_column_letter(col_i)].width < 14:
                                    ws.column_dimensions[get_column_letter(col_i)].width = 16
                                col_i += 1
            cwb.save(str(combined_xlsx_path))
            cwb.close()
        print(f"[aggregate] {combined_xlsx_path.name}")

    # Phase 3: manifest (with combined files + per-split files)
    final_files = [combined_json_path, combined_md_path, combined_xlsx_path]
    for split in result.keys():
        for ext in ("json", "md", "xlsx"):
            p = out_dir / f"model-comparison_{split}.{ext}"
            if p.exists():
                final_files.append(p)
    manifest = {
        "stage": "session_comparison",
        "schema_version": "1",
        "produced_by": produced_by,
        "created_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "files": [
            {"name": f.name, "size": f.stat().st_size} for f in final_files if f.exists()
        ],
        "compare_engine": "classification-model-comparison/scripts/compare_models.py",
        "session_dir": str(session_dir),
        "included_runs": included_runs,
        "splits": list(result.keys()),
        "skipped": skipped,
    }
    (out_dir / "_manifest.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"[aggregate] _manifest.json")
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="会话级 N-way 对比聚合")
    parser.add_argument(
        "--session-dir", required=True,
        help="session_dir (含 new-models/ 和 model-recommend/ 两类子目录)",
    )
    parser.add_argument(
        "--produced-by", default="skills/model-comparison",
        help="manifest 中的 produced_by 字段",
    )
    args = parser.parse_args()

    res = aggregate(Path(args.session_dir).resolve(), produced_by=args.produced_by)
    print(f"[aggregate] 完成: {len(res)} splits")
    for split, files in res.items():
        print(f"  {split}: {[str(f) for f in files]}")


if __name__ == "__main__":
    main()
